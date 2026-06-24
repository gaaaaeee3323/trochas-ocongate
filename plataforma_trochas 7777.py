# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  PLATAFORMA WEB — GESTIÓN DE TROCHAS CARROZABLES                            ║
║  Sistema de Registro, Inspección Técnica y Gestión de Mantenimiento Vial    ║
║  Municipalidad Distrital de Ocongate — Quispicanchi, Cusco, Perú            ║
║                                                                               ║
║  Autor  : Frank Puma Mamani | Código: 202220055                             ║
║  Curso  : Proyecto Preprofesional — Teoría 7, UTEC 2026-I                    ║
║  Docente: Mg. Fernandez Choquepuma, Miguel Ángel                            ║
╚══════════════════════════════════════════════════════════════════════════════╝

INSTALACIÓN:
    pip install streamlit plotly pandas pillow openpyxl fpdf2

EJECUCIÓN LOCAL:
    streamlit run plataforma_trochas.py

ESTRUCTURA (5 módulos según especificación):
    1. 🏠 INICIO                    — Presentación, KPIs rápidos, gráficos resumen
    2. 🛣️ REGISTRO DE TRAMOS         — Alta de tramos + evidencia fotográfica + adjuntos
    3. 📂 CONSULTA DE TRAMOS         — Buscar / filtrar / editar / eliminar / descargar
    4. 📋 FICHA TÉCNICA DE DAÑOS     — Ficha tipo Provías + clasificación automática + ICT
    5. 📊 REPORTES                  — Reportes automáticos + gráficos Plotly + export Excel/PDF

Persistencia: SQLite local (trochas_ocongate.db) — se crea automáticamente.
"""

import streamlit as st
import pandas as pd
import sqlite3
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
from io import BytesIO
import base64
import os

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN GENERAL
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Gestión de Trochas Carrozables — Ocongate",
    page_icon="🛣️",
    layout="wide",
    initial_sidebar_state="expanded",
)

DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "trochas_ocongate.db")

# ──────────────────────────────────────────────────────────────────────────────
# BASE DE DATOS
# ──────────────────────────────────────────────────────────────────────────────
def get_conn():
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    conn.execute("PRAGMA foreign_keys = ON")
    return conn

# ──────────────────────────────────────────────────────────────────────────────
# SIEMBRA INICIAL — Los 5 tramos reales del proyecto (M1-M5, Ocongate)
# ──────────────────────────────────────────────────────────────────────────────
TRAMOS_PROYECTO = [
    {
        "codigo": "M1", "nombre": "Pacchanta Alta – UPIS", "comunidad": "Pacchanta Alta",
        "longitud_km": 21.12, "ancho_m": 4.5, "estado_actual": "Malo",
        "altitud_msnm": 4250, "estado_drenaje": "Deficiente — cunetas colmatadas",
        "senalizacion": "Ausente", "n_puentes": 2, "n_badenes": 5,
        "accesibilidad": "Estacional (cerrado dic-mar)", "cantera_fuente": "Sí — Cantera Pacchanta (2.3 km)",
        "fuente_lastre": "Río Ausangate (material aluvial)", "ultima_intervencion": "2023-06-15",
        "costo_rutinario": 184000, "costo_periodico": 336000, "gps_lat": -13.6458, "gps_lon": -71.1234,
    },
    {
        "codigo": "M2", "nombre": "Mahuayani – Ocongate", "comunidad": "Mahuayani",
        "longitud_km": 15.00, "ancho_m": 4.5, "estado_actual": "Regular",
        "altitud_msnm": 4180, "estado_drenaje": "Regular — cunetas parcialmente operativas",
        "senalizacion": "Parcial (hitos kilométricos)", "n_puentes": 1, "n_badenes": 3,
        "accesibilidad": "Permanente", "cantera_fuente": "Sí — Cantera Mahuayani (1.8 km)",
        "fuente_lastre": "Quebrada local (material aluvial)", "ultima_intervencion": "2022-09-10",
        "costo_rutinario": 169575, "costo_periodico": 399000, "gps_lat": -13.6010, "gps_lon": -71.3520,
    },
    {
        "codigo": "M3", "nombre": "Palcca Central – Alta", "comunidad": "Palcca Central",
        "longitud_km": 30.00, "ancho_m": 4.0, "estado_actual": "Muy Malo",
        "altitud_msnm": 4400, "estado_drenaje": "Crítico — sin cunetas en la mayor parte del tramo",
        "senalizacion": "Ausente", "n_puentes": 3, "n_badenes": 8,
        "accesibilidad": "Estacional (cerrado dic-abr)", "cantera_fuente": "No identificada",
        "fuente_lastre": "Por definir (requiere estudio de canteras)", "ultima_intervencion": "2022-04-22",
        "costo_rutinario": 405000, "costo_periodico": 630000, "gps_lat": -13.7200, "gps_lon": -71.0850,
    },
    {
        "codigo": "M4", "nombre": "Pacchanta Baja – Cruce UPIS", "comunidad": "Pacchanta Baja",
        "longitud_km": 18.50, "ancho_m": 4.5, "estado_actual": "Malo",
        "altitud_msnm": 4150, "estado_drenaje": "Deficiente — alcantarillas obstruidas",
        "senalizacion": "Parcial (señalización informativa)", "n_puentes": 2, "n_badenes": 4,
        "accesibilidad": "Estacional (cerrado ene-mar)", "cantera_fuente": "Sí — Cantera Pacchanta Baja (3.1 km)",
        "fuente_lastre": "Río Mapacho (material aluvial)", "ultima_intervencion": "2023-02-18",
        "costo_rutinario": 238625, "costo_periodico": 483000, "gps_lat": -13.6520, "gps_lon": -71.1450,
    },
    {
        "codigo": "M5", "nombre": "Ausangate – Comunidad", "comunidad": "Ausangate",
        "longitud_km": 12.00, "ancho_m": 4.0, "estado_actual": "Regular",
        "altitud_msnm": 4600, "estado_drenaje": "Regular — mantenimiento comunal periódico",
        "senalizacion": "Preventiva (gestión comunal)", "n_puentes": 1, "n_badenes": 2,
        "accesibilidad": "Permanente (acceso turístico)", "cantera_fuente": "Sí — Cantera Ausangate (1.2 km)",
        "fuente_lastre": "Glaciar Ausangate (material morrénico)", "ultima_intervencion": "2024-05-30",
        "costo_rutinario": 94929, "costo_periodico": 220000, "gps_lat": -13.7850, "gps_lon": -71.1650,
    },
]


def seed_tramos_proyecto():
    """Inserta los 5 tramos reales del proyecto si la tabla está vacía o si falta alguno."""
    conn = get_conn()
    c = conn.cursor()
    existentes = {row[0] for row in c.execute("SELECT codigo FROM tramos").fetchall()}
    for t in TRAMOS_PROYECTO:
        if t["codigo"] in existentes:
            continue
        c.execute("""
            INSERT INTO tramos (codigo, nombre, comunidad, distrito, provincia, departamento,
                longitud_km, ancho_m, tipo_superficie, estado_actual, responsable,
                telefono_responsable, fecha_registro, foto_inicio, foto_intermedia,
                foto_final, foto_danos, adjunto_pdf, adjunto_pdf_nombre,
                altitud_msnm, estado_drenaje, senalizacion, n_puentes, n_badenes,
                accesibilidad, cantera_fuente, fuente_lastre, ultima_intervencion,
                costo_rutinario, costo_periodico, gps_lat, gps_lon)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            t["codigo"], t["nombre"], t["comunidad"], "Ocongate", "Quispicanchi", "Cusco",
            t["longitud_km"], t["ancho_m"], "Afirmado (AF)", t["estado_actual"],
            "Subgerencia de Gestión de Riesgos y Mantenimiento", "",
            date.today().strftime("%Y-%m-%d"), None, None, None, None, None, None,
            t["altitud_msnm"], t["estado_drenaje"], t["senalizacion"], t["n_puentes"], t["n_badenes"],
            t["accesibilidad"], t["cantera_fuente"], t["fuente_lastre"], t["ultima_intervencion"],
            t["costo_rutinario"], t["costo_periodico"], t["gps_lat"], t["gps_lon"],
        ))
    conn.commit()
    conn.close()


def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS tramos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            codigo TEXT UNIQUE,
            nombre TEXT,
            comunidad TEXT,
            distrito TEXT,
            provincia TEXT,
            departamento TEXT,
            longitud_km REAL,
            ancho_m REAL,
            tipo_superficie TEXT,
            estado_actual TEXT,
            responsable TEXT,
            telefono_responsable TEXT,
            fecha_registro TEXT,
            foto_inicio BLOB,
            foto_intermedia BLOB,
            foto_final BLOB,
            foto_danos BLOB,
            adjunto_pdf BLOB,
            adjunto_pdf_nombre TEXT,
            altitud_msnm REAL,
            estado_drenaje TEXT,
            senalizacion TEXT,
            n_puentes INTEGER,
            n_badenes INTEGER,
            accesibilidad TEXT,
            cantera_fuente TEXT,
            fuente_lastre TEXT,
            ultima_intervencion TEXT,
            costo_rutinario REAL,
            costo_periodico REAL,
            gps_lat REAL,
            gps_lon REAL
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS danos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tramo_id INTEGER,
            progresiva_inicial REAL,
            progresiva_final REAL,
            longitud_afectada REAL,
            tipo_dano TEXT,
            tipo_falla TEXT,
            nivel_gravedad TEXT,
            clase_densidad TEXT,
            estado_tramo TEXT,
            transitabilidad TEXT,
            necesidad_intervencion TEXT,
            tiempo_estimado_dias REAL,
            fecha_inspeccion TEXT,
            observaciones TEXT,
            ict REAL,
            pct_deterioro REAL,
            prioridad TEXT,
            tipo_mantenimiento TEXT,
            foto_dano BLOB,
            FOREIGN KEY (tramo_id) REFERENCES tramos(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS intervenciones (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tramo_id INTEGER,
            dano_id INTEGER,
            tipo_mantenimiento TEXT,
            actividades TEXT,
            expediente TEXT,
            fecha_programada TEXT,
            duracion_dias REAL,
            costo_estimado REAL,
            estado TEXT,
            responsable TEXT,
            observaciones TEXT,
            FOREIGN KEY (tramo_id) REFERENCES tramos(id),
            FOREIGN KEY (dano_id) REFERENCES danos(id)
        )
    """)
    # Migración suave: si la BD ya existía sin la columna tipo_mantenimiento
    try:
        c.execute("ALTER TABLE danos ADD COLUMN tipo_mantenimiento TEXT")
    except sqlite3.OperationalError:
        pass
    for col_def in [
        "altitud_msnm REAL", "estado_drenaje TEXT", "senalizacion TEXT",
        "n_puentes INTEGER", "n_badenes INTEGER", "accesibilidad TEXT",
        "cantera_fuente TEXT", "fuente_lastre TEXT", "ultima_intervencion TEXT",
        "costo_rutinario REAL", "costo_periodico REAL", "gps_lat REAL", "gps_lon REAL",
    ]:
        try:
            c.execute(f"ALTER TABLE tramos ADD COLUMN {col_def}")
        except sqlite3.OperationalError:
            pass
    conn.commit()
    conn.close()


init_db()
seed_tramos_proyecto()

# ──────────────────────────────────────────────────────────────────────────────
# DICCIONARIOS / REGLAS DE CLASIFICACIÓN AUTOMÁTICA
# ──────────────────────────────────────────────────────────────────────────────
TIPOS_SUPERFICIE = ["Asfaltado (AS)", "Afirmado (AF)", "Sin Afirmar (SA)", "Trocha (T)"]
ESTADOS_TRAMO = ["Bueno", "Regular", "Malo", "Muy Malo"]

TIPOS_DANO = [
    "Deformación", "Erosión", "Baches o Huecos", "Encalaminado", "Lodazal",
    "Cruce de Agua", "Hundimiento", "Pérdida de Plataforma", "Deslizamiento",
    "Falla de Drenaje",
]

# Clasificación de falla: Superficial vs Estructural (criterio MTC MCV-2014)
FALLA_SUPERFICIAL = {"Deformación", "Erosión", "Encalaminado", "Lodazal", "Cruce de Agua"}
FALLA_ESTRUCTURAL = {"Baches o Huecos", "Hundimiento", "Pérdida de Plataforma",
                      "Deslizamiento", "Falla de Drenaje"}

# Severidad base por tipo de daño (peso 0-3) usado en el cálculo del ICT
PESO_DANO = {
    "Deformación": 1, "Erosión": 1, "Encalaminado": 1, "Lodazal": 2,
    "Cruce de Agua": 1, "Baches o Huecos": 2, "Hundimiento": 3,
    "Pérdida de Plataforma": 3, "Deslizamiento": 3, "Falla de Drenaje": 2,
}

NIVELES_GRAVEDAD = ["Sin Deterioro", "Leve", "Moderada", "Severa"]
CLASES_DENSIDAD = ["No aplica", "Baja", "Media", "Alta"]

# ──────────────────────────────────────────────────────────────────────────────
# TIPOS DE MANTENIMIENTO VIAL (5 tipos — según informe del proyecto)
# ──────────────────────────────────────────────────────────────────────────────
MANTENIMIENTOS = {
    "Mantenimiento Rutinario": {
        "color": "#2e7d32",
        "descripcion": "Intervención preventiva de baja complejidad, ejecutada de forma periódica y continua sobre toda la vía.",
        "actividades": [
            "Limpieza de cunetas", "Limpieza de alcantarillas", "Roce y desbroce",
            "Eliminación de derrumbes menores", "Perfilado ligero",
        ],
    },
    "Mantenimiento Periódico": {
        "color": "#1565c0",
        "descripcion": "Intervención programada cada 1-3 años para restituir las condiciones estructurales y funcionales de la vía.",
        "actividades": [
            "Reposición de material granular", "Reconformación de plataforma",
            "Perfilado general", "Mejoramiento de drenajes",
        ],
    },
    "Mantenimiento de Emergencia": {
        "color": "#e65100",
        "descripcion": "Intervención inmediata ante eventos imprevistos que interrumpen la transitabilidad de la vía.",
        "actividades": [
            "Atención de derrumbes", "Atención de inundaciones",
            "Restitución de interrupciones de tránsito", "Atención de eventos climáticos extremos",
        ],
    },
    "Rehabilitación": {
        "color": "#8e24aa",
        "descripcion": "Intervención de mayor envergadura para recuperar tramos con deterioro avanzado, sin llegar a reconstrucción total.",
        "actividades": [
            "Recuperación de tramos deteriorados", "Reconstrucción parcial de plataforma",
            "Estabilización de taludes", "Recuperación de drenajes",
        ],
    },
    "Reconstrucción": {
        "color": "#b71c1c",
        "descripcion": "Intervención integral cuando la vía presenta pérdida total de plataforma o daños estructurales severos irreversibles.",
        "actividades": [
            "Reconstrucción integral de la vía", "Reposición por pérdida total de plataforma",
            "Atención de daños estructurales severos",
        ],
    },
}

ESTADOS_INTERVENCION = ["Programada", "En ejecución", "Ejecutada", "Postergada"]

# Expedientes históricos ejecutados — referencia institucional (Municipalidad de Ocongate)
EXPEDIENTES_HISTORICOS = [
    {"Expediente": "ET-MDO-001", "Tramo / Descripción": "Rehabilitación y mejoramiento de trocha carrozable Ocongate – Ccatcca (Km 0+000 al Km 12+500)", "Año": 2021, "Costo total ejecutado (S/.)": 424650.00},
    {"Expediente": "ET-MDO-002", "Tramo / Descripción": "Mantenimiento correctivo trocha carrozable Ccatcca – Marcapata (Km 0+000 al Km 18+200)", "Año": 2022, "Costo total ejecutado (S/.)": 618320.00},
    {"Expediente": "ET-MDO-003", "Tramo / Descripción": "Rehabilitación trocha carrozable Marcapata – Nuñoa (Km 0+000 al Km 22+400)", "Año": 2022, "Costo total ejecutado (S/.)": 761440.00},
    {"Expediente": "ET-MDO-004", "Tramo / Descripción": "Mantenimiento correctivo trocha carrozable Ocongate – Tinki (Km 0+000 al Km 15+600)", "Año": 2023, "Costo total ejecutado (S/.)": 530720.00},
    {"Expediente": "ET-MDO-005", "Tramo / Descripción": "Rehabilitación trocha carrozable Tinki – Ausangate (Km 0+000 al Km 26+800)", "Año": 2024, "Costo total ejecutado (S/.)": 574521.00},
]


def determinar_tipo_mantenimiento(tipo_dano: str, gravedad: str, ict: float, transitabilidad: str) -> str:
    """Determina automáticamente el tipo de mantenimiento (de los 5 definidos) según
    la severidad del daño, el ICT calculado y la transitabilidad resultante."""
    if transitabilidad == "Interrumpida" and gravedad == "Severa" and tipo_dano in {
        "Deslizamiento", "Cruce de Agua", "Lodazal", "Falla de Drenaje", "Hundimiento"
    }:
        return "Mantenimiento de Emergencia"
    if tipo_dano == "Pérdida de Plataforma" and gravedad == "Severa":
        return "Reconstrucción"
    if ict < 30:
        return "Rehabilitación"
    if ict < 60:
        return "Mantenimiento Periódico"
    return "Mantenimiento Rutinario"


def clasificar_falla(tipo_dano: str) -> str:
    if tipo_dano in FALLA_ESTRUCTURAL:
        return "Estructural"
    return "Superficial"


def calcular_gravedad(longitud_afectada: float, longitud_tramo: float, peso_dano: int) -> str:
    """Gravedad en función de % de longitud afectada y severidad propia del daño."""
    if longitud_tramo <= 0:
        pct = 0
    else:
        pct = (longitud_afectada / longitud_tramo) * 100
    score = pct * 0.5 + peso_dano * 10
    if score < 10:
        return "Sin Deterioro"
    elif score < 25:
        return "Leve"
    elif score < 50:
        return "Moderada"
    else:
        return "Severa"


def calcular_estado(gravedad: str) -> str:
    return {
        "Sin Deterioro": "Bueno",
        "Leve": "Regular",
        "Moderada": "Malo",
        "Severa": "Muy Malo",
    }.get(gravedad, "Regular")


def calcular_transitabilidad(estado: str) -> str:
    return {
        "Bueno": "Libre",
        "Regular": "Libre",
        "Malo": "Restringida",
        "Muy Malo": "Interrumpida",
    }.get(estado, "Libre")


def calcular_prioridad(estado: str, transitabilidad: str) -> str:
    if transitabilidad == "Interrumpida":
        return "Urgente"
    if estado == "Malo":
        return "Alta"
    if estado == "Regular":
        return "Media"
    return "Baja"


def calcular_tiempo_reparacion(tipo_dano: str, longitud_afectada: float) -> float:
    """Días estimados de reparación según tipo de daño y longitud afectada (km)."""
    rendimiento_km_dia = {  # rendimiento referencial MCV-2014 (km/día por cuadrilla)
        "Deformación": 0.8, "Erosión": 0.6, "Baches o Huecos": 1.0,
        "Encalaminado": 1.2, "Lodazal": 0.4, "Cruce de Agua": 0.3,
        "Hundimiento": 0.25, "Pérdida de Plataforma": 0.2,
        "Deslizamiento": 0.15, "Falla de Drenaje": 0.35,
    }.get(tipo_dano, 0.5)
    dias = max(1.0, longitud_afectada / rendimiento_km_dia)
    return round(dias, 1)


def calcular_ict(gravedad: str, densidad: str, falla: str) -> float:
    """Índice de Condición de Trocha (0-100, 100 = óptimo)."""
    base = {"Sin Deterioro": 100, "Leve": 80, "Moderada": 55, "Severa": 25}.get(gravedad, 80)
    ajuste_densidad = {"No aplica": 0, "Baja": -2, "Media": -8, "Alta": -15}.get(densidad, 0)
    ajuste_falla = -10 if falla == "Estructural" else -3
    ict = base + ajuste_densidad + ajuste_falla
    return max(0.0, min(100.0, round(ict, 1)))


def calcular_necesidad_intervencion(prioridad: str) -> str:
    return {
        "Urgente": "Intervención inmediata (correctivo de emergencia)",
        "Alta": "Intervención correctiva en el corto plazo (≤ 30 días)",
        "Media": "Mantenimiento periódico programado",
        "Baja": "Mantenimiento rutinario / monitoreo",
    }.get(prioridad, "Monitoreo")


def ict_base_por_estado(estado_actual: str) -> float:
    """Estimación referencial del IEC/ICT actual del tramo a partir de su estado declarado."""
    return {"Bueno": 90, "Regular": 70, "Malo": 45, "Muy Malo": 20}.get(estado_actual, 70)


def anios_desde(fecha_str) -> float:
    if not fecha_str:
        return 0.0
    try:
        fecha = datetime.strptime(str(fecha_str), "%Y-%m-%d").date()
        return max(0.0, (date.today() - fecha).days / 365.25)
    except Exception:
        return 0.0


def curva_iec_deterioro(tramo_dict: dict = None):
    """Construye la figura Plotly de la curva de deterioro del IEC con y sin mantenimiento,
    marcando la posición actual del tramo si se provee."""
    t = list(range(0, 26))
    sin_mant = [max(0, round(100 * (1 - (ti / 23) ** 2.2), 1)) for ti in t]
    rutinario = [round(100 - 0.78 * ti, 1) for ti in t]
    periodico = [round(100 - 0.40 * ti, 1) for ti in t]

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=t, y=sin_mant, mode="lines", name="Sin mantenimiento",
                              line=dict(color="#c62828", dash="dash", width=2.5)))
    fig.add_trace(go.Scatter(x=t, y=rutinario, mode="lines", name="Con mant. rutinario",
                              line=dict(color="#2e7d32", width=2.5)))
    fig.add_trace(go.Scatter(x=t, y=periodico, mode="lines", name="Con mant. periódico",
                              line=dict(color="#1565c0", dash="dot", width=2.5)))

    fig.add_hrect(y0=80, y1=100, fillcolor="#2e7d32", opacity=0.06, line_width=0)
    fig.add_hrect(y0=50, y1=80, fillcolor="#fdd835", opacity=0.08, line_width=0)
    fig.add_hrect(y0=0, y1=50, fillcolor="#c62828", opacity=0.06, line_width=0)
    fig.add_annotation(x=25, y=90, text="Bueno / Muy Bueno", showarrow=False, font=dict(size=10, color="#2e7d32"), xanchor="right")
    fig.add_annotation(x=25, y=65, text="Regular", showarrow=False, font=dict(size=10, color="#a08000"), xanchor="right")
    fig.add_annotation(x=25, y=10, text="Malo / Crítico", showarrow=False, font=dict(size=10, color="#c62828"), xanchor="right")

    if tramo_dict is not None:
        x_actual = round(anios_desde(tramo_dict.get("ultima_intervencion")), 1)
        y_actual = ict_base_por_estado(tramo_dict.get("estado_actual"))
        fig.add_trace(go.Scatter(
            x=[x_actual], y=[y_actual], mode="markers+text", name=f"Tramo {tramo_dict.get('codigo','')} (actual)",
            marker=dict(color="#0b2545", size=14, symbol="star", line=dict(color="white", width=1)),
            text=[f"  {tramo_dict.get('codigo','')}"], textposition="middle right",
        ))

    fig.update_layout(
        xaxis_title="Años desde la última intervención", yaxis_title="IEC (%)",
        yaxis_range=[0, 102], height=380, margin=dict(t=20, b=40, l=10, r=10),
        legend=dict(orientation="h", yanchor="bottom", y=-0.28, xanchor="center", x=0.5),
        plot_bgcolor="white",
    )
    return fig


# ──────────────────────────────────────────────────────────────────────────────
# UTILIDADES DE DATOS
# ──────────────────────────────────────────────────────────────────────────────
def img_to_blob(uploaded_file):
    if uploaded_file is None:
        return None
    return uploaded_file.getvalue()


def blob_to_img(blob):
    if blob is None:
        return None
    return BytesIO(blob)


def df_tramos() -> pd.DataFrame:
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT id, codigo, nombre, comunidad, distrito, provincia, departamento,
               longitud_km, ancho_m, tipo_superficie, estado_actual, responsable,
               telefono_responsable, fecha_registro, altitud_msnm, estado_drenaje,
               senalizacion, n_puentes, n_badenes, accesibilidad, cantera_fuente,
               fuente_lastre, ultima_intervencion, costo_rutinario, costo_periodico,
               gps_lat, gps_lon
        FROM tramos ORDER BY id DESC
    """, conn)
    conn.close()
    return df


def df_danos() -> pd.DataFrame:
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT d.*, t.codigo AS codigo_tramo, t.nombre AS nombre_tramo,
               t.longitud_km AS longitud_tramo_km
        FROM danos d
        LEFT JOIN tramos t ON d.tramo_id = t.id
        ORDER BY d.id DESC
    """, conn)
    conn.close()
    return df


def insertar_tramo(data: dict):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        INSERT INTO tramos (codigo, nombre, comunidad, distrito, provincia, departamento,
            longitud_km, ancho_m, tipo_superficie, estado_actual, responsable,
            telefono_responsable, fecha_registro, foto_inicio, foto_intermedia,
            foto_final, foto_danos, adjunto_pdf, adjunto_pdf_nombre,
            altitud_msnm, estado_drenaje, senalizacion, n_puentes, n_badenes,
            accesibilidad, cantera_fuente, fuente_lastre, ultima_intervencion,
            costo_rutinario, costo_periodico, gps_lat, gps_lon)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        data["codigo"], data["nombre"], data["comunidad"], data["distrito"],
        data["provincia"], data["departamento"], data["longitud_km"], data["ancho_m"],
        data["tipo_superficie"], data["estado_actual"], data["responsable"],
        data["telefono_responsable"], data["fecha_registro"], data["foto_inicio"],
        data["foto_intermedia"], data["foto_final"], data["foto_danos"],
        data["adjunto_pdf"], data["adjunto_pdf_nombre"],
        data.get("altitud_msnm"), data.get("estado_drenaje"), data.get("senalizacion"),
        data.get("n_puentes"), data.get("n_badenes"), data.get("accesibilidad"),
        data.get("cantera_fuente"), data.get("fuente_lastre"), data.get("ultima_intervencion"),
        data.get("costo_rutinario"), data.get("costo_periodico"), data.get("gps_lat"), data.get("gps_lon"),
    ))
    conn.commit()
    conn.close()


def actualizar_tramo(tramo_id: int, data: dict):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        UPDATE tramos SET codigo=?, nombre=?, comunidad=?, distrito=?, provincia=?,
            departamento=?, longitud_km=?, ancho_m=?, tipo_superficie=?, estado_actual=?,
            responsable=?, telefono_responsable=?
        WHERE id=?
    """, (
        data["codigo"], data["nombre"], data["comunidad"], data["distrito"],
        data["provincia"], data["departamento"], data["longitud_km"], data["ancho_m"],
        data["tipo_superficie"], data["estado_actual"], data["responsable"],
        data["telefono_responsable"], tramo_id,
    ))
    conn.commit()
    conn.close()


def eliminar_tramo(tramo_id: int):
    conn = get_conn()
    c = conn.cursor()
    c.execute("DELETE FROM danos WHERE tramo_id=?", (tramo_id,))
    c.execute("DELETE FROM tramos WHERE id=?", (tramo_id,))
    conn.commit()
    conn.close()


def insertar_dano(data: dict) -> int:
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        INSERT INTO danos (tramo_id, progresiva_inicial, progresiva_final, longitud_afectada,
            tipo_dano, tipo_falla, nivel_gravedad, clase_densidad, estado_tramo,
            transitabilidad, necesidad_intervencion, tiempo_estimado_dias, fecha_inspeccion,
            observaciones, ict, pct_deterioro, prioridad, tipo_mantenimiento, foto_dano)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
    """, (
        data["tramo_id"], data["progresiva_inicial"], data["progresiva_final"],
        data["longitud_afectada"], data["tipo_dano"], data["tipo_falla"],
        data["nivel_gravedad"], data["clase_densidad"], data["estado_tramo"],
        data["transitabilidad"], data["necesidad_intervencion"], data["tiempo_estimado_dias"],
        data["fecha_inspeccion"], data["observaciones"], data["ict"], data["pct_deterioro"],
        data["prioridad"], data["tipo_mantenimiento"], data["foto_dano"],
    ))
    conn.commit()
    new_id = c.lastrowid
    conn.close()
    return new_id


def insertar_intervencion(data: dict):
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        INSERT INTO intervenciones (tramo_id, dano_id, tipo_mantenimiento, actividades,
            expediente, fecha_programada, duracion_dias, costo_estimado, estado,
            responsable, observaciones)
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
    """, (
        data["tramo_id"], data.get("dano_id"), data["tipo_mantenimiento"], data["actividades"],
        data["expediente"], data["fecha_programada"], data["duracion_dias"],
        data["costo_estimado"], data["estado"], data["responsable"], data["observaciones"],
    ))
    conn.commit()
    conn.close()


def df_intervenciones() -> pd.DataFrame:
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT i.*, t.codigo AS codigo_tramo, t.nombre AS nombre_tramo
        FROM intervenciones i
        LEFT JOIN tramos t ON i.tramo_id = t.id
        ORDER BY i.fecha_programada ASC
    """, conn)
    conn.close()
    return df


# ──────────────────────────────────────────────────────────────────────────────
# EXPORTACIÓN
# ──────────────────────────────────────────────────────────────────────────────
def exportar_excel(dfs: dict) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    return output.getvalue()


def _pdf_safe(texto) -> str:
    """Convierte cualquier texto a algo seguro para las fuentes core de FPDF
    (latin-1), sustituyendo caracteres no soportados (guion largo, tildes raras, etc.)."""
    if texto is None:
        return ""
    s = str(texto)
    reemplazos = {
        "–": "-", "—": "-", "“": '"', "”": '"', "‘": "'", "’": "'",
        "…": "...", "•": "-", "´": "'", "`": "'",
    }
    for k, v in reemplazos.items():
        s = s.replace(k, v)
    return s.encode("latin-1", "replace").decode("latin-1")


def exportar_pdf_resumen(titulo: str, df_resumen: pd.DataFrame, texto_intro: str = "") -> bytes:
    try:
        from fpdf import FPDF
    except ImportError:
        return None
    pdf = FPDF()
    pdf.add_page()
    epw = pdf.w - 2 * pdf.l_margin  # ancho útil de página
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(epw, 8, _pdf_safe(titulo))
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "", 9)
    if texto_intro:
        pdf.multi_cell(epw, 5, _pdf_safe(texto_intro))
        pdf.set_x(pdf.l_margin)
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 8)
    cols = list(df_resumen.columns)
    col_w = max(20, int(190 / max(1, len(cols))))
    for col in cols:
        pdf.cell(col_w, 6, _pdf_safe(col)[:18], border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7)
    for _, row in df_resumen.iterrows():
        for col in cols:
            pdf.cell(col_w, 6, _pdf_safe(row[col])[:18], border=1)
        pdf.ln()
    return bytes(pdf.output(dest="S"))


def download_button_bytes(data: bytes, filename: str, label: str, mime: str, key: str):
    if data is None:
        st.warning("Para exportar a PDF instala la librería `fpdf2` (pip install fpdf2).")
        return
    st.download_button(label, data=data, file_name=filename, mime=mime, key=key)


# ──────────────────────────────────────────────────────────────────────────────
# FICHAS TÉCNICAS — Formato tipo Provías Descentralizado (1-B Itinerario / 1-D Daños)
# ──────────────────────────────────────────────────────────────────────────────
def generar_grid_progresiva(longitud_km: float, paso: float = 0.5) -> list:
    """Genera los tramos de progresiva (Del Km / Al Km) cada `paso` km."""
    tramos = []
    actual = 0.0
    longitud_km = max(longitud_km, paso)
    while actual < longitud_km:
        siguiente = min(round(actual + paso, 3), longitud_km)
        tramos.append((actual, siguiente))
        actual = siguiente
    return tramos


def df_ficha_itinerario(tramo_dict: dict) -> pd.DataFrame:
    """Construye la tabla 1-B (Ficha del Itinerario del Camino Vecinal) para un tramo,
    incluyendo coordenadas UTM/GPS y obras de arte en el primer registro (punto de referencia)."""
    grid = generar_grid_progresiva(tramo_dict["longitud_km"])
    obras_arte = []
    if tramo_dict.get("n_puentes"):
        obras_arte.append(f"{tramo_dict['n_puentes']} puente(s)")
    if tramo_dict.get("n_badenes"):
        obras_arte.append(f"{tramo_dict['n_badenes']} badén(es)")
    obras_arte_txt = ", ".join(obras_arte) if obras_arte else ""

    filas = []
    for idx, (del_km, al_km) in enumerate(grid):
        filas.append({
            "Del Km": f"{del_km:.3f}", "Al Km": f"{al_km:.3f}",
            "Tipo de Superficie": tramo_dict["tipo_superficie"],
            "Estado de Transitabilidad": tramo_dict["estado_actual"],
            "Ancho de Plataforma (m)": tramo_dict["ancho_m"],
            "Norte (WGS84)": tramo_dict.get("gps_lat") if idx == 0 else "",
            "Este (WGS84)": tramo_dict.get("gps_lon") if idx == 0 else "",
            "Altitud (msnm)": tramo_dict.get("altitud_msnm") if idx == 0 else "",
            "Obras Arte / Drenaje / Señalización": obras_arte_txt if idx == 0 else "",
            "Fotos N°": "1" if idx == 0 else "",
        })
    return pd.DataFrame(filas)


def df_ficha_danos(danos_tramo: pd.DataFrame) -> pd.DataFrame:
    """Construye la tabla 1-D (Ficha Técnica de Daños en Camino Vecinal) con datos reales registrados."""
    if danos_tramo.empty:
        return pd.DataFrame(columns=["Del Km", "Al Km", "Longitud (Km)", "Tipo de Daño",
                                      "Nivel de Gravedad", "Clase de Densidad", "Fecha"])
    filas = []
    for r in danos_tramo.itertuples():
        filas.append({
            "Del Km": f"{r.progresiva_inicial:.3f}", "Al Km": f"{r.progresiva_final:.3f}",
            "Longitud (Km)": f"{r.longitud_afectada:.3f}", "Tipo de Daño": r.tipo_dano,
            "Nivel de Gravedad": r.nivel_gravedad, "Clase de Densidad": r.clase_densidad,
            "Fecha": r.fecha_inspeccion,
        })
    return pd.DataFrame(filas)


def _primera_foto_disponible(tramo_dict: dict):
    """Devuelve (bytes, etiqueta) de la primera fotografía disponible del tramo."""
    for campo, etiqueta in [
        ("foto_inicio", "Foto de inicio del tramo"), ("foto_intermedia", "Foto intermedia"),
        ("foto_final", "Foto final del tramo"), ("foto_danos", "Fotografía de daños"),
    ]:
        if tramo_dict.get(campo):
            return tramo_dict[campo], etiqueta
    return None, None


def exportar_ficha_excel(tramo_dict: dict, danos_tramo: pd.DataFrame) -> bytes:
    """Genera el Excel de la ficha técnica completa: Datos Generales (con foto), 1-B Itinerario, 1-D Daños."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    azul = "0B2545"
    dorado = "C9A227"
    gris = "F4F6F9"
    borde = Border(*[Side(style="thin", color="B0B8C4")] * 4)

    def estilo_header(ws, fila, col_ini, col_fin, texto, color_fondo=azul, color_texto="FFFFFF", alto=22):
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
        celda = ws.cell(row=fila, column=col_ini, value=texto)
        celda.font = Font(name="Calibri", size=12, bold=True, color=color_texto)
        celda.fill = PatternFill("solid", fgColor=color_fondo)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[fila].height = alto

    # ---- HOJA 1: Datos Generales + Foto ----
    ws0 = wb.active
    ws0.title = "Datos Generales"
    for col, w in zip("ABCDE", [22, 30, 4, 28, 30]):
        ws0.column_dimensions[col].width = w
    estilo_header(ws0, 1, 1, 5, "FICHA TÉCNICA DEL TRAMO VIAL", alto=28)
    estilo_header(ws0, 2, 1, 5, "Municipalidad Distrital de Ocongate — Subgerencia de Gestión de Riesgos y Mantenimiento",
                  color_fondo=dorado, color_texto="0B2545", alto=22)

    campos = [
        ("Código de tramo", tramo_dict["codigo"]), ("Nombre del tramo", tramo_dict["nombre"]),
        ("Comunidad / Sector", tramo_dict["comunidad"]),
        ("Distrito / Provincia / Departamento", f"{tramo_dict['distrito']} / {tramo_dict['provincia']} / {tramo_dict['departamento']}"),
        ("Longitud (km)", tramo_dict["longitud_km"]), ("Ancho de plataforma (m)", tramo_dict["ancho_m"]),
        ("Tipo de superficie", tramo_dict["tipo_superficie"]), ("Estado actual", tramo_dict["estado_actual"]),
        ("Responsable de registro", tramo_dict["responsable"]),
        ("Teléfono del responsable", tramo_dict["telefono_responsable"] or "—"),
        ("Fecha de registro", tramo_dict["fecha_registro"]),
    ]
    fila = 4
    for etiqueta, valor in campos:
        c1 = ws0.cell(row=fila, column=1, value=etiqueta)
        c1.font = Font(name="Calibri", size=10, bold=True, color=azul)
        c1.fill = PatternFill("solid", fgColor=gris)
        c1.border = borde
        ws0.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
        c2 = ws0.cell(row=fila, column=2, value=valor)
        c2.font = Font(name="Calibri", size=10)
        c2.border = borde
        ws0.row_dimensions[fila].height = 20
        fila += 1

    foto_bytes, foto_label = _primera_foto_disponible(tramo_dict)
    if foto_bytes:
        try:
            img_buf = BytesIO(foto_bytes)
            xl_img = XLImage(img_buf)
            xl_img.width, xl_img.height = 360, 260
            ws0.add_image(xl_img, "D4")
            ws0.cell(row=fila + 1, column=4, value=f"📷 {foto_label}").font = Font(italic=True, size=9, color="555555")
        except Exception:
            pass

    # ---- HOJA 2: Ficha 1-B Itinerario ----
    ws1 = wb.create_sheet("1-B Itinerario")
    df_b = df_ficha_itinerario(tramo_dict)
    estilo_header(ws1, 1, 1, len(df_b.columns), f"1-B: FICHA DEL ITINERARIO DEL CAMINO VECINAL — Tramo {tramo_dict['codigo']}", alto=26)
    for j, col_name in enumerate(df_b.columns, start=1):
        c = ws1.cell(row=2, column=j, value=col_name)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borde
        ws1.column_dimensions[get_column_letter(j)].width = 16
    for i, row in df_b.iterrows():
        for j, val in enumerate(row, start=1):
            c = ws1.cell(row=i + 3, column=j, value=val)
            c.border = borde
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="center")
            c.fill = PatternFill("solid", fgColor="FFFFFF" if i % 2 == 0 else gris)
    ws1.freeze_panes = "A3"

    # ---- HOJA 3: Ficha 1-D Daños ----
    ws2 = wb.create_sheet("1-D Daños")
    df_d = df_ficha_danos(danos_tramo)
    estilo_header(ws2, 1, 1, max(len(df_d.columns), 1), f"1-D: FICHA TÉCNICA DE DAÑOS EN CAMINO VECINAL — Tramo {tramo_dict['codigo']}", alto=26)
    if not df_d.empty:
        for j, col_name in enumerate(df_d.columns, start=1):
            c = ws2.cell(row=2, column=j, value=col_name)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=azul)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = borde
            ws2.column_dimensions[get_column_letter(j)].width = 18
        for i, row in df_d.iterrows():
            for j, val in enumerate(row, start=1):
                c = ws2.cell(row=i + 3, column=j, value=val)
                c.border = borde
                c.font = Font(size=9)
                c.alignment = Alignment(horizontal="center")
                c.fill = PatternFill("solid", fgColor="FFFFFF" if i % 2 == 0 else gris)
    else:
        ws2.cell(row=3, column=1, value="Sin fichas de daño registradas para este tramo.").font = Font(italic=True, size=10)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def exportar_ficha_pdf(tramo_dict: dict, danos_tramo: pd.DataFrame) -> bytes:
    """Genera el PDF de la ficha técnica completa con foto incrustada."""
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    pdf = FPDF()
    epw = pdf.w - 2 * pdf.l_margin

    # ---- Página 1: Datos generales + foto ----
    pdf.add_page()
    pdf.set_fill_color(11, 37, 69)
    pdf.rect(0, 0, pdf.w, 22, "F")
    pdf.set_xy(10, 6)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, _pdf_safe(f"FICHA TÉCNICA DEL TRAMO {tramo_dict['codigo']}"))
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(10, 26)
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(epw, 5, _pdf_safe("Municipalidad Distrital de Ocongate — Subgerencia de Gestión de Riesgos y Mantenimiento"))
    pdf.set_x(pdf.l_margin)
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(epw, 7, _pdf_safe("DATOS GENERALES"), border="B")
    pdf.ln(9)
    pdf.set_font("Helvetica", "", 9)
    campos = [
        ("Nombre del tramo", tramo_dict["nombre"]), ("Comunidad / Sector", tramo_dict["comunidad"]),
        ("Ubicación", f"{tramo_dict['distrito']} / {tramo_dict['provincia']} / {tramo_dict['departamento']}"),
        ("Longitud", f"{tramo_dict['longitud_km']} km"), ("Ancho de plataforma", f"{tramo_dict['ancho_m']} m"),
        ("Tipo de superficie", tramo_dict["tipo_superficie"]), ("Estado actual", tramo_dict["estado_actual"]),
        ("Responsable", tramo_dict["responsable"]), ("Fecha de registro", tramo_dict["fecha_registro"]),
    ]
    y_inicio_tabla = pdf.get_y()
    for etiqueta, valor in campos:
        pdf.set_x(pdf.l_margin)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(55, 6, _pdf_safe(etiqueta), border=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(epw - 55 - 75, 6, _pdf_safe(valor), border=1)
        pdf.ln()

    foto_bytes, foto_label = _primera_foto_disponible(tramo_dict)
    if foto_bytes:
        try:
            from PIL import Image as PILImage
            pil_img = PILImage.open(BytesIO(foto_bytes)).convert("RGB")
            tmp_path = "/tmp/_ficha_foto_tmp.jpg"
            pil_img.save(tmp_path, format="JPEG", quality=85)
            pdf.image(tmp_path, x=140, y=y_inicio_tabla, w=58)
            pdf.set_xy(140, y_inicio_tabla + 58 * pil_img.height / pil_img.width + 2)
            pdf.set_font("Helvetica", "I", 7)
            pdf.cell(58, 4, _pdf_safe(foto_label))
            os.remove(tmp_path)
        except Exception:
            pass

    # ---- Página 2: Ficha 1-B Itinerario ----
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(epw, 8, _pdf_safe(f"1-B: FICHA DEL ITINERARIO DEL CAMINO VECINAL — Tramo {tramo_dict['codigo']}"))
    pdf.ln(10)
    df_b = df_ficha_itinerario(tramo_dict)
    cols_b = ["Del Km", "Al Km", "Tipo de Superficie", "Estado de Transitabilidad", "Ancho (m)"]
    anchos_b = [22, 22, 45, 50, 25]
    pdf.set_font("Helvetica", "B", 8)
    for c, w in zip(cols_b, anchos_b):
        pdf.cell(w, 7, _pdf_safe(c), border=1, align="C")
    pdf.ln()
    pdf.set_font("Helvetica", "", 8)
    for _, row in df_b.iterrows():
        if pdf.get_y() > 270:
            pdf.add_page()
        pdf.cell(anchos_b[0], 6, _pdf_safe(row["Del Km"]), border=1, align="C")
        pdf.cell(anchos_b[1], 6, _pdf_safe(row["Al Km"]), border=1, align="C")
        pdf.cell(anchos_b[2], 6, _pdf_safe(row["Tipo de Superficie"])[:22], border=1, align="C")
        pdf.cell(anchos_b[3], 6, _pdf_safe(row["Estado de Transitabilidad"])[:24], border=1, align="C")
        pdf.cell(anchos_b[4], 6, _pdf_safe(row["Ancho de Plataforma (m)"]), border=1, align="C")
        pdf.ln()

    # ---- Página 3: Ficha 1-D Daños ----
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(epw, 8, _pdf_safe(f"1-D: FICHA TÉCNICA DE DAÑOS EN CAMINO VECINAL — Tramo {tramo_dict['codigo']}"))
    pdf.ln(10)
    df_d = df_ficha_danos(danos_tramo)
    if df_d.empty:
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(epw, 7, _pdf_safe("Sin fichas de daño registradas para este tramo."))
    else:
        cols_d = list(df_d.columns)
        ancho_d = epw / len(cols_d)
        pdf.set_font("Helvetica", "B", 8)
        for c in cols_d:
            pdf.cell(ancho_d, 7, _pdf_safe(c), border=1, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for _, row in df_d.iterrows():
            if pdf.get_y() > 270:
                pdf.add_page()
            for c in cols_d:
                pdf.cell(ancho_d, 6, _pdf_safe(row[c])[:20], border=1, align="C")
            pdf.ln()

    return bytes(pdf.output(dest="S"))


# ──────────────────────────────────────────────────────────────────────────────
# ESTILOS
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}

    .stApp {
        background-color: #f4f6f9;
    }

    .main-header {
        background: linear-gradient(135deg, #0b2545 0%, #13315c 60%, #0b2545 100%);
        padding: 1.6rem 2rem; border-radius: 6px; color: white; margin-bottom: 1.3rem;
        border-bottom: 4px solid #c9a227;
        box-shadow: 0 2px 8px rgba(0,0,0,0.15);
    }
    .main-header h2 { font-family: 'Georgia', serif; letter-spacing: 0.3px; }
    .main-header .subtitle { opacity:0.92; font-size:0.92rem; margin-top:0.3rem; }
    .main-header .tagline {
        display:inline-block; background:#c9a227; color:#0b2545; font-weight:700;
        font-size:0.72rem; padding:2px 10px; border-radius:3px; letter-spacing:0.5px;
        text-transform:uppercase; margin-top:0.6rem;
    }

    .kpi-card {
        background: #ffffff; border: 1px solid #dfe3ea; border-top: 4px solid #0b2545;
        border-radius: 6px; padding: 1rem 1.1rem; box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    }
    .kpi-value { font-size: 1.9rem; font-weight: 700; color: #0b2545; font-family:'Georgia',serif;}
    .kpi-label { font-size: 0.8rem; color: #5a6472; text-transform:uppercase; letter-spacing:0.3px;}

    .badge-bueno { background:#dcedc8; color:#1b5e20; padding:3px 12px; border-radius:3px; font-weight:700; font-size:0.78rem;}
    .badge-regular { background:#fff3cd; color:#7a6500; padding:3px 12px; border-radius:3px; font-weight:700; font-size:0.78rem;}
    .badge-malo { background:#ffe0b2; color:#a04b00; padding:3px 12px; border-radius:3px; font-weight:700; font-size:0.78rem;}
    .badge-muymalo { background:#f5c6cb; color:#7a0c14; padding:3px 12px; border-radius:3px; font-weight:700; font-size:0.78rem;}

    section[data-testid="stSidebar"] {
        background-color: #0b2545;
    }
    section[data-testid="stSidebar"] * { color: #f0f2f6 !important; }
    section[data-testid="stSidebar"] .stRadio label { font-size: 0.92rem; }
    section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.18); }

    div[data-baseweb="tab-list"] { gap: 4px; }
    button[data-baseweb="tab"] { font-weight:600; }

    h1, h2, h3, h4, h5 { color:#0b2545; font-family:'Georgia',serif; }
</style>
""", unsafe_allow_html=True)


def badge_estado(estado: str) -> str:
    clase = {"Bueno": "badge-bueno", "Regular": "badge-regular",
             "Malo": "badge-malo", "Muy Malo": "badge-muymalo"}.get(estado, "badge-regular")
    return f'<span class="{clase}">{estado}</span>'


# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR — MENÚ PRINCIPAL (5 módulos)
# ──────────────────────────────────────────────────────────────────────────────
st.sidebar.markdown("""
<div style="text-align:center; padding:0.5rem 0 0.8rem 0;">
    <div style="font-size:2.2rem;">🛣️</div>
    <div style="font-weight:700; font-size:1.05rem; letter-spacing:0.3px;">TROCHAS OCONGATE</div>
    <div style="font-size:0.72rem; opacity:0.8; margin-top:0.2rem; line-height:1.3;">
    Municipalidad Distrital de Ocongate<br>Subgerencia de Gestión de Riesgos y Mantenimiento
    </div>
</div>
""", unsafe_allow_html=True)
st.sidebar.markdown("---")

menu = st.sidebar.radio(
    "Menú principal",
    ["🏠 INICIO", "🛣️ REGISTRO DE TRAMOS", "📂 CONSULTA DE TRAMOS",
     "📋 FICHA TÉCNICA DE DAÑOS", "🗓️ PROGRAMACIÓN DE MANTENIMIENTO", "📊 REPORTES"],
    label_visibility="collapsed",
)

st.sidebar.markdown("---")
st.sidebar.caption(f"Sesión: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
st.sidebar.caption("Proyecto Preprofesional — UTEC 2026-I")

# Selector global de tramo (M1-M5) — disponible para filtrar en todos los módulos
_tramos_sidebar = df_tramos()
st.sidebar.markdown("---")
st.sidebar.markdown("**📍 Tramo activo**")
if not _tramos_sidebar.empty:
    _opciones_sb = ["Todos los tramos"] + [
        f"{r.codigo} — {r.nombre}" for r in _tramos_sidebar.sort_values("codigo").itertuples()
    ]
    _sel_sb = st.sidebar.selectbox("Filtrar por tramo", _opciones_sb, label_visibility="collapsed")
    if _sel_sb == "Todos los tramos":
        st.session_state["tramo_activo_codigo"] = None
    else:
        st.session_state["tramo_activo_codigo"] = _sel_sb.split(" — ")[0]
else:
    st.session_state["tramo_activo_codigo"] = None


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 1 — INICIO
# ══════════════════════════════════════════════════════════════════════════════
if menu == "🏠 INICIO":
    st.markdown("""
    <div class="main-header">
        <h2 style="margin:0;">Sistema de Gestión Preventiva de Trochas Carrozables</h2>
        <p class="subtitle" style="margin:0.3rem 0 0 0;">
        Municipalidad Distrital de Ocongate · Provincia de Quispicanchi · Región Cusco
        </p>
        <span class="tagline">Subgerencia de Gestión de Riesgos y Mantenimiento</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    Esta plataforma apoya a la **Subgerencia de Gestión de Riesgos y Mantenimiento** en el
    registro, inspección técnica y seguimiento del estado de la red vial no pavimentada
    (trochas carrozables) del distrito, siguiendo los criterios del **Manual de Conservación
    Vial MCV-2014 (MTC)** y los formatos técnicos de **Provías Descentralizado**.
    """)

    tramos = df_tramos()
    danos = df_danos()

    _tac = st.session_state.get("tramo_activo_codigo")
    if _tac:
        tramos = tramos[tramos["codigo"] == _tac]
        if not danos.empty:
            danos = danos[danos["codigo_tramo"] == _tac]
        st.info(f"📍 Mostrando datos del tramo **{_tac}** únicamente. Cambia el filtro en el panel lateral para ver todos los tramos.")

    total_tramos = len(tramos)
    km_inventariados = round(tramos["longitud_km"].sum(), 2) if not tramos.empty else 0.0

    if not tramos.empty:
        conteo_estado = tramos["estado_actual"].value_counts()
        n_bueno = int(conteo_estado.get("Bueno", 0))
        n_regular = int(conteo_estado.get("Regular", 0))
        n_malo = int(conteo_estado.get("Malo", 0)) + int(conteo_estado.get("Muy Malo", 0))
    else:
        n_bueno = n_regular = n_malo = 0

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{total_tramos}</div>'
                     f'<div class="kpi-label">Tramos registrados</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{km_inventariados}</div>'
                     f'<div class="kpi-label">Km inventariados</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{n_bueno}</div>'
                     f'<div class="kpi-label">Tramos en buen estado</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{n_malo}</div>'
                     f'<div class="kpi-label">Tramos en mal / muy mal estado</div></div>', unsafe_allow_html=True)

    st.markdown("###")
    colA, colB = st.columns(2)

    with colA:
        st.markdown("##### Distribución de tramos por estado")
        if not tramos.empty:
            fig = px.pie(
                tramos, names="estado_actual", hole=0.45,
                color="estado_actual",
                color_discrete_map={"Bueno": "#43a047", "Regular": "#fdd835",
                                     "Malo": "#fb8c00", "Muy Malo": "#e53935"},
            )
            fig.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=320)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Aún no hay tramos registrados. Ve a **🛣️ Registro de Tramos**.")

    with colB:
        st.markdown("##### Kilómetros por comunidad / sector")
        if not tramos.empty:
            agg = tramos.groupby("comunidad")["longitud_km"].sum().reset_index()
            fig2 = px.bar(agg, x="comunidad", y="longitud_km", color="longitud_km",
                          color_continuous_scale="Greens", text_auto=".1f")
            fig2.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=320,
                               yaxis_title="km", xaxis_title="")
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Sin datos aún.")

    st.markdown("###")
    st.markdown("##### Indicadores rápidos de inspección")
    if not danos.empty:
        c1, c2, c3 = st.columns(3)
        c1.metric("Daños registrados", len(danos))
        c2.metric("ICT promedio de la red", f"{danos['ict'].mean():.1f} / 100")
        c3.metric("Tramos con prioridad Urgente/Alta",
                  int(danos["prioridad"].isin(["Urgente", "Alta"]).sum()))
    else:
        st.info("Aún no se han registrado fichas de daños. Ve a **📋 Ficha Técnica de Daños**.")


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 2 — REGISTRO DE TRAMOS
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "🛣️ REGISTRO DE TRAMOS":
    st.markdown("## 🛣️ Registro de Tramos")
    st.caption("Registra los datos generales, evidencia fotográfica y archivos adjuntos de un tramo vial.")

    with st.form("form_registro_tramo", clear_on_submit=True):
        st.markdown("#### Datos Generales")
        c1, c2, c3 = st.columns(3)
        codigo = c1.text_input("Código de tramo *", placeholder="Ej: M-01")
        nombre = c2.text_input("Nombre del tramo *", placeholder="Ej: Ocongate - Pacchanta")
        comunidad = c3.text_input("Comunidad / Sector *", placeholder="Ej: Pacchanta Baja")

        c4, c5, c6 = st.columns(3)
        distrito = c4.text_input("Distrito *", value="Ocongate")
        provincia = c5.text_input("Provincia *", value="Quispicanchi")
        departamento = c6.text_input("Departamento *", value="Cusco")

        c7, c8, c9 = st.columns(3)
        longitud_km = c7.number_input("Longitud (km) *", min_value=0.0, step=0.1, format="%.2f")
        ancho_m = c8.number_input("Ancho de plataforma (m) *", min_value=0.0, step=0.1, format="%.2f")
        tipo_superficie = c9.selectbox("Tipo de superficie *", TIPOS_SUPERFICIE)

        c10, c11, c12 = st.columns(3)
        estado_actual = c10.selectbox("Estado actual *", ESTADOS_TRAMO)
        responsable = c11.text_input("Responsable de registro *")
        telefono_responsable = c12.text_input("Número telefónico del responsable", placeholder="9XXXXXXXX")

        fecha_registro = st.date_input("Fecha de registro *", value=date.today())

        with st.expander("🔧 Datos técnicos avanzados y recursos (opcional)"):
            a1, a2, a3 = st.columns(3)
            altitud_msnm = a1.number_input("Altitud (m s.n.m.)", min_value=0.0, step=10.0, value=0.0)
            estado_drenaje = a2.text_input("Estado del drenaje", placeholder="Ej: Deficiente — cunetas colmatadas")
            senalizacion = a3.text_input("Señalización", placeholder="Ej: Ausente / Parcial / Completa")

            a4, a5, a6 = st.columns(3)
            n_puentes = a4.number_input("N° de puentes", min_value=0, step=1, value=0)
            n_badenes = a5.number_input("N° de badenes", min_value=0, step=1, value=0)
            accesibilidad = a6.text_input("Accesibilidad", placeholder="Ej: Estacional (cerrado dic-mar)")

            a7, a8 = st.columns(2)
            cantera_fuente = a7.text_input("Cantera / fuente de material", placeholder="Ej: Sí — Cantera Pacchanta (2.3 km)")
            fuente_lastre = a8.text_input("Fuente de lastre", placeholder="Ej: Río Ausangate (material aluvial)")

            a9, a10, a11 = st.columns(3)
            ultima_intervencion = a9.date_input("Fecha de última intervención", value=None)
            costo_rutinario = a10.number_input("Costo est. mant. rutinario (S/.)", min_value=0.0, step=1000.0, value=0.0)
            costo_periodico = a11.number_input("Costo est. mant. periódico (S/.)", min_value=0.0, step=1000.0, value=0.0)

            a12, a13 = st.columns(2)
            gps_lat = a12.number_input("Coordenada GPS — Latitud", value=0.0, format="%.4f")
            gps_lon = a13.number_input("Coordenada GPS — Longitud", value=0.0, format="%.4f")

        st.markdown("#### Evidencia Fotográfica")
        f1, f2, f3, f4 = st.columns(4)
        foto_inicio = f1.file_uploader("Foto inicio del tramo", type=["jpg", "jpeg", "png"], key="fi")
        foto_intermedia = f2.file_uploader("Foto intermedia", type=["jpg", "jpeg", "png"], key="fm")
        foto_final = f3.file_uploader("Foto final del tramo", type=["jpg", "jpeg", "png"], key="ff")
        foto_danos = f4.file_uploader("Fotografías de daños", type=["jpg", "jpeg", "png"], key="fd")

        st.markdown("#### Archivos Adjuntos")
        adjunto_pdf = st.file_uploader("Adjuntar PDF (expediente, ficha, plano, etc.)", type=["pdf"])

        enviado = st.form_submit_button("💾 Guardar tramo en la base de datos", use_container_width=True)

    if enviado:
        if not codigo or not nombre or not comunidad or not responsable:
            st.error("Completa los campos obligatorios marcados con *.")
        else:
            try:
                data = {
                    "codigo": codigo, "nombre": nombre, "comunidad": comunidad,
                    "distrito": distrito, "provincia": provincia, "departamento": departamento,
                    "longitud_km": longitud_km, "ancho_m": ancho_m,
                    "tipo_superficie": tipo_superficie, "estado_actual": estado_actual,
                    "responsable": responsable, "telefono_responsable": telefono_responsable,
                    "fecha_registro": fecha_registro.strftime("%Y-%m-%d"),
                    "foto_inicio": img_to_blob(foto_inicio),
                    "foto_intermedia": img_to_blob(foto_intermedia),
                    "foto_final": img_to_blob(foto_final),
                    "foto_danos": img_to_blob(foto_danos),
                    "adjunto_pdf": adjunto_pdf.getvalue() if adjunto_pdf else None,
                    "adjunto_pdf_nombre": adjunto_pdf.name if adjunto_pdf else None,
                    "altitud_msnm": altitud_msnm if altitud_msnm else None,
                    "estado_drenaje": estado_drenaje or None,
                    "senalizacion": senalizacion or None,
                    "n_puentes": n_puentes if n_puentes else None,
                    "n_badenes": n_badenes if n_badenes else None,
                    "accesibilidad": accesibilidad or None,
                    "cantera_fuente": cantera_fuente or None,
                    "fuente_lastre": fuente_lastre or None,
                    "ultima_intervencion": ultima_intervencion.strftime("%Y-%m-%d") if ultima_intervencion else None,
                    "costo_rutinario": costo_rutinario if costo_rutinario else None,
                    "costo_periodico": costo_periodico if costo_periodico else None,
                    "gps_lat": gps_lat if gps_lat else None,
                    "gps_lon": gps_lon if gps_lon else None,
                }
                insertar_tramo(data)
                st.success(f"✅ Tramo **{codigo} — {nombre}** guardado correctamente en la base de datos.")
            except sqlite3.IntegrityError:
                st.error(f"⚠️ Ya existe un tramo con el código **{codigo}**. Usa un código distinto.")

    st.markdown("---")
    st.caption(f"Tramos registrados actualmente: **{len(df_tramos())}**")


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 3 — CONSULTA DE TRAMOS
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📂 CONSULTA DE TRAMOS":
    st.markdown("## 📂 Consulta de Tramos")
    tramos = df_tramos()

    if tramos.empty:
        st.info("No hay tramos registrados todavía.")
    else:
        st.markdown("#### Filtros")
        f1, f2, f3, f4 = st.columns(4)
        busq_codigo = f1.text_input("Buscar por código")
        busq_nombre = f2.text_input("Buscar por nombre")
        filtro_comunidad = f3.selectbox("Filtrar por comunidad", ["Todas"] + sorted(tramos["comunidad"].dropna().unique().tolist()))
        filtro_estado = f4.selectbox("Filtrar por estado", ["Todos"] + ESTADOS_TRAMO)

        min_km, max_km = float(tramos["longitud_km"].min()), float(tramos["longitud_km"].max())
        if min_km == max_km:
            max_km += 1
        rango_long = st.slider("Filtrar por longitud (km)", min_value=float(min_km), max_value=float(max_km),
                                value=(float(min_km), float(max_km)))

        df_f = tramos.copy()
        if busq_codigo:
            df_f = df_f[df_f["codigo"].str.contains(busq_codigo, case=False, na=False)]
        if busq_nombre:
            df_f = df_f[df_f["nombre"].str.contains(busq_nombre, case=False, na=False)]
        if filtro_comunidad != "Todas":
            df_f = df_f[df_f["comunidad"] == filtro_comunidad]
        if filtro_estado != "Todos":
            df_f = df_f[df_f["estado_actual"] == filtro_estado]
        df_f = df_f[(df_f["longitud_km"] >= rango_long[0]) & (df_f["longitud_km"] <= rango_long[1])]

        st.markdown(f"#### Resultados ({len(df_f)} tramo(s))")
        st.dataframe(
            df_f[["codigo", "nombre", "comunidad", "distrito", "longitud_km", "ancho_m",
                  "tipo_superficie", "estado_actual", "responsable", "fecha_registro"]],
            use_container_width=True, hide_index=True,
        )

        st.markdown("---")
        st.markdown("#### Detalle / Edición / Eliminación de un tramo")
        opciones = {f"{row.codigo} — {row.nombre}": row.id for row in df_f.itertuples()}
        if opciones:
            seleccion = st.selectbox("Selecciona un tramo", list(opciones.keys()))
            tramo_id = opciones[seleccion]

            conn = get_conn()
            row = conn.execute("SELECT * FROM tramos WHERE id=?", (tramo_id,)).fetchone()
            cols = [d[0] for d in conn.execute("SELECT * FROM tramos LIMIT 1").description]
            conn.close()
            tramo_dict = dict(zip(cols, row))

            tabs = st.tabs(["📄 Ficha", "📷 Fotografías", "✏️ Editar", "🗑️ Eliminar", "⬇️ Descargar"])

            with tabs[0]:
                c1, c2 = st.columns(2)
                with c1:
                    st.write(f"**Código:** {tramo_dict['codigo']}")
                    st.write(f"**Nombre:** {tramo_dict['nombre']}")
                    st.write(f"**Comunidad:** {tramo_dict['comunidad']}")
                    st.write(f"**Distrito / Provincia / Dpto:** {tramo_dict['distrito']} / {tramo_dict['provincia']} / {tramo_dict['departamento']}")
                    st.write(f"**Longitud:** {tramo_dict['longitud_km']} km")
                with c2:
                    st.write(f"**Ancho de plataforma:** {tramo_dict['ancho_m']} m")
                    st.write(f"**Tipo de superficie:** {tramo_dict['tipo_superficie']}")
                    st.markdown(f"**Estado actual:** {badge_estado(tramo_dict['estado_actual'])}", unsafe_allow_html=True)
                    st.write(f"**Responsable:** {tramo_dict['responsable']} ({tramo_dict['telefono_responsable'] or 's/n'})")
                    st.write(f"**Fecha de registro:** {tramo_dict['fecha_registro']}")
                if tramo_dict["adjunto_pdf"]:
                    st.download_button("⬇️ Descargar PDF adjunto", data=tramo_dict["adjunto_pdf"],
                                        file_name=tramo_dict["adjunto_pdf_nombre"] or "adjunto.pdf",
                                        mime="application/pdf", key=f"pdf_{tramo_id}")

            with tabs[1]:
                fc1, fc2, fc3, fc4 = st.columns(4)
                for col, label, container in [
                    ("foto_inicio", "Inicio", fc1), ("foto_intermedia", "Intermedia", fc2),
                    ("foto_final", "Final", fc3), ("foto_danos", "Daños", fc4),
                ]:
                    with container:
                        st.caption(label)
                        if tramo_dict[col]:
                            st.image(blob_to_img(tramo_dict[col]), use_container_width=True)
                        else:
                            st.caption("Sin foto")

            with tabs[2]:
                with st.form(f"editar_{tramo_id}"):
                    e1, e2, e3 = st.columns(3)
                    n_codigo = e1.text_input("Código", value=tramo_dict["codigo"])
                    n_nombre = e2.text_input("Nombre", value=tramo_dict["nombre"])
                    n_comunidad = e3.text_input("Comunidad", value=tramo_dict["comunidad"])
                    e4, e5, e6 = st.columns(3)
                    n_distrito = e4.text_input("Distrito", value=tramo_dict["distrito"])
                    n_provincia = e5.text_input("Provincia", value=tramo_dict["provincia"])
                    n_departamento = e6.text_input("Departamento", value=tramo_dict["departamento"])
                    e7, e8, e9 = st.columns(3)
                    n_long = e7.number_input("Longitud (km)", value=float(tramo_dict["longitud_km"]), step=0.1)
                    n_ancho = e8.number_input("Ancho (m)", value=float(tramo_dict["ancho_m"]), step=0.1)
                    n_sup = e9.selectbox("Tipo de superficie", TIPOS_SUPERFICIE,
                                          index=TIPOS_SUPERFICIE.index(tramo_dict["tipo_superficie"]) if tramo_dict["tipo_superficie"] in TIPOS_SUPERFICIE else 0)
                    e10, e11 = st.columns(2)
                    n_estado = e10.selectbox("Estado actual", ESTADOS_TRAMO,
                                              index=ESTADOS_TRAMO.index(tramo_dict["estado_actual"]) if tramo_dict["estado_actual"] in ESTADOS_TRAMO else 0)
                    n_resp = e11.text_input("Responsable", value=tramo_dict["responsable"])
                    n_tel = st.text_input("Teléfono", value=tramo_dict["telefono_responsable"] or "")
                    guardar = st.form_submit_button("💾 Guardar cambios")
                if guardar:
                    actualizar_tramo(tramo_id, {
                        "codigo": n_codigo, "nombre": n_nombre, "comunidad": n_comunidad,
                        "distrito": n_distrito, "provincia": n_provincia, "departamento": n_departamento,
                        "longitud_km": n_long, "ancho_m": n_ancho, "tipo_superficie": n_sup,
                        "estado_actual": n_estado, "responsable": n_resp, "telefono_responsable": n_tel,
                    })
                    st.success("✅ Tramo actualizado. Recarga la página o vuelve a filtrar para ver los cambios.")

            with tabs[3]:
                st.warning("Esta acción eliminará el tramo y todas sus fichas de daño asociadas.")
                if st.button("🗑️ Confirmar eliminación", key=f"del_{tramo_id}"):
                    eliminar_tramo(tramo_id)
                    st.success("Tramo eliminado. Vuelve a cargar la consulta.")

            with tabs[4]:
                danos_tramo_dl = df_danos()
                danos_tramo_dl = (danos_tramo_dl[danos_tramo_dl["tramo_id"] == tramo_id]
                                   if not danos_tramo_dl.empty else danos_tramo_dl)

                st.markdown(
                    f'<div style="background:#0b254508; border-left:4px solid #0b2545; '
                    f'border-radius:6px; padding:0.7rem 1rem; margin-bottom:0.8rem;">'
                    f'<b>Vista previa — Ficha Técnica del Tramo {tramo_dict["codigo"]}</b></div>',
                    unsafe_allow_html=True,
                )

                col_prev1, col_prev2 = st.columns([1, 1.4])
                with col_prev1:
                    foto_bytes_prev, foto_label_prev = _primera_foto_disponible(tramo_dict)
                    if foto_bytes_prev:
                        st.image(blob_to_img(foto_bytes_prev), caption=foto_label_prev, use_container_width=True)
                    else:
                        st.info("Este tramo no tiene fotografías cargadas todavía. Agrega una en el módulo de Registro.")

                with col_prev2:
                    st.markdown(f"**Código:** {tramo_dict['codigo']} &nbsp;·&nbsp; **Nombre:** {tramo_dict['nombre']}", unsafe_allow_html=True)
                    st.markdown(f"**Comunidad:** {tramo_dict['comunidad']} &nbsp;·&nbsp; **Longitud:** {tramo_dict['longitud_km']} km", unsafe_allow_html=True)
                    st.markdown(f"**Estado actual:** {badge_estado(tramo_dict['estado_actual'])}", unsafe_allow_html=True)
                    st.markdown(f"**Superficie:** {tramo_dict['tipo_superficie']} &nbsp;·&nbsp; **Ancho:** {tramo_dict['ancho_m']} m", unsafe_allow_html=True)
                    st.caption(f"Fichas de daño registradas: {len(danos_tramo_dl)}")

                st.markdown("##### Vista previa — 1-B Ficha del Itinerario del Camino Vecinal")
                st.dataframe(df_ficha_itinerario(tramo_dict).head(8), use_container_width=True, hide_index=True)
                st.caption(f"Mostrando 8 de {len(generar_grid_progresiva(tramo_dict['longitud_km']))} tramos de progresiva. El archivo descargable incluye la tabla completa.")

                st.markdown("##### Vista previa — 1-D Ficha Técnica de Daños")
                df_d_prev = df_ficha_danos(danos_tramo_dl)
                if df_d_prev.empty:
                    st.caption("Sin fichas de daño registradas aún para este tramo.")
                else:
                    st.dataframe(df_d_prev, use_container_width=True, hide_index=True)

                st.markdown("---")
                st.markdown("##### Descargar ficha técnica completa")
                d1, d2 = st.columns(2)
                with d1:
                    excel_bytes = exportar_ficha_excel(tramo_dict, danos_tramo_dl)
                    st.download_button(
                        "📊 Descargar ficha completa (Excel)",
                        data=excel_bytes,
                        file_name=f"Ficha_Tecnica_{tramo_dict['codigo']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"xlsx_{tramo_id}",
                        use_container_width=True,
                    )
                with d2:
                    pdf_ficha_bytes = exportar_ficha_pdf(tramo_dict, danos_tramo_dl)
                    download_button_bytes(
                        pdf_ficha_bytes,
                        f"Ficha_Tecnica_{tramo_dict['codigo']}.pdf",
                        "📄 Descargar ficha completa (PDF)",
                        "application/pdf",
                        f"pdf_{tramo_id}",
                    )
                st.caption("Ambos archivos incluyen: datos generales, fotografía principal, ficha 1-B (itinerario) y ficha 1-D (daños registrados).")


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 4 — FICHA TÉCNICA DE DAÑOS
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📋 FICHA TÉCNICA DE DAÑOS":
    st.markdown("## 📋 Ficha Técnica de Daños en Camino Vecinal")
    st.caption("Formato basado en las fichas 1-B / 1-D de Provías Descentralizado (Itinerario y Daños).")

    tramos = df_tramos()
    if tramos.empty:
        st.info("Primero registra al menos un tramo en **🛣️ Registro de Tramos**.")
    else:
        opciones = {f"{row.codigo} — {row.nombre} ({row.longitud_km} km)": (row.id, row.longitud_km)
                    for row in tramos.itertuples()}
        seleccion = st.selectbox("Tramo a inspeccionar", list(opciones.keys()))
        tramo_id, longitud_tramo = opciones[seleccion]

        conn = get_conn()
        row = conn.execute("SELECT * FROM tramos WHERE id=?", (tramo_id,)).fetchone()
        cols_t = [d[0] for d in conn.execute("SELECT * FROM tramos LIMIT 1").description]
        conn.close()
        tramo_dict_ficha = dict(zip(cols_t, row))

        def _fmt(v, sufijo=""):
            if v is None or v == "":
                return "No registrado"
            return f"{v}{sufijo}"

        st.markdown("---")
        st.markdown(f"### 📍 Tramo {tramo_dict_ficha['codigo']} — {tramo_dict_ficha['nombre']}")

        col_dt, col_rl = st.columns(2)
        with col_dt:
            st.markdown("##### 🔧 Datos técnicos")
            st.markdown(f"""
            <div class="kpi-card" style="padding:0.9rem 1.1rem;">
            <table style="width:100%; font-size:0.88rem;">
            <tr><td style="color:#5a6472;">Longitud del tramo</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['longitud_km'], ' km')}</td></tr>
            <tr><td style="color:#5a6472;">Ancho de plataforma</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['ancho_m'], ' m')}</td></tr>
            <tr><td style="color:#5a6472;">Altitud</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['altitud_msnm'], ' m s.n.m.')}</td></tr>
            <tr><td style="color:#5a6472;">Tipo de superficie</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['tipo_superficie'])}</td></tr>
            <tr><td style="color:#5a6472;">Estado del drenaje</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['estado_drenaje'])}</td></tr>
            <tr><td style="color:#5a6472;">Señalización</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['senalizacion'])}</td></tr>
            <tr><td style="color:#5a6472;">N° de puentes</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['n_puentes'])}</td></tr>
            <tr><td style="color:#5a6472;">N° de badenes</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['n_badenes'])}</td></tr>
            </table>
            </div>
            """, unsafe_allow_html=True)

        with col_rl:
            st.markdown("##### 🚧 Recursos y logística")
            costo_km_rut = (tramo_dict_ficha["costo_rutinario"] / tramo_dict_ficha["longitud_km"]
                             if tramo_dict_ficha["costo_rutinario"] and tramo_dict_ficha["longitud_km"] else None)
            st.markdown(f"""
            <div class="kpi-card" style="padding:0.9rem 1.1rem;">
            <table style="width:100%; font-size:0.88rem;">
            <tr><td style="color:#5a6472;">Accesibilidad</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['accesibilidad'])}</td></tr>
            <tr><td style="color:#5a6472;">Cantera / fuente mat.</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['cantera_fuente'])}</td></tr>
            <tr><td style="color:#5a6472;">Fuente de lastre</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['fuente_lastre'])}</td></tr>
            <tr><td style="color:#5a6472;">Última intervención</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['ultima_intervencion'])}</td></tr>
            <tr><td style="color:#5a6472;">Costo est. rutinario</td><td style="text-align:right; font-weight:600;">{('S/. ' + format(tramo_dict_ficha['costo_rutinario'], ',.0f')) if tramo_dict_ficha['costo_rutinario'] else 'No registrado'}</td></tr>
            <tr><td style="color:#5a6472;">Costo est. periódico</td><td style="text-align:right; font-weight:600;">{('S/. ' + format(tramo_dict_ficha['costo_periodico'], ',.0f')) if tramo_dict_ficha['costo_periodico'] else 'No registrado'}</td></tr>
            <tr><td style="color:#5a6472;">Costo/km rutinario</td><td style="text-align:right; font-weight:600;">{('S/. ' + format(costo_km_rut, ',.0f')) if costo_km_rut else 'No registrado'}</td></tr>
            <tr><td style="color:#5a6472;">Coordenadas GPS</td><td style="text-align:right; font-weight:600;">{(str(tramo_dict_ficha['gps_lat']) + ', ' + str(tramo_dict_ficha['gps_lon'])) if tramo_dict_ficha['gps_lat'] else 'No registrado'}</td></tr>
            </table>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("##### 📉 Curva de deterioro del IEC (Índice de Estado del Camino)")
        st.plotly_chart(curva_iec_deterioro(tramo_dict_ficha), use_container_width=True)
        st.caption(
            "Curva referencial según el Manual de Conservación Vial MCV-2014 (MTC). La estrella marca la "
            "posición estimada del tramo según su estado actual y el tiempo transcurrido desde su última intervención."
        )

        st.markdown("---")

        with st.form("form_ficha_dano", clear_on_submit=True):
            st.markdown("#### Progresiva y ubicación del daño")
            c1, c2, c3 = st.columns(3)
            prog_inicial = c1.number_input("Progresiva inicial (km)", min_value=0.0, step=0.05, format="%.3f")
            prog_final = c2.number_input("Progresiva final (km)", min_value=0.0, step=0.05, format="%.3f")
            longitud_afectada = c3.number_input("Longitud afectada (km)", min_value=0.0, step=0.01, format="%.3f")

            st.markdown("#### Características del daño")
            c4, c5, c6 = st.columns(3)
            tipo_dano = c4.selectbox("Tipo de daño", TIPOS_DANO)
            clase_densidad = c5.selectbox("Clase de densidad", CLASES_DENSIDAD,
                                           help="Solo aplica obligatoriamente si el tipo de daño es 'Baches o Huecos'.")
            fecha_inspeccion = c6.date_input("Fecha de inspección", value=date.today())

            observaciones = st.text_area("Observaciones técnicas", placeholder="Describe lo observado en campo...")
            foto_dano = st.file_uploader("Fotografía del daño", type=["jpg", "jpeg", "png"])

            calcular = st.form_submit_button("⚙️ Calcular clasificación automática y guardar", use_container_width=True)

        if calcular:
            falla = clasificar_falla(tipo_dano)
            peso = PESO_DANO.get(tipo_dano, 1)
            gravedad = calcular_gravedad(longitud_afectada, longitud_tramo, peso)
            estado_tramo_calc = calcular_estado(gravedad)
            transitabilidad = calcular_transitabilidad(estado_tramo_calc)
            prioridad = calcular_prioridad(estado_tramo_calc, transitabilidad)
            tiempo_rep = calcular_tiempo_reparacion(tipo_dano, max(longitud_afectada, 0.05))
            ict = calcular_ict(gravedad, clase_densidad, falla)
            pct_deterioro = round(100 - ict, 1)
            necesidad = calcular_necesidad_intervencion(prioridad)
            tipo_mantenimiento = determinar_tipo_mantenimiento(tipo_dano, gravedad, ict, transitabilidad)

            data = {
                "tramo_id": tramo_id, "progresiva_inicial": prog_inicial, "progresiva_final": prog_final,
                "longitud_afectada": longitud_afectada, "tipo_dano": tipo_dano, "tipo_falla": falla,
                "nivel_gravedad": gravedad, "clase_densidad": clase_densidad, "estado_tramo": estado_tramo_calc,
                "transitabilidad": transitabilidad, "necesidad_intervencion": necesidad,
                "tiempo_estimado_dias": tiempo_rep, "fecha_inspeccion": fecha_inspeccion.strftime("%Y-%m-%d"),
                "observaciones": observaciones, "ict": ict, "pct_deterioro": pct_deterioro,
                "prioridad": prioridad, "tipo_mantenimiento": tipo_mantenimiento,
                "foto_dano": img_to_blob(foto_dano),
            }
            dano_id = insertar_dano(data)
            st.session_state["ultimo_dano_id"] = dano_id
            st.session_state["ultimo_dano_tramo"] = tramo_id
            st.session_state["ultimo_dano_mant"] = tipo_mantenimiento
            st.session_state["ultimo_dano_dias"] = tiempo_rep

            st.success("✅ Ficha de daño registrada. Resultado de la clasificación automática:")
            r1, r2, r3, r4 = st.columns(4)
            r1.metric("Tipo de falla", falla)
            r2.metric("Nivel de gravedad", gravedad)
            r3.markdown(f"**Estado del tramo**<br>{badge_estado(estado_tramo_calc)}", unsafe_allow_html=True)
            r4.metric("Transitabilidad", transitabilidad)

            r5, r6, r7, r8 = st.columns(4)
            r5.metric("Prioridad de intervención", prioridad)
            r6.metric("Tiempo estimado de reparación", f"{tiempo_rep} días")
            r7.metric("ICT (Índice de Condición)", f"{ict} / 100")
            r8.metric("% de deterioro", f"{pct_deterioro}%")

            color_mant = MANTENIMIENTOS[tipo_mantenimiento]["color"]
            st.markdown(
                f'<div style="background:{color_mant}15; border-left:5px solid {color_mant}; '
                f'border-radius:8px; padding:0.8rem 1rem; margin-top:0.5rem;">'
                f'<b style="color:{color_mant};">Tipo de mantenimiento recomendado: {tipo_mantenimiento}</b><br>'
                f'<span style="font-size:0.88rem;">{MANTENIMIENTOS[tipo_mantenimiento]["descripcion"]}</span></div>',
                unsafe_allow_html=True,
            )
            st.info(f"**Necesidad de intervención:** {necesidad}")
            st.caption("👉 Ve al módulo **🗓️ Programación de Mantenimiento** para programar esta intervención con expediente, fecha y costo.")

        st.markdown("---")
        st.markdown("#### Histórico de fichas de daños registradas")
        danos = df_danos()
        if danos.empty:
            st.caption("Aún no hay fichas registradas.")
        else:
            cols_show = ["codigo_tramo", "nombre_tramo", "progresiva_inicial", "progresiva_final",
                         "longitud_afectada", "tipo_dano", "tipo_falla", "nivel_gravedad",
                         "estado_tramo", "transitabilidad", "prioridad", "tipo_mantenimiento", "ict", "pct_deterioro",
                         "tiempo_estimado_dias", "fecha_inspeccion"]
            st.dataframe(danos[cols_show].rename(columns={
                "codigo_tramo": "Código", "nombre_tramo": "Tramo", "progresiva_inicial": "Prog. Inicial",
                "progresiva_final": "Prog. Final", "longitud_afectada": "Long. Afectada (km)",
                "tipo_dano": "Tipo de Daño", "tipo_falla": "Falla", "nivel_gravedad": "Gravedad",
                "estado_tramo": "Estado", "transitabilidad": "Transitabilidad", "prioridad": "Prioridad",
                "tipo_mantenimiento": "Mantenimiento Sugerido",
                "ict": "ICT", "pct_deterioro": "% Deterioro", "tiempo_estimado_dias": "Días Rep.",
                "fecha_inspeccion": "Fecha",
            }), use_container_width=True, hide_index=True)

            st.markdown("##### 🏆 Ranking de tramos críticos (menor ICT)")
            ranking = danos.groupby(["codigo_tramo", "nombre_tramo"])["ict"].mean().reset_index()
            ranking = ranking.sort_values("ict").rename(columns={"ict": "ICT promedio"})
            ranking.index = range(1, len(ranking) + 1)
            st.dataframe(ranking, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 5 — PROGRAMACIÓN DE MANTENIMIENTO (5 tipos de mantenimiento)
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "🗓️ PROGRAMACIÓN DE MANTENIMIENTO":
    st.markdown("## 🗓️ Programación de Mantenimiento Vial")
    st.caption("Programa intervenciones según los 5 tipos de mantenimiento definidos en el informe del proyecto.")

    with st.expander("ℹ️ Tipos de mantenimiento vial y actividades asociadas", expanded=False):
        for nombre, info in MANTENIMIENTOS.items():
            st.markdown(
                f'<div style="border-left:5px solid {info["color"]}; padding:0.4rem 0.8rem; margin-bottom:0.5rem;">'
                f'<b style="color:{info["color"]};">{nombre}</b><br>'
                f'<span style="font-size:0.86rem;">{info["descripcion"]}</span><br>'
                f'<span style="font-size:0.82rem; color:#444;">• ' + "<br>• ".join(info["actividades"]) + '</span>'
                f'</div>', unsafe_allow_html=True,
            )

    tramos = df_tramos()
    danos = df_danos()

    if tramos.empty:
        st.info("Primero registra al menos un tramo.")
    else:
        st.markdown("---")
        st.markdown("#### Nueva intervención programada")

        opciones_tramo = {f"{row.codigo} — {row.nombre}": row.id for row in tramos.itertuples()}

        dano_default = None
        tramo_default_id = None
        if "ultimo_dano_id" in st.session_state:
            dano_default = st.session_state["ultimo_dano_id"]
            tramo_default_id = st.session_state["ultimo_dano_tramo"]

        with st.form("form_programacion", clear_on_submit=True):
            c1, c2 = st.columns(2)
            sel_tramo_label = c1.selectbox(
                "Tramo", list(opciones_tramo.keys()),
                index=list(opciones_tramo.values()).index(tramo_default_id) if tramo_default_id in opciones_tramo.values() else 0,
            )
            tramo_id_sel = opciones_tramo[sel_tramo_label]

            # Fichas de daño asociadas a ese tramo (opcional, para vincular)
            danos_tramo = danos[danos["tramo_id"] == tramo_id_sel] if not danos.empty else pd.DataFrame()
            opciones_dano = {"— Sin vincular a ficha de daño —": None}
            if not danos_tramo.empty:
                for row in danos_tramo.itertuples():
                    opciones_dano[f"#{row.id} · {row.tipo_dano} · ICT {row.ict} · sugerido: {row.tipo_mantenimiento}"] = row.id
            sel_dano_label = c2.selectbox("Ficha de daño vinculada (opcional)", list(opciones_dano.keys()))
            dano_id_sel = opciones_dano[sel_dano_label]

            sugerido = None
            if dano_id_sel is not None:
                sugerido = danos_tramo[danos_tramo["id"] == dano_id_sel]["tipo_mantenimiento"].values[0]

            tipo_mant = st.selectbox(
                "Tipo de mantenimiento *", list(MANTENIMIENTOS.keys()),
                index=list(MANTENIMIENTOS.keys()).index(sugerido) if sugerido in MANTENIMIENTOS else 0,
                help="Si vinculas una ficha de daño, se preselecciona el tipo sugerido automáticamente (puedes cambiarlo).",
            )
            actividades_sel = st.multiselect(
                "Actividades a ejecutar *", MANTENIMIENTOS[tipo_mant]["actividades"],
                default=MANTENIMIENTOS[tipo_mant]["actividades"],
            )

            c3, c4, c5 = st.columns(3)
            expediente = c3.text_input("N° de Expediente Técnico", placeholder="Ej: ET-MDO-006")
            fecha_programada = c4.date_input("Fecha programada de inicio", value=date.today())
            duracion_dias = c5.number_input("Duración estimada (días)", min_value=1.0, value=15.0, step=1.0)

            c6, c7 = st.columns(2)
            costo_estimado = c6.number_input("Costo estimado (S/.)", min_value=0.0, step=100.0, value=10000.0)
            estado_int = c7.selectbox("Estado de la intervención", ESTADOS_INTERVENCION)

            responsable_int = st.text_input("Responsable de la intervención",
                                             value="Subgerencia de Gestión de Riesgos y Mantenimiento")
            observaciones_int = st.text_area("Observaciones")

            guardar_int = st.form_submit_button("💾 Guardar programación", use_container_width=True)

        if guardar_int:
            insertar_intervencion({
                "tramo_id": tramo_id_sel, "dano_id": dano_id_sel, "tipo_mantenimiento": tipo_mant,
                "actividades": "; ".join(actividades_sel), "expediente": expediente,
                "fecha_programada": fecha_programada.strftime("%Y-%m-%d"), "duracion_dias": duracion_dias,
                "costo_estimado": costo_estimado, "estado": estado_int, "responsable": responsable_int,
                "observaciones": observaciones_int,
            })
            st.success(f"✅ Intervención de **{tipo_mant}** programada para el tramo seleccionado.")
            for k in ["ultimo_dano_id", "ultimo_dano_tramo", "ultimo_dano_mant", "ultimo_dano_dias"]:
                st.session_state.pop(k, None)

        st.markdown("---")
        st.markdown("#### Cronograma de intervenciones programadas")
        intervenciones = df_intervenciones()
        if intervenciones.empty:
            st.caption("Aún no hay intervenciones programadas.")
        else:
            intervenciones["fecha_fin"] = pd.to_datetime(intervenciones["fecha_programada"]) + pd.to_timedelta(
                intervenciones["duracion_dias"], unit="D")
            fig_g = px.timeline(
                intervenciones, x_start="fecha_programada", x_end="fecha_fin",
                y="nombre_tramo", color="tipo_mantenimiento",
                color_discrete_map={k: v["color"] for k, v in MANTENIMIENTOS.items()},
                hover_data=["expediente", "estado", "costo_estimado"],
            )
            fig_g.update_yaxes(autorange="reversed", title="")
            fig_g.update_layout(height=380, legend_title="Tipo de mantenimiento")
            st.plotly_chart(fig_g, use_container_width=True)

            st.dataframe(
                intervenciones[["codigo_tramo", "nombre_tramo", "tipo_mantenimiento", "expediente",
                                 "fecha_programada", "duracion_dias", "costo_estimado", "estado", "responsable"]]
                .rename(columns={"codigo_tramo": "Código", "nombre_tramo": "Tramo",
                                  "tipo_mantenimiento": "Tipo Mant.", "fecha_programada": "Inicio",
                                  "duracion_dias": "Días", "costo_estimado": "Costo (S/.)"}),
                use_container_width=True, hide_index=True,
            )

            st.markdown("##### Distribución de costo programado por tipo de mantenimiento")
            costo_mant = intervenciones.groupby("tipo_mantenimiento")["costo_estimado"].sum().reset_index()
            fig_pie = px.pie(costo_mant, names="tipo_mantenimiento", values="costo_estimado", hole=0.4,
                              color="tipo_mantenimiento",
                              color_discrete_map={k: v["color"] for k, v in MANTENIMIENTOS.items()})
            st.plotly_chart(fig_pie, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 6 — REPORTES
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📊 REPORTES":
    st.markdown("## 📊 Reportes")
    tramos = df_tramos()
    danos = df_danos()

    _tac_r = st.session_state.get("tramo_activo_codigo")
    if _tac_r:
        tramos = tramos[tramos["codigo"] == _tac_r]
        if not danos.empty:
            danos = danos[danos["codigo_tramo"] == _tac_r]
        st.info(f"📍 Reportes filtrados al tramo **{_tac_r}**. Cambia el filtro lateral para ver la red completa.")

    if tramos.empty:
        st.info("No hay datos suficientes para generar reportes todavía.")
    else:
        tab1, tab2, tab3, tab4, tab5 = st.tabs(
            ["📦 Inventario y Estado", "⚠️ Daños", "💰 Costos e Intervenciones",
             "🗂️ Histórico de Expedientes", "⬇️ Exportar"])

        # ---- TAB 1: Inventario y estado de la red ----
        with tab1:
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("##### Estado de la red vial (N° de tramos)")
                fig = px.bar(tramos["estado_actual"].value_counts().reset_index(),
                             x="estado_actual", y="count",
                             color="estado_actual",
                             color_discrete_map={"Bueno": "#43a047", "Regular": "#fdd835",
                                                  "Malo": "#fb8c00", "Muy Malo": "#e53935"},
                             text_auto=True)
                fig.update_layout(showlegend=False, xaxis_title="", yaxis_title="N° tramos", height=340)
                st.plotly_chart(fig, use_container_width=True)
            with c2:
                st.markdown("##### Kilómetros afectados por estado")
                km_estado = tramos.groupby("estado_actual")["longitud_km"].sum().reset_index()
                fig2 = px.pie(km_estado, names="estado_actual", values="longitud_km", hole=0.4,
                              color="estado_actual",
                              color_discrete_map={"Bueno": "#43a047", "Regular": "#fdd835",
                                                   "Malo": "#fb8c00", "Muy Malo": "#e53935"})
                fig2.update_layout(height=340)
                st.plotly_chart(fig2, use_container_width=True)

            st.markdown("##### Inventario completo de tramos")
            st.dataframe(tramos, use_container_width=True, hide_index=True)

        # ---- TAB 2: Daños ----
        with tab2:
            if danos.empty:
                st.info("No hay fichas de daños registradas aún.")
            else:
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("##### Daños por tipo")
                    fig3 = px.bar(danos["tipo_dano"].value_counts().reset_index(),
                                  x="tipo_dano", y="count", color="count",
                                  color_continuous_scale="Reds", text_auto=True)
                    fig3.update_layout(xaxis_title="", yaxis_title="N° de casos", height=360)
                    st.plotly_chart(fig3, use_container_width=True)
                with c2:
                    st.markdown("##### Daños por nivel de gravedad")
                    fig4 = px.pie(danos, names="nivel_gravedad", hole=0.4,
                                  color="nivel_gravedad",
                                  color_discrete_map={"Sin Deterioro": "#43a047", "Leve": "#9ccc65",
                                                       "Moderada": "#fb8c00", "Severa": "#e53935"})
                    fig4.update_layout(height=360)
                    st.plotly_chart(fig4, use_container_width=True)

                st.markdown("##### Histograma de % de deterioro (ICT invertido)")
                fig5 = px.histogram(danos, x="pct_deterioro", nbins=10, color_discrete_sequence=["#c62828"])
                fig5.update_layout(xaxis_title="% de deterioro", yaxis_title="Frecuencia", height=320)
                st.plotly_chart(fig5, use_container_width=True)

                st.markdown("##### Kilómetros afectados por tipo de daño")
                km_dano = danos.groupby("tipo_dano")["longitud_afectada"].sum().reset_index()
                fig6 = px.bar(km_dano.sort_values("longitud_afectada", ascending=False),
                              x="tipo_dano", y="longitud_afectada", color="longitud_afectada",
                              color_continuous_scale="Oranges", text_auto=".2f")
                fig6.update_layout(xaxis_title="", yaxis_title="km afectados", height=340)
                st.plotly_chart(fig6, use_container_width=True)

                st.markdown("##### Tramos críticos (Prioridad Urgente / Alta)")
                criticos = danos[danos["prioridad"].isin(["Urgente", "Alta"])][
                    ["codigo_tramo", "nombre_tramo", "tipo_dano", "estado_tramo",
                     "transitabilidad", "prioridad", "ict", "tiempo_estimado_dias"]]
                st.dataframe(criticos, use_container_width=True, hide_index=True)

        # ---- TAB 3: Costos / necesidades de mantenimiento ----
        with tab3:
            if danos.empty:
                st.info("No hay fichas de daños registradas aún.")
            else:
                COSTO_REFERENCIAL_KM = {  # S/. por km — referencial MCV-2014 / benchmarks de campo
                    "Deformación": 9500, "Erosión": 11000, "Baches o Huecos": 14500,
                    "Encalaminado": 8000, "Lodazal": 17500, "Cruce de Agua": 22000,
                    "Hundimiento": 26000, "Pérdida de Plataforma": 31000,
                    "Deslizamiento": 35000, "Falla de Drenaje": 19500,
                }
                danos["costo_estimado"] = danos.apply(
                    lambda r: round(r["longitud_afectada"] * COSTO_REFERENCIAL_KM.get(r["tipo_dano"], 12000), 0),
                    axis=1,
                )
                c1, c2, c3 = st.columns(3)
                c1.metric("Costo estimado total de intervención", f"S/. {danos['costo_estimado'].sum():,.0f}")
                c2.metric("Km totales afectados", f"{danos['longitud_afectada'].sum():.2f} km")
                c3.metric("Días totales estimados de reparación", f"{danos['tiempo_estimado_dias'].sum():.0f} días")

                st.markdown("##### Costo estimado por tipo de daño")
                costo_tipo = danos.groupby("tipo_dano")["costo_estimado"].sum().reset_index()
                fig7 = px.bar(costo_tipo.sort_values("costo_estimado", ascending=False),
                              x="tipo_dano", y="costo_estimado", color="costo_estimado",
                              color_continuous_scale="Blues", text_auto=".2s")
                fig7.update_layout(xaxis_title="", yaxis_title="S/.", height=340)
                st.plotly_chart(fig7, use_container_width=True)

                st.markdown("##### Necesidades de mantenimiento por tramo")
                necesidades = danos.groupby(["codigo_tramo", "nombre_tramo", "necesidad_intervencion"]).size().reset_index(name="N° fichas")
                st.dataframe(necesidades, use_container_width=True, hide_index=True)

        # ---- TAB 4: Histórico de expedientes ----
        with tab4:
            st.markdown("##### Expedientes técnicos ejecutados — Municipalidad Distrital de Ocongate")
            df_exp = pd.DataFrame(EXPEDIENTES_HISTORICOS)
            st.dataframe(
                df_exp.style.format({"Costo total ejecutado (S/.)": "S/. {:,.2f}"}),
                use_container_width=True, hide_index=True,
            )
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("##### Inversión ejecutada por año")
                inv_anio = df_exp.groupby("Año")["Costo total ejecutado (S/.)"].sum().reset_index()
                fig_exp1 = px.bar(inv_anio, x="Año", y="Costo total ejecutado (S/.)",
                                  color="Costo total ejecutado (S/.)", color_continuous_scale="Blues",
                                  text_auto=".2s")
                fig_exp1.update_layout(height=340)
                st.plotly_chart(fig_exp1, use_container_width=True)
            with c2:
                st.markdown("##### Costo ejecutado por expediente")
                fig_exp2 = px.bar(df_exp.sort_values("Costo total ejecutado (S/.)", ascending=True),
                                  x="Costo total ejecutado (S/.)", y="Expediente", orientation="h",
                                  color="Costo total ejecutado (S/.)", color_continuous_scale="Tealgrn",
                                  text_auto=".2s")
                fig_exp2.update_layout(height=340)
                st.plotly_chart(fig_exp2, use_container_width=True)

            total_historico = df_exp["Costo total ejecutado (S/.)"].sum()
            st.metric("Inversión histórica total ejecutada (2021-2024)", f"S/. {total_historico:,.2f}")
            st.caption(
                "Esta información histórica sirve como línea base de costos reales de intervención "
                "correctiva, en contraste con el ahorro proyectado mediante el enfoque preventivo (PMA)."
            )

        # ---- TAB 5: Exportar ----
        with tab5:
            st.markdown("#### Exportar reportes")
            dfs_excel = {"Inventario_Tramos": tramos}
            if not danos.empty:
                dfs_excel["Fichas_Danos"] = danos.drop(columns=["foto_dano"], errors="ignore")
                dfs_excel["Ranking_Criticos"] = (
                    danos.groupby(["codigo_tramo", "nombre_tramo"])["ict"].mean()
                    .reset_index().sort_values("ict").rename(columns={"ict": "ICT_promedio"})
                )

            c1, c2 = st.columns(2)
            with c1:
                st.download_button(
                    "⬇️ Descargar reporte completo (Excel)",
                    data=exportar_excel(dfs_excel),
                    file_name=f"Reporte_Trochas_Ocongate_{date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with c2:
                pdf_bytes = exportar_pdf_resumen(
                    "Reporte de Inventario — Trochas Carrozables, Ocongate",
                    tramos[["codigo", "nombre", "comunidad", "longitud_km", "estado_actual"]],
                    texto_intro=f"Total de tramos: {len(tramos)} | Km inventariados: {tramos['longitud_km'].sum():.2f} km",
                )
                download_button_bytes(
                    pdf_bytes,
                    f"Reporte_Trochas_Ocongate_{date.today().strftime('%Y%m%d')}.pdf",
                    "⬇️ Descargar reporte resumen (PDF)",
                    "application/pdf",
                    "pdf_main",
                )


st.markdown("---")
st.caption("Plataforma desarrollada como demostración técnica — Proyecto Preprofesional UTEC 2026-I · "
           "Frank Puma Mamani (202220055) · Municipalidad Distrital de Ocongate.")
