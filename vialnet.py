# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║  VIALNET                                                                     ║
║  Plataforma Inteligente de Gestión de Mantenimiento Vial Rural              ║
║  Municipalidad Distrital de Ocongate — Quispicanchi, Cusco, Perú            ║
║                                                                               ║
║  Autor  : Frank Puma Mamani | Código: 202220055                             ║
║  Curso  : Proyecto Preprofesional — Teoría 7, UTEC 2026-I                    ║
║  Docente: Mg. Fernandez Choquepuma, Miguel Ángel                            ║
╚══════════════════════════════════════════════════════════════════════════════╝

INSTALACIÓN:
    pip install -r requirements.txt

EJECUCIÓN LOCAL:
    streamlit run vialnet.py

ESTRUCTURA (6 módulos):
    1. 🏠 INICIO                          — Presentación, KPIs rápidos, gráficos resumen
    2. 🛣️ REGISTRO DE TRAMOS               — Alta de tramos + evidencia fotográfica + adjuntos
    3. 📂 CONSULTA DE TRAMOS               — Buscar / filtrar / editar / eliminar / descargar
    4. 📋 FICHA TÉCNICA DE DAÑOS           — Ficha tipo Provías + clasificación automática + ICT
    5. 🗓️ PROGRAMACIÓN DE MANTENIMIENTO    — Cronograma y programación de intervenciones
    6. 📊 REPORTES                        — Reportes automáticos + gráficos Plotly + export Excel/PDF

Persistencia: PostgreSQL en la nube (Supabase) — se crea automáticamente al iniciar.
Archivos (fotos, PDFs adjuntos): Supabase Storage — accesibles por URL pública desde cualquier dispositivo.
"""

import streamlit as st
import pandas as pd
import psycopg2
import psycopg2.pool
import psycopg2.errors
import psycopg2.extensions
import plotly.express as px
import plotly.graph_objects as go
from datetime import datetime, date
from io import BytesIO
import base64
import os
import re
import uuid
import requests
import warnings

warnings.filterwarnings("ignore", message="pandas only supports SQLAlchemy")

# ──────────────────────────────────────────────────────────────────────────────
# CONFIGURACIÓN GENERAL
# ──────────────────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="VialNet — Gestión de Mantenimiento Vial",
    page_icon="🛣️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ──────────────────────────────────────────────────────────────────────────────
# BASE DE DATOS EN LA NUBE — POSTGRESQL (SUPABASE)
#
# Requiere en st.secrets (Streamlit Cloud → Settings → Secrets, o
# .streamlit/secrets.toml en local, NUNCA subido al repositorio):
#
#   SUPABASE_DB_URL = "postgresql://postgres:[PASSWORD]@db.xxxx.supabase.co:5432/postgres"
#   SUPABASE_URL           = "https://xxxx.supabase.co"
#   SUPABASE_SERVICE_KEY   = "eyJ..."   # service_role key (solo backend, nunca en frontend)
#   SUPABASE_STORAGE_BUCKET = "vialnet-archivos"   # opcional, default abajo
# ──────────────────────────────────────────────────────────────────────────────
def _secret(nombre: str, default=None):
    """Lee una variable desde st.secrets sin reventar si no existe (útil en dev local)."""
    try:
        return st.secrets[nombre]
    except (KeyError, FileNotFoundError):
        return os.environ.get(nombre, default)


DEFAULT_BUCKET = _secret("SUPABASE_STORAGE_BUCKET", "vialnet-archivos")


def _validar_configuracion_cloud():
    faltantes = [k for k in ("SUPABASE_DB_URL", "SUPABASE_URL", "SUPABASE_SERVICE_KEY") if not _secret(k)]
    if faltantes:
        st.error(
            "⚠️ **VialNet no está configurado para la nube.** Faltan estas claves en `st.secrets`: "
            + ", ".join(f"`{k}`" for k in faltantes) +
            ".\n\nConfigúralas en *Settings → Secrets* (Streamlit Community Cloud) o en "
            "`.streamlit/secrets.toml` en local. Revisa el README para el detalle paso a paso."
        )
        st.stop()


_validar_configuracion_cloud()


@st.cache_resource(show_spinner=False)
def _pool_conexiones():
    """Pool de conexiones reutilizables a PostgreSQL — evita abrir un socket nuevo en cada rerun."""
    return psycopg2.pool.SimpleConnectionPool(
        1, 10, dsn=_secret("SUPABASE_DB_URL"), sslmode="require", connect_timeout=10,
    )


@st.cache_resource(show_spinner=False)
def _cliente_supabase():
    """Cliente de Supabase (solo se usa para Storage: subida/lectura de fotos y adjuntos)."""
    from supabase import create_client
    return create_client(_secret("SUPABASE_URL"), _secret("SUPABASE_SERVICE_KEY"))


_QMARK_RE = re.compile(r"\?")


def _a_postgres(sql: str) -> str:
    """Traduce los placeholders `?` (estilo sqlite) a `%s` (estilo psycopg2/Postgres)."""
    return _QMARK_RE.sub("%s", sql)


class _CursorPG:
    """Envoltura ligera sobre un cursor de psycopg2 que traduce `?` → `%s` automáticamente,
    para que el resto del código (escrito originalmente para sqlite3) no tenga que cambiar."""

    def __init__(self, cursor):
        self._cursor = cursor
        self.lastrowid = None

    def execute(self, sql, params=None):
        self._cursor.execute(_a_postgres(sql), params or None)
        return self

    def executemany(self, sql, seq_params):
        self._cursor.executemany(_a_postgres(sql), seq_params)
        return self

    def fetchone(self):
        return self._cursor.fetchone()

    def fetchall(self):
        return self._cursor.fetchall()

    @property
    def description(self):
        return self._cursor.description


class _ConnPG:
    """Envoltura sobre una conexión de psycopg2 que imita la interfaz de sqlite3.Connection
    usada en el resto del código (`conn.execute(...)`, `conn.cursor()`, `conn.commit()`,
    `conn.close()`), devolviendo la conexión física al pool en vez de cerrarla de verdad."""

    def __init__(self, raw_conn, pool_):
        self._conn = raw_conn
        self._pool = pool_

    def cursor(self):
        return _CursorPG(self._conn.cursor())

    def execute(self, sql, params=None):
        cur = self._conn.cursor()
        cur.execute(_a_postgres(sql), params or None)
        return cur

    def commit(self):
        self._conn.commit()

    def rollback(self):
        self._conn.rollback()

    def close(self):
        # Devuelve la conexión al pool en vez de cerrarla; si quedó en una transacción
        # fallida (p. ej. tras un IntegrityError no controlado), se revierte primero.
        try:
            if self._conn.get_transaction_status() != psycopg2.extensions.TRANSACTION_STATUS_IDLE:
                self._conn.rollback()
        except Exception:
            pass
        self._pool.putconn(self._conn)


def get_conn() -> _ConnPG:
    raw_conn = _pool_conexiones().getconn()
    return _ConnPG(raw_conn, _pool_conexiones())


# ──────────────────────────────────────────────────────────────────────────────
# ALMACENAMIENTO EN LA NUBE — SUPABASE STORAGE (fotos, PDFs adjuntos)
# ──────────────────────────────────────────────────────────────────────────────
def subir_archivo_storage(file_bytes: bytes, nombre_original: str, carpeta: str) -> str | None:
    """Sube un archivo (foto o PDF) a Supabase Storage y devuelve su URL pública.

    `carpeta` agrupa los archivos por tipo (ej. 'fotos-tramos', 'adjuntos-pdf', 'fotos-danos').
    Devuelve None si la subida falla (se maneja con un mensaje de error en la UI, sin romper el flujo).
    """
    if not file_bytes:
        return None
    try:
        ext = (nombre_original.rsplit(".", 1)[-1] if "." in nombre_original else "bin").lower()
        ruta = f"{carpeta}/{date.today().strftime('%Y/%m')}/{uuid.uuid4().hex}.{ext}"
        content_type = {
            "jpg": "image/jpeg", "jpeg": "image/jpeg", "png": "image/png", "pdf": "application/pdf",
        }.get(ext, "application/octet-stream")

        cliente = _cliente_supabase()
        cliente.storage.from_(DEFAULT_BUCKET).upload(
            ruta, file_bytes, {"content-type": content_type, "upsert": "true"}
        )
        return cliente.storage.from_(DEFAULT_BUCKET).get_public_url(ruta)
    except Exception as e:
        st.error(f"⚠️ No se pudo subir el archivo **{nombre_original}** a Supabase Storage: {e}")
        return None


def eliminar_archivo_storage(url: str):
    """Elimina un archivo de Supabase Storage a partir de su URL pública (best-effort)."""
    if not url:
        return
    try:
        marcador = f"/object/public/{DEFAULT_BUCKET}/"
        if marcador in url:
            ruta = url.split(marcador, 1)[1]
            _cliente_supabase().storage.from_(DEFAULT_BUCKET).remove([ruta])
    except Exception:
        pass  # Eliminación de archivo huérfano no es crítica; no interrumpe el flujo del usuario.


@st.cache_data(show_spinner=False, ttl=3600)
def descargar_bytes_storage(url: str):
    """Descarga los bytes de un archivo alojado en Storage (para incrustarlo en PDF/Excel).
    Se cachea por 1 hora para no re-descargar la misma imagen varias veces en una sesión."""
    if not url:
        return None
    try:
        resp = requests.get(url, timeout=15)
        resp.raise_for_status()
        return resp.content
    except Exception:
        return None

# ──────────────────────────────────────────────────────────────────────────────
# SIEMBRA INICIAL — Los 5 tramos reales del proyecto (M1-M5, Ocongate)
# ──────────────────────────────────────────────────────────────────────────────
TRAMOS_PROYECTO = [
    {
        "codigo": "M1", "nombre": "Pacchanta Alta – UPIS", "comunidad": "Pacchanta Alta",
        "longitud_km": 21.12, "ancho_m": 4.5, "estado_actual": "Malo",
        "altitud_msnm": 4250, "estado_drenaje": "Deficiente — cunetas colmatadas",
        "senalizacion": "Ausente", "n_puentes": 2, "n_badenes": 5,
        "accesibilidad": "Estacional (cerrado dic-mar)", "cantera_fuente": "Sí — Cantera Pacchanta (2.3 km)",
        "fuente_lastre": "Río Ausangate (material aluvial)", "ultima_intervencion": "2023-06-15",
        "costo_rutinario": 184000, "costo_periodico": 336000, "gps_lat": -13.6458, "gps_lon": -71.1234,
    },
    {
        "codigo": "M2", "nombre": "Mahuayani – Ocongate", "comunidad": "Mahuayani",
        "longitud_km": 15.00, "ancho_m": 4.5, "estado_actual": "Regular",
        "altitud_msnm": 4180, "estado_drenaje": "Regular — cunetas parcialmente operativas",
        "senalizacion": "Parcial (hitos kilométricos)", "n_puentes": 1, "n_badenes": 3,
        "accesibilidad": "Permanente", "cantera_fuente": "Sí — Cantera Mahuayani (1.8 km)",
        "fuente_lastre": "Quebrada local (material aluvial)", "ultima_intervencion": "2022-09-10",
        "costo_rutinario": 169575, "costo_periodico": 399000, "gps_lat": -13.6010, "gps_lon": -71.3520,
    },
    {
        "codigo": "M3", "nombre": "Palcca Central – Alta", "comunidad": "Palcca Central",
        "longitud_km": 30.00, "ancho_m": 4.0, "estado_actual": "Muy Malo",
        "altitud_msnm": 4400, "estado_drenaje": "Crítico — sin cunetas en la mayor parte del tramo",
        "senalizacion": "Ausente", "n_puentes": 3, "n_badenes": 8,
        "accesibilidad": "Estacional (cerrado dic-abr)", "cantera_fuente": "No identificada",
        "fuente_lastre": "Por definir (requiere estudio de canteras)", "ultima_intervencion": "2022-04-22",
        "costo_rutinario": 405000, "costo_periodico": 630000, "gps_lat": -13.7200, "gps_lon": -71.0850,
    },
    {
        "codigo": "M4", "nombre": "Pacchanta Baja – Cruce UPIS", "comunidad": "Pacchanta Baja",
        "longitud_km": 18.50, "ancho_m": 4.5, "estado_actual": "Malo",
        "altitud_msnm": 4150, "estado_drenaje": "Deficiente — alcantarillas obstruidas",
        "senalizacion": "Parcial (señalización informativa)", "n_puentes": 2, "n_badenes": 4,
        "accesibilidad": "Estacional (cerrado ene-mar)", "cantera_fuente": "Sí — Cantera Pacchanta Baja (3.1 km)",
        "fuente_lastre": "Río Mapacho (material aluvial)", "ultima_intervencion": "2023-02-18",
        "costo_rutinario": 238625, "costo_periodico": 483000, "gps_lat": -13.6520, "gps_lon": -71.1450,
    },
    {
        "codigo": "M5", "nombre": "Ausangate – Comunidad", "comunidad": "Ausangate",
        "longitud_km": 12.00, "ancho_m": 4.0, "estado_actual": "Regular",
        "altitud_msnm": 4600, "estado_drenaje": "Regular — mantenimiento comunal periódico",
        "senalizacion": "Preventiva (gestión comunal)", "n_puentes": 1, "n_badenes": 2,
        "accesibilidad": "Permanente (acceso turístico)", "cantera_fuente": "Sí — Cantera Ausangate (1.2 km)",
        "fuente_lastre": "Glaciar Ausangate (material morrénico)", "ultima_intervencion": "2024-05-30",
        "costo_rutinario": 94929, "costo_periodico": 220000, "gps_lat": -13.7850, "gps_lon": -71.1650,
    },
]


def seed_tramos_proyecto():
    """Inserta los 5 tramos reales del proyecto si la tabla está vacía o si falta alguno."""
    conn = get_conn()
    c = conn.cursor()
    existentes = {row[0] for row in c.execute("SELECT codigo FROM tramos").fetchall()}
    for t in TRAMOS_PROYECTO:
        if t["codigo"] in existentes:
            continue
        c.execute("""
            INSERT INTO tramos (codigo, nombre, comunidad, distrito, provincia, departamento,
                longitud_km, ancho_m, tipo_superficie, estado_actual, responsable,
                telefono_responsable, fecha_registro, adjunto_pdf, adjunto_pdf_nombre,
                altitud_msnm, estado_drenaje, senalizacion, n_puentes, n_badenes,
                accesibilidad, cantera_fuente, fuente_lastre, ultima_intervencion,
                costo_rutinario, costo_periodico, gps_lat, gps_lon)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            t["codigo"], t["nombre"], t["comunidad"], "Ocongate", "Quispicanchi", "Cusco",
            t["longitud_km"], t["ancho_m"], "Afirmado (AF)", t["estado_actual"],
            "Subgerencia de Gestión de Riesgos y Mantenimiento", "",
            date.today().strftime("%Y-%m-%d"), None, None,
            t["altitud_msnm"], t["estado_drenaje"], t["senalizacion"], t["n_puentes"], t["n_badenes"],
            t["accesibilidad"], t["cantera_fuente"], t["fuente_lastre"], t["ultima_intervencion"],
            t["costo_rutinario"], t["costo_periodico"], t["gps_lat"], t["gps_lon"],
        ))
    conn.commit()
    conn.close()


def init_db():
    conn = get_conn()
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS tramos (
            id SERIAL PRIMARY KEY,
            codigo TEXT UNIQUE,
            nombre TEXT,
            comunidad TEXT,
            distrito TEXT,
            provincia TEXT,
            departamento TEXT,
            longitud_km REAL,
            ancho_m REAL,
            tipo_superficie TEXT,
            estado_actual TEXT,
            responsable TEXT,
            telefono_responsable TEXT,
            fecha_registro TEXT,
            adjunto_pdf TEXT,
            adjunto_pdf_nombre TEXT,
            altitud_msnm REAL,
            estado_drenaje TEXT,
            senalizacion TEXT,
            n_puentes INTEGER,
            n_badenes INTEGER,
            accesibilidad TEXT,
            cantera_fuente TEXT,
            fuente_lastre TEXT,
            ultima_intervencion TEXT,
            costo_rutinario REAL,
            costo_periodico REAL,
            gps_lat REAL,
            gps_lon REAL
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS danos (
            id SERIAL PRIMARY KEY,
            tramo_id INTEGER,
            progresiva_inicial REAL,
            progresiva_final REAL,
            longitud_afectada REAL,
            tipo_dano TEXT,
            tipo_falla TEXT,
            nivel_gravedad TEXT,
            clase_densidad TEXT,
            estado_tramo TEXT,
            transitabilidad TEXT,
            necesidad_intervencion TEXT,
            tiempo_estimado_dias REAL,
            fecha_inspeccion TEXT,
            observaciones TEXT,
            ict REAL,
            pct_deterioro REAL,
            prioridad TEXT,
            tipo_mantenimiento TEXT,
            foto_dano TEXT,
            FOREIGN KEY (tramo_id) REFERENCES tramos(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS intervenciones (
            id SERIAL PRIMARY KEY,
            tramo_id INTEGER,
            dano_id INTEGER,
            tipo_mantenimiento TEXT,
            componente TEXT,
            prioridad TEXT,
            actividades TEXT,
            expediente TEXT,
            fecha_programada TEXT,
            fecha_fin TEXT,
            duracion_dias REAL,
            costo_estimado REAL,
            estado TEXT,
            responsable TEXT,
            observaciones TEXT,
            foto_evidencia TEXT,
            FOREIGN KEY (tramo_id) REFERENCES tramos(id),
            FOREIGN KEY (dano_id) REFERENCES danos(id)
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS tramo_fotos (
            id SERIAL PRIMARY KEY,
            tramo_id INTEGER NOT NULL,
            nombre_archivo TEXT,
            imagen TEXT NOT NULL,
            pie_foto TEXT,
            fecha_carga TEXT,
            orden INTEGER,
            FOREIGN KEY (tramo_id) REFERENCES tramos(id)
        )
    """)
    # Migraciones incrementales idempotentes — seguras de re-ejecutar en cada arranque.
    # IF NOT EXISTS evita el patrón try/except propio de SQLite.
    c.execute("ALTER TABLE danos ADD COLUMN IF NOT EXISTS tipo_mantenimiento TEXT")
    for col_def in [
        "altitud_msnm REAL", "estado_drenaje TEXT", "senalizacion TEXT",
        "n_puentes INTEGER", "n_badenes INTEGER", "accesibilidad TEXT",
        "cantera_fuente TEXT", "fuente_lastre TEXT", "ultima_intervencion TEXT",
        "costo_rutinario REAL", "costo_periodico REAL", "gps_lat REAL", "gps_lon REAL",
        "observaciones TEXT",
    ]:
        c.execute(f"ALTER TABLE tramos ADD COLUMN IF NOT EXISTS {col_def}")
    for col_def in ["componente TEXT", "prioridad TEXT", "foto_evidencia TEXT", "fecha_fin TEXT"]:
        c.execute(f"ALTER TABLE intervenciones ADD COLUMN IF NOT EXISTS {col_def}")
    conn.commit()
    conn.close()


try:
    init_db()
    seed_tramos_proyecto()
except Exception as e:
    st.error(
        "⚠️ **No se pudo inicializar la base de datos en la nube.** Verifica que `SUPABASE_DB_URL` "
        f"esté correctamente configurado y que el proyecto de Supabase esté activo.\n\nDetalle técnico: {e}"
    )
    st.stop()

# ──────────────────────────────────────────────────────────────────────────────
# DICCIONARIOS / REGLAS DE CLASIFICACIÓN AUTOMÁTICA
# ──────────────────────────────────────────────────────────────────────────────
TIPOS_SUPERFICIE = ["Asfaltado (AS)", "Afirmado (AF)", "Sin Afirmar (SA)", "Trocha (T)"]
ESTADOS_TRAMO = ["Bueno", "Regular", "Malo", "Muy Malo"]

TIPOS_DANO = [
    "Deformación", "Erosión", "Baches o Huecos", "Encalaminado", "Lodazal",
    "Cruce de Agua", "Hundimiento", "Pérdida de Plataforma", "Deslizamiento",
    "Falla de Drenaje",
]

# Clasificación de falla: Superficial vs Estructural (criterio MTC MCV-2014)
FALLA_SUPERFICIAL = {"Deformación", "Erosión", "Encalaminado", "Lodazal", "Cruce de Agua"}
FALLA_ESTRUCTURAL = {"Baches o Huecos", "Hundimiento", "Pérdida de Plataforma",
                      "Deslizamiento", "Falla de Drenaje"}

# Severidad base por tipo de daño (peso 0-3) usado en el cálculo del ICT
PESO_DANO = {
    "Deformación": 1, "Erosión": 1, "Encalaminado": 1, "Lodazal": 2,
    "Cruce de Agua": 1, "Baches o Huecos": 2, "Hundimiento": 3,
    "Pérdida de Plataforma": 3, "Deslizamiento": 3, "Falla de Drenaje": 2,
}

NIVELES_GRAVEDAD = ["Sin Deterioro", "Leve", "Moderada", "Severa"]
CLASES_DENSIDAD = ["No aplica", "Baja", "Media", "Alta"]

# ──────────────────────────────────────────────────────────────────────────────
# TIPOS DE MANTENIMIENTO VIAL (5 tipos — según informe del proyecto)
# ──────────────────────────────────────────────────────────────────────────────
MANTENIMIENTOS = {
    "Mantenimiento Rutinario": {
        "color": "#2e7d32",
        "descripcion": "Intervención preventiva de baja complejidad, ejecutada de forma periódica y continua sobre toda la vía.",
        "actividades": [
            "Limpieza de cunetas", "Limpieza de alcantarillas", "Roce y desbroce",
            "Eliminación de derrumbes menores", "Perfilado ligero",
        ],
    },
    "Mantenimiento Periódico": {
        "color": "#1565c0",
        "descripcion": "Intervención programada cada 1-3 años para restituir las condiciones estructurales y funcionales de la vía.",
        "actividades": [
            "Reposición de material granular", "Reconformación de plataforma",
            "Perfilado general", "Mejoramiento de drenajes",
        ],
    },
    "Mantenimiento de Emergencia": {
        "color": "#e65100",
        "descripcion": "Intervención inmediata ante eventos imprevistos que interrumpen la transitabilidad de la vía.",
        "actividades": [
            "Atención de derrumbes", "Atención de inundaciones",
            "Restitución de interrupciones de tránsito", "Atención de eventos climáticos extremos",
        ],
    },
    "Rehabilitación": {
        "color": "#8e24aa",
        "descripcion": "Intervención de mayor envergadura para recuperar tramos con deterioro avanzado, sin llegar a reconstrucción total.",
        "actividades": [
            "Recuperación de tramos deteriorados", "Reconstrucción parcial de plataforma",
            "Estabilización de taludes", "Recuperación de drenajes",
        ],
    },
    "Reconstrucción": {
        "color": "#b71c1c",
        "descripcion": "Intervención integral cuando la vía presenta pérdida total de plataforma o daños estructurales severos irreversibles.",
        "actividades": [
            "Reconstrucción integral de la vía", "Reposición por pérdida total de plataforma",
            "Atención de daños estructurales severos",
        ],
    },
}

ESTADOS_INTERVENCION = ["Pendiente", "Programada", "En ejecución", "Ejecutada / Finalizada", "Reprogramada", "Postergada"]
COMPONENTES_VIA = ["Superficie", "Estructura", "Drenaje", "Señalización", "Obras Complementarias"]
PRIORIDADES_INTERVENCION = ["Baja", "Media", "Alta", "Urgente"]

# Expedientes históricos ejecutados — referencia institucional (Municipalidad de Ocongate)
EXPEDIENTES_HISTORICOS = [
    {"Expediente": "ET-MDO-001", "Tramo / Descripción": "Rehabilitación y mejoramiento de trocha carrozable Ocongate – Ccatcca (Km 0+000 al Km 12+500)", "Año": 2021, "Costo total ejecutado (S/.)": 424650.00},
    {"Expediente": "ET-MDO-002", "Tramo / Descripción": "Mantenimiento correctivo trocha carrozable Ccatcca – Marcapata (Km 0+000 al Km 18+200)", "Año": 2022, "Costo total ejecutado (S/.)": 618320.00},
    {"Expediente": "ET-MDO-003", "Tramo / Descripción": "Rehabilitación trocha carrozable Marcapata – Nuñoa (Km 0+000 al Km 22+400)", "Año": 2022, "Costo total ejecutado (S/.)": 761440.00},
    {"Expediente": "ET-MDO-004", "Tramo / Descripción": "Mantenimiento correctivo trocha carrozable Ocongate – Tinki (Km 0+000 al Km 15+600)", "Año": 2023, "Costo total ejecutado (S/.)": 530720.00},
    {"Expediente": "ET-MDO-005", "Tramo / Descripción": "Rehabilitación trocha carrozable Tinki – Ausangate (Km 0+000 al Km 26+800)", "Año": 2024, "Costo total ejecutado (S/.)": 574521.00},
]


def determinar_tipo_mantenimiento(tipo_dano: str, gravedad: str, ict: float, transitabilidad: str) -> str:
    """Determina automáticamente el tipo de mantenimiento (de los 5 definidos) según
    la severidad del daño, el ICT calculado y la transitabilidad resultante."""
    if transitabilidad == "Interrumpida" and gravedad == "Severa" and tipo_dano in {
        "Deslizamiento", "Cruce de Agua", "Lodazal", "Falla de Drenaje", "Hundimiento"
    }:
        return "Mantenimiento de Emergencia"
    if tipo_dano == "Pérdida de Plataforma" and gravedad == "Severa":
        return "Reconstrucción"
    if ict < 30:
        return "Rehabilitación"
    if ict < 60:
        return "Mantenimiento Periódico"
    return "Mantenimiento Rutinario"


def clasificar_falla(tipo_dano: str) -> str:
    if tipo_dano in FALLA_ESTRUCTURAL:
        return "Estructural"
    return "Superficial"


def calcular_gravedad(longitud_afectada: float, longitud_tramo: float, peso_dano: int) -> str:
    """Gravedad en función de % de longitud afectada y severidad propia del daño."""
    if longitud_tramo <= 0:
        pct = 0
    else:
        pct = (longitud_afectada / longitud_tramo) * 100
    score = pct * 0.5 + peso_dano * 10
    if score < 10:
        return "Sin Deterioro"
    elif score < 25:
        return "Leve"
    elif score < 50:
        return "Moderada"
    else:
        return "Severa"


def calcular_estado(gravedad: str) -> str:
    return {
        "Sin Deterioro": "Bueno",
        "Leve": "Regular",
        "Moderada": "Malo",
        "Severa": "Muy Malo",
    }.get(gravedad, "Regular")


def calcular_transitabilidad(estado: str) -> str:
    return {
        "Bueno": "Libre",
        "Regular": "Libre",
        "Malo": "Restringida",
        "Muy Malo": "Interrumpida",
    }.get(estado, "Libre")


def calcular_prioridad(estado: str, transitabilidad: str) -> str:
    if transitabilidad == "Interrumpida":
        return "Urgente"
    if estado == "Malo":
        return "Alta"
    if estado == "Regular":
        return "Media"
    return "Baja"


def calcular_tiempo_reparacion(tipo_dano: str, longitud_afectada: float) -> float:
    """Días estimados de reparación según tipo de daño y longitud afectada (km)."""
    rendimiento_km_dia = {  # rendimiento referencial MCV-2014 (km/día por cuadrilla)
        "Deformación": 0.8, "Erosión": 0.6, "Baches o Huecos": 1.0,
        "Encalaminado": 1.2, "Lodazal": 0.4, "Cruce de Agua": 0.3,
        "Hundimiento": 0.25, "Pérdida de Plataforma": 0.2,
        "Deslizamiento": 0.15, "Falla de Drenaje": 0.35,
    }.get(tipo_dano, 0.5)
    dias = max(1.0, longitud_afectada / rendimiento_km_dia)
    return round(dias, 1)


def calcular_ict(gravedad: str, densidad: str, falla: str) -> float:
    """Índice de Condición de Trocha (0-100, 100 = óptimo)."""
    base = {"Sin Deterioro": 100, "Leve": 80, "Moderada": 55, "Severa": 25}.get(gravedad, 80)
    ajuste_densidad = {"No aplica": 0, "Baja": -2, "Media": -8, "Alta": -15}.get(densidad, 0)
    ajuste_falla = -10 if falla == "Estructural" else -3
    ict = base + ajuste_densidad + ajuste_falla
    return max(0.0, min(100.0, round(ict, 1)))


def calcular_necesidad_intervencion(prioridad: str) -> str:
    return {
        "Urgente": "Intervención inmediata (correctivo de emergencia)",
        "Alta": "Intervención correctiva en el corto plazo (≤ 30 días)",
        "Media": "Mantenimiento periódico programado",
        "Baja": "Mantenimiento rutinario / monitoreo",
    }.get(prioridad, "Monitoreo")


def ict_base_por_estado(estado_actual: str) -> float:
    """Estimación referencial del IEC/ICT actual del tramo a partir de su estado declarado."""
    return {"Bueno": 90, "Regular": 70, "Malo": 45, "Muy Malo": 20}.get(estado_actual, 70)


def anios_desde(fecha_str) -> float:
    if not fecha_str:
        return 0.0
    try:
        fecha = datetime.strptime(str(fecha_str), "%Y-%m-%d").date()
        return max(0.0, (date.today() - fecha).days / 365.25)
    except Exception:
        return 0.0


def curva_iec_deterioro(tramo_dict: dict = None):
    """Construye la figura Plotly de la curva de deterioro del IEC con y sin mantenimiento,
    marcando la posición actual del tramo si se provee."""
    t = list(range(0, 26))
    sin_mant = [max(0, round(100 * (1 - (ti / 23) ** 2.2), 1)) for ti in t]
    rutinario = [round(100 - 0.78 * ti, 1) for ti in t]
    periodico = [round(100 - 0.40 * ti, 1) for ti in t]

    fig = go.Figure()
    fig.add_trace(go.Scatter(x=t, y=sin_mant, mode="lines", name="Sin mantenimiento",
                              line=dict(color="#c62828", dash="dash", width=2.5)))
    fig.add_trace(go.Scatter(x=t, y=rutinario, mode="lines", name="Con mant. rutinario",
                              line=dict(color="#2e7d32", width=2.5)))
    fig.add_trace(go.Scatter(x=t, y=periodico, mode="lines", name="Con mant. periódico",
                              line=dict(color="#1565c0", dash="dot", width=2.5)))

    fig.add_hrect(y0=80, y1=100, fillcolor="#2e7d32", opacity=0.06, line_width=0)
    fig.add_hrect(y0=50, y1=80, fillcolor="#fdd835", opacity=0.08, line_width=0)
    fig.add_hrect(y0=0, y1=50, fillcolor="#c62828", opacity=0.06, line_width=0)
    fig.add_annotation(x=25, y=90, text="Bueno / Muy Bueno", showarrow=False, font=dict(size=10, color="#2e7d32"), xanchor="right")
    fig.add_annotation(x=25, y=65, text="Regular", showarrow=False, font=dict(size=10, color="#a08000"), xanchor="right")
    fig.add_annotation(x=25, y=10, text="Malo / Crítico", showarrow=False, font=dict(size=10, color="#c62828"), xanchor="right")

    if tramo_dict is not None:
        x_actual = round(anios_desde(tramo_dict.get("ultima_intervencion")), 1)
        y_actual = ict_base_por_estado(tramo_dict.get("estado_actual"))
        fig.add_trace(go.Scatter(
            x=[x_actual], y=[y_actual], mode="markers+text", name=f"Tramo {tramo_dict.get('codigo','')} (actual)",
            marker=dict(color="#0b2545", size=14, symbol="star", line=dict(color="white", width=1)),
            text=[f"  {tramo_dict.get('codigo','')}"], textposition="middle right",
        ))

    fig.update_layout(
        xaxis_title="Años desde la última intervención", yaxis_title="IEC (%)",
        yaxis_range=[0, 102], height=380, margin=dict(t=20, b=40, l=10, r=10),
        legend=dict(orientation="h", yanchor="bottom", y=-0.28, xanchor="center", x=0.5),
        plot_bgcolor="white",
    )
    return fig


# ──────────────────────────────────────────────────────────────────────────────
# UTILIDADES DE DATOS
# ──────────────────────────────────────────────────────────────────────────────
def img_to_blob(uploaded_file, carpeta: str = "adjuntos"):
    """Sube el archivo cargado por el usuario a Supabase Storage y devuelve su URL pública.
    Se conserva el nombre `img_to_blob` por compatibilidad con el resto del código."""
    if uploaded_file is None:
        return None
    return subir_archivo_storage(uploaded_file.getvalue(), uploaded_file.name, carpeta)


def blob_to_img(valor):
    """Devuelve un BytesIO listo para st.image()/incrustar en PDF, a partir de:
    - bytes crudos (vista previa antes de guardar), o
    - una URL de Supabase Storage (después de guardado, se descarga con caché)."""
    if valor is None:
        return None
    if isinstance(valor, (bytes, bytearray)):
        return BytesIO(valor)
    contenido = descargar_bytes_storage(valor)
    return BytesIO(contenido) if contenido else None


def df_tramos() -> pd.DataFrame:
    conn = get_conn()
    try:
        return pd.read_sql_query("""
            SELECT id, codigo, nombre, comunidad, distrito, provincia, departamento,
                   longitud_km, ancho_m, tipo_superficie, estado_actual, responsable,
                   telefono_responsable, fecha_registro, altitud_msnm, estado_drenaje,
                   senalizacion, n_puentes, n_badenes, accesibilidad, cantera_fuente,
                   fuente_lastre, ultima_intervencion, costo_rutinario, costo_periodico,
                   gps_lat, gps_lon, observaciones
            FROM tramos ORDER BY id DESC
        """, conn)
    finally:
        conn.close()


def df_danos() -> pd.DataFrame:
    conn = get_conn()
    try:
        return pd.read_sql_query("""
            SELECT d.*, t.codigo AS codigo_tramo, t.nombre AS nombre_tramo,
                   t.longitud_km AS longitud_tramo_km
            FROM danos d
            LEFT JOIN tramos t ON d.tramo_id = t.id
            ORDER BY d.id DESC
        """, conn)
    finally:
        conn.close()


def insertar_tramo(data: dict) -> int:
    conn = get_conn()
    try:
        c = conn.cursor()

        # El PDF adjunto (si existe) se sube a Supabase Storage; en la BD solo se guarda la URL.
        adjunto_url = None
        if data.get("adjunto_pdf"):
            adjunto_url = subir_archivo_storage(
                data["adjunto_pdf"], data.get("adjunto_pdf_nombre") or "adjunto.pdf", "adjuntos-pdf"
            )

        c.execute("""
            INSERT INTO tramos (codigo, nombre, comunidad, distrito, provincia, departamento,
                longitud_km, ancho_m, tipo_superficie, estado_actual, responsable,
                telefono_responsable, fecha_registro, adjunto_pdf, adjunto_pdf_nombre,
                altitud_msnm, estado_drenaje, senalizacion, n_puentes, n_badenes,
                accesibilidad, cantera_fuente, fuente_lastre, ultima_intervencion,
                costo_rutinario, costo_periodico, gps_lat, gps_lon, observaciones)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            RETURNING id
        """, (
            data["codigo"], data["nombre"], data["comunidad"], data["distrito"],
            data["provincia"], data["departamento"], data["longitud_km"], data["ancho_m"],
            data["tipo_superficie"], data["estado_actual"], data["responsable"],
            data["telefono_responsable"], data["fecha_registro"],
            adjunto_url, data["adjunto_pdf_nombre"],
            data.get("altitud_msnm"), data.get("estado_drenaje"), data.get("senalizacion"),
            data.get("n_puentes"), data.get("n_badenes"), data.get("accesibilidad"),
            data.get("cantera_fuente"), data.get("fuente_lastre"), data.get("ultima_intervencion"),
            data.get("costo_rutinario"), data.get("costo_periodico"), data.get("gps_lat"), data.get("gps_lon"),
            data.get("observaciones"),
        ))
        new_id = c.fetchone()[0]
        conn.commit()
        return new_id
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ──────────────────────────────────────────────────────────────────────────────
# FOTOGRAFÍAS DEL MANTENIMIENTO (tabla tramo_fotos — relación 1 a N con tramos)
# ──────────────────────────────────────────────────────────────────────────────
MAX_FOTOS_TRAMO = 5


def guardar_fotos_tramo(tramo_id: int, fotos: list):
    """Sube a Supabase Storage e inserta en tramo_fotos una o varias fotografías asociadas a un tramo.

    `fotos` es una lista de dicts con las llaves: nombre_archivo, imagen (bytes crudos),
    pie_foto (opcional) y fecha_carga (opcional, YYYY-MM-DD).
    """
    if not fotos:
        return
    conn = get_conn()
    try:
        c = conn.cursor()
        orden_actual = c.execute(
            "SELECT COALESCE(MAX(orden), 0) FROM tramo_fotos WHERE tramo_id=?", (tramo_id,)
        ).fetchone()[0]
        for i, foto in enumerate(fotos[:MAX_FOTOS_TRAMO], start=1):
            url_imagen = subir_archivo_storage(
                foto["imagen"], foto.get("nombre_archivo") or f"foto_{i}.jpg", "fotos-tramos"
            )
            if not url_imagen:
                continue  # Si la subida a Storage falla, se omite esa foto sin interrumpir el resto.
            c.execute(
                "INSERT INTO tramo_fotos (tramo_id, nombre_archivo, imagen, pie_foto, fecha_carga, orden) "
                "VALUES (?,?,?,?,?,?)",
                (
                    tramo_id, foto.get("nombre_archivo"), url_imagen,
                    foto.get("pie_foto") or None,
                    foto.get("fecha_carga") or date.today().strftime("%Y-%m-%d"),
                    orden_actual + i,
                ),
            )
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def obtener_fotos_tramo(tramo_id: int) -> list:
    """Devuelve la lista de fotografías de un tramo (con su URL de Storage), ordenadas."""
    conn = get_conn()
    try:
        filas = conn.execute(
            "SELECT id, nombre_archivo, imagen, pie_foto, fecha_carga FROM tramo_fotos "
            "WHERE tramo_id=? ORDER BY orden ASC, id ASC", (tramo_id,)
        ).fetchall()
        return [
            {"id": f[0], "nombre_archivo": f[1], "imagen": f[2], "pie_foto": f[3], "fecha_carga": f[4]}
            for f in filas
        ]
    finally:
        conn.close()


def eliminar_foto_tramo(foto_id: int):
    """Elimina el registro en BD y, si es posible, el archivo físico en Supabase Storage."""
    conn = get_conn()
    try:
        fila = conn.execute("SELECT imagen FROM tramo_fotos WHERE id=?", (foto_id,)).fetchone()
        conn.execute("DELETE FROM tramo_fotos WHERE id=?", (foto_id,))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()
    if fila and fila[0]:
        eliminar_archivo_storage(fila[0])


def actualizar_tramo(tramo_id: int, data: dict):
    conn = get_conn()
    try:
        c = conn.cursor()
        c.execute("""
            UPDATE tramos SET codigo=?, nombre=?, comunidad=?, distrito=?, provincia=?,
                departamento=?, longitud_km=?, ancho_m=?, tipo_superficie=?, estado_actual=?,
                responsable=?, telefono_responsable=?
            WHERE id=?
        """, (
            data["codigo"], data["nombre"], data["comunidad"], data["distrito"],
            data["provincia"], data["departamento"], data["longitud_km"], data["ancho_m"],
            data["tipo_superficie"], data["estado_actual"], data["responsable"],
            data["telefono_responsable"], tramo_id,
        ))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def eliminar_tramo(tramo_id: int):
    conn = get_conn()
    try:
        c = conn.cursor()
        # Recolecta URLs de Storage asociadas antes de borrar los registros, para limpiarlas después.
        fotos_urls = [f[0] for f in c.execute(
            "SELECT imagen FROM tramo_fotos WHERE tramo_id=?", (tramo_id,)
        ).fetchall()]
        fila_adjunto = c.execute("SELECT adjunto_pdf FROM tramos WHERE id=?", (tramo_id,)).fetchone()

        c.execute("DELETE FROM tramo_fotos WHERE tramo_id=?", (tramo_id,))
        c.execute("DELETE FROM intervenciones WHERE tramo_id=?", (tramo_id,))
        c.execute("DELETE FROM danos WHERE tramo_id=?", (tramo_id,))
        c.execute("DELETE FROM tramos WHERE id=?", (tramo_id,))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    for url in fotos_urls:
        eliminar_archivo_storage(url)
    if fila_adjunto and fila_adjunto[0]:
        eliminar_archivo_storage(fila_adjunto[0])


def insertar_dano(data: dict) -> int:
    conn = get_conn()
    try:
        c = conn.cursor()
        c.execute("""
            INSERT INTO danos (tramo_id, progresiva_inicial, progresiva_final, longitud_afectada,
                tipo_dano, tipo_falla, nivel_gravedad, clase_densidad, estado_tramo,
                transitabilidad, necesidad_intervencion, tiempo_estimado_dias, fecha_inspeccion,
                observaciones, ict, pct_deterioro, prioridad, tipo_mantenimiento, foto_dano)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
            RETURNING id
        """, (
            data["tramo_id"], data["progresiva_inicial"], data["progresiva_final"],
            data["longitud_afectada"], data["tipo_dano"], data["tipo_falla"],
            data["nivel_gravedad"], data["clase_densidad"], data["estado_tramo"],
            data["transitabilidad"], data["necesidad_intervencion"], data["tiempo_estimado_dias"],
            data["fecha_inspeccion"], data["observaciones"], data["ict"], data["pct_deterioro"],
            data["prioridad"], data["tipo_mantenimiento"], data["foto_dano"],
        ))
        new_id = c.fetchone()[0]
        conn.commit()
        return new_id
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


# ──────────────────────────────────────────────────────────────────────────────
# MOTOR DE PLANIFICACIÓN INTELIGENTE DE MANTENIMIENTO
# ──────────────────────────────────────────────────────────────────────────────
# Meses de temporada de lluvias en la sierra sur del Perú (Cusco / Ocongate, >4000 msnm)
MESES_LLUVIA = {12, 1, 2, 3}

# Frecuencia recomendada (días) desde la última intervención, y duración/costo referencial por tipo
REGLAS_MANTENIMIENTO = {
    "Mantenimiento Rutinario": {"frecuencia_dias": 90, "duracion_dias": 7, "sensible_lluvia": True, "lead_time_dias": 7},
    "Mantenimiento Periódico": {"frecuencia_dias": 545, "duracion_dias": 30, "sensible_lluvia": True, "lead_time_dias": 30},
    "Mantenimiento de Emergencia": {"frecuencia_dias": 0, "duracion_dias": 5, "sensible_lluvia": False, "lead_time_dias": 0},
    "Rehabilitación": {"frecuencia_dias": 1095, "duracion_dias": 60, "sensible_lluvia": True, "lead_time_dias": 45},
    "Reconstrucción": {"frecuencia_dias": 1825, "duracion_dias": 120, "sensible_lluvia": True, "lead_time_dias": 90},
}


def es_temporada_lluvias(fecha) -> bool:
    return fecha.month in MESES_LLUVIA


def proxima_fecha_seca(fecha):
    """Devuelve la fecha del 15 de abril más próxima (inicio de temporada seca) a partir de `fecha`."""
    anio = fecha.year if fecha.month <= 4 else fecha.year + 1
    candidata = date(anio, 4, 15)
    if candidata < fecha:
        candidata = date(anio + 1, 4, 15)
    return candidata


def sugerir_programacion(tipo_mantenimiento: str, componente: str, tramo_dict: dict, prioridad: str) -> dict:
    """Sugiere fecha tentativa de inicio/fin, duración y costo referencial, con advertencias
    según temporada de lluvias, frecuencia recomendada y criticidad del tramo."""
    regla = REGLAS_MANTENIMIENTO.get(tipo_mantenimiento, REGLAS_MANTENIMIENTO["Mantenimiento Rutinario"])
    hoy = date.today()
    advertencias = []

    # 1) Emergencia → siempre inmediato, sin restricción de temporada
    if tipo_mantenimiento == "Mantenimiento de Emergencia" or prioridad == "Urgente":
        fecha_inicio = hoy
        advertencias.append("⚡ Intervención de emergencia: se sugiere ejecución inmediata, sin restricción por temporada.")
    else:
        # 2) Fecha base según frecuencia recomendada desde la última intervención
        ultima = tramo_dict.get("ultima_intervencion")
        if ultima:
            try:
                fecha_ultima = datetime.strptime(str(ultima), "%Y-%m-%d").date()
            except Exception:
                fecha_ultima = hoy
            fecha_base = fecha_ultima + pd.Timedelta(days=regla["frecuencia_dias"])
            fecha_base = fecha_base if isinstance(fecha_base, date) else fecha_base.date()
        else:
            fecha_base = hoy + pd.Timedelta(days=regla["lead_time_dias"])
            fecha_base = fecha_base.date() if hasattr(fecha_base, "date") else fecha_base

        fecha_inicio = max(fecha_base, hoy)

        # 3) Evitar temporada de lluvias para componentes/tipos sensibles
        if regla["sensible_lluvia"] and componente in {"Superficie", "Drenaje", "Estructura"} and es_temporada_lluvias(fecha_inicio):
            fecha_original = fecha_inicio
            fecha_inicio = proxima_fecha_seca(fecha_inicio)
            advertencias.append(
                f"🌧️ La fecha calculada ({fecha_original.strftime('%d/%m/%Y')}) cae en temporada de lluvias "
                f"(dic-mar). Se reprogramó al inicio de temporada seca: {fecha_inicio.strftime('%d/%m/%Y')}."
            )

    duracion = regla["duracion_dias"]
    fecha_fin = fecha_inicio + pd.Timedelta(days=duracion)
    fecha_fin = fecha_fin.date() if hasattr(fecha_fin, "date") else fecha_fin

    # 4) Alerta de criticidad del tramo
    if tramo_dict.get("estado_actual") in {"Malo", "Muy Malo"} or prioridad in {"Alta", "Urgente"}:
        advertencias.append(
            f"⚠️ Tramo con criticidad {('alta' if prioridad in {'Alta','Urgente'} else 'por estado ' + tramo_dict.get('estado_actual',''))}: "
            f"se recomienda no postergar esta intervención."
        )

    # 5) Costo referencial según campos del tramo, si existen
    costo_ref = None
    if tipo_mantenimiento == "Mantenimiento Rutinario" and tramo_dict.get("costo_rutinario"):
        costo_ref = tramo_dict["costo_rutinario"]
    elif tipo_mantenimiento == "Mantenimiento Periódico" and tramo_dict.get("costo_periodico"):
        costo_ref = tramo_dict["costo_periodico"]

    return {
        "fecha_inicio": fecha_inicio, "fecha_fin": fecha_fin, "duracion_dias": duracion,
        "costo_referencial": costo_ref, "advertencias": advertencias,
        "frecuencia_dias": regla["frecuencia_dias"],
    }


def determinar_componente(tipo_dano: str) -> str:
    """Determina el componente vial afectado a partir del tipo de daño registrado."""
    if tipo_dano in {"Falla de Drenaje", "Cruce de Agua", "Lodazal"}:
        return "Drenaje"
    if tipo_dano in {"Pérdida de Plataforma", "Hundimiento", "Deslizamiento"}:
        return "Estructura"
    return "Superficie"


def insertar_intervencion(data: dict):
    conn = get_conn()
    try:
        c = conn.cursor()
        c.execute("""
            INSERT INTO intervenciones (tramo_id, dano_id, tipo_mantenimiento, componente, prioridad,
                actividades, expediente, fecha_programada, fecha_fin, duracion_dias, costo_estimado, estado,
                responsable, observaciones, foto_evidencia)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            data["tramo_id"], data.get("dano_id"), data["tipo_mantenimiento"], data.get("componente"),
            data.get("prioridad"), data["actividades"], data["expediente"], data["fecha_programada"],
            data.get("fecha_fin"), data["duracion_dias"], data["costo_estimado"], data["estado"],
            data["responsable"], data["observaciones"], data.get("foto_evidencia"),
        ))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def actualizar_estado_intervencion(intervencion_id: int, nuevo_estado: str):
    conn = get_conn()
    try:
        conn.execute("UPDATE intervenciones SET estado=? WHERE id=?", (nuevo_estado, intervencion_id))
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def df_intervenciones() -> pd.DataFrame:
    conn = get_conn()
    try:
        return pd.read_sql_query("""
            SELECT i.*, t.codigo AS codigo_tramo, t.nombre AS nombre_tramo, t.comunidad AS comunidad_tramo,
                   t.distrito AS distrito_tramo, t.longitud_km AS longitud_tramo_km,
                   t.fecha_registro AS fecha_registro_tramo, t.ultima_intervencion AS ultima_intervencion_tramo
            FROM intervenciones i
            LEFT JOIN tramos t ON i.tramo_id = t.id
            ORDER BY i.fecha_programada ASC
        """, conn)
    finally:
        conn.close()


# ──────────────────────────────────────────────────────────────────────────────
# SIEMBRA INICIAL — Una intervención de ejemplo por cada uno de los 5 tipos
# de mantenimiento, para que el panel de Programación nunca se vea vacío.
# ──────────────────────────────────────────────────────────────────────────────
INTERVENCIONES_PROYECTO = [
    {
        "codigo_tramo": "M5", "tipo_mantenimiento": "Mantenimiento Rutinario", "componente": "Superficie",
        "prioridad": "Media", "actividades": "Limpieza de cunetas; Roce y desbroce; Perfilado ligero",
        "expediente": "ET-MDO-RUT-01", "fecha_programada": "2026-06-15", "fecha_fin": "2026-06-22",
        "duracion_dias": 7, "costo_estimado": 18500.0, "estado": "En ejecución",
        "responsable": "Subgerencia de Gestión de Riesgos y Mantenimiento",
        "observaciones": "Mantenimiento preventivo programado dentro del ciclo rutinario trimestral del tramo.",
    },
    {
        "codigo_tramo": "M2", "tipo_mantenimiento": "Mantenimiento Periódico", "componente": "Superficie",
        "prioridad": "Media", "actividades": "Reposición de material granular; Reconformación de plataforma; Perfilado general",
        "expediente": "ET-MDO-PER-01", "fecha_programada": "2026-07-15", "fecha_fin": "2026-08-14",
        "duracion_dias": 30, "costo_estimado": 95000.0, "estado": "Programada",
        "responsable": "Subgerencia de Gestión de Riesgos y Mantenimiento",
        "observaciones": "Programado al inicio de temporada seca para restituir condición estructural de la plataforma.",
    },
    {
        "codigo_tramo": "M4", "tipo_mantenimiento": "Mantenimiento de Emergencia", "componente": "Drenaje",
        "prioridad": "Urgente", "actividades": "Atención de derrumbes; Restitución de interrupciones de tránsito",
        "expediente": "ET-MDO-EMG-01", "fecha_programada": "2026-06-10", "fecha_fin": "2026-06-15",
        "duracion_dias": 5, "costo_estimado": 32000.0, "estado": "Ejecutada / Finalizada",
        "responsable": "Subgerencia de Gestión de Riesgos y Mantenimiento",
        "observaciones": "Atención inmediata por obstrucción de cuneta tras evento de lluvia intensa.",
    },
    {
        "codigo_tramo": "M1", "tipo_mantenimiento": "Rehabilitación", "componente": "Estructura",
        "prioridad": "Alta", "actividades": "Recuperación de tramos deteriorados; Estabilización de taludes; Recuperación de drenajes",
        "expediente": "ET-MDO-REH-01", "fecha_programada": "2026-08-20", "fecha_fin": "2026-10-19",
        "duracion_dias": 60, "costo_estimado": 420000.0, "estado": "Pendiente",
        "responsable": "Subgerencia de Gestión de Riesgos y Mantenimiento",
        "observaciones": "Tramo con deterioro avanzado (estado Malo); requiere recuperación estructural antes de la próxima temporada de lluvias.",
    },
    {
        "codigo_tramo": "M3", "tipo_mantenimiento": "Reconstrucción", "componente": "Estructura",
        "prioridad": "Urgente", "actividades": "Reconstrucción integral de la vía; Reposición por pérdida total de plataforma",
        "expediente": "ET-MDO-REC-01", "fecha_programada": "2027-04-15", "fecha_fin": "2027-07-14",
        "duracion_dias": 90, "costo_estimado": 980000.0, "estado": "Pendiente",
        "responsable": "Subgerencia de Gestión de Riesgos y Mantenimiento",
        "observaciones": "Tramo en estado Muy Malo con pérdida estructural severa; requiere expediente técnico de reconstrucción integral.",
    },
]


def seed_intervenciones_proyecto():
    """Inserta una intervención de ejemplo por cada tipo de mantenimiento (5 en total)
    si la tabla de intervenciones está vacía, para que el panel nunca se vea vacío."""
    conn = get_conn()
    c = conn.cursor()
    total = c.execute("SELECT COUNT(*) FROM intervenciones").fetchone()[0]
    conn.close()
    if total > 0:
        return
    for item in INTERVENCIONES_PROYECTO:
        fila_tramo = c_tramo = None
        conn2 = get_conn()
        fila_tramo = conn2.execute("SELECT id FROM tramos WHERE codigo=?", (item["codigo_tramo"],)).fetchone()
        conn2.close()
        if not fila_tramo:
            continue
        insertar_intervencion({
            "tramo_id": fila_tramo[0], "dano_id": None, "tipo_mantenimiento": item["tipo_mantenimiento"],
            "componente": item["componente"], "prioridad": item["prioridad"], "actividades": item["actividades"],
            "expediente": item["expediente"], "fecha_programada": item["fecha_programada"],
            "fecha_fin": item["fecha_fin"], "duracion_dias": item["duracion_dias"],
            "costo_estimado": item["costo_estimado"], "estado": item["estado"],
            "responsable": item["responsable"], "observaciones": item["observaciones"], "foto_evidencia": None,
        })


try:
    seed_intervenciones_proyecto()
except Exception as e:
    st.warning(f"No se pudieron sembrar las intervenciones de ejemplo (no es crítico): {e}")


# ──────────────────────────────────────────────────────────────────────────────
# EXPORTACIÓN
# ──────────────────────────────────────────────────────────────────────────────
def exportar_excel(dfs: dict) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for sheet_name, df in dfs.items():
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)
    return output.getvalue()


def _pdf_safe(texto) -> str:
    """Convierte cualquier texto a algo seguro para las fuentes core de FPDF
    (latin-1), sustituyendo caracteres no soportados (guion largo, tildes raras, etc.)."""
    if texto is None:
        return ""
    s = str(texto)
    reemplazos = {
        "–": "-", "—": "-", "“": '"', "”": '"', "‘": "'", "’": "'",
        "…": "...", "•": "-", "´": "'", "`": "'",
    }
    for k, v in reemplazos.items():
        s = s.replace(k, v)
    return s.encode("latin-1", "replace").decode("latin-1")


def exportar_pdf_resumen(titulo: str, df_resumen: pd.DataFrame, texto_intro: str = "") -> bytes:
    try:
        from fpdf import FPDF
    except ImportError:
        return None
    pdf = FPDF()
    pdf.add_page()
    epw = pdf.w - 2 * pdf.l_margin  # ancho útil de página
    pdf.set_font("Helvetica", "B", 14)
    pdf.multi_cell(epw, 8, _pdf_safe(titulo))
    pdf.set_x(pdf.l_margin)
    pdf.set_font("Helvetica", "", 9)
    if texto_intro:
        pdf.multi_cell(epw, 5, _pdf_safe(texto_intro))
        pdf.set_x(pdf.l_margin)
    pdf.ln(2)
    pdf.set_font("Helvetica", "B", 8)
    cols = list(df_resumen.columns)
    col_w = max(20, int(190 / max(1, len(cols))))
    for col in cols:
        pdf.cell(col_w, 6, _pdf_safe(col)[:18], border=1)
    pdf.ln()
    pdf.set_font("Helvetica", "", 7)
    for _, row in df_resumen.iterrows():
        for col in cols:
            pdf.cell(col_w, 6, _pdf_safe(row[col])[:18], border=1)
        pdf.ln()
    return bytes(pdf.output(dest="S"))


def download_button_bytes(data: bytes, filename: str, label: str, mime: str, key: str):
    if data is None:
        st.warning("Para exportar a PDF instala la librería `fpdf2` (pip install fpdf2).")
        return
    st.download_button(label, data=data, file_name=filename, mime=mime, key=key)


# ──────────────────────────────────────────────────────────────────────────────
# FICHAS TÉCNICAS — Formato tipo Provías Descentralizado (1-B Itinerario / 1-D Daños)
# ──────────────────────────────────────────────────────────────────────────────
def generar_grid_progresiva(longitud_km: float, paso: float = 0.5) -> list:
    """Genera los tramos de progresiva (Del Km / Al Km) cada `paso` km."""
    tramos = []
    actual = 0.0
    longitud_km = max(longitud_km, paso)
    while actual < longitud_km:
        siguiente = min(round(actual + paso, 3), longitud_km)
        tramos.append((actual, siguiente))
        actual = siguiente
    return tramos


def df_ficha_itinerario(tramo_dict: dict) -> pd.DataFrame:
    """Construye la tabla 1-B (Ficha del Itinerario del Camino Vecinal) para un tramo,
    incluyendo coordenadas UTM/GPS y obras de arte en el primer registro (punto de referencia)."""
    grid = generar_grid_progresiva(tramo_dict["longitud_km"])
    obras_arte = []
    if tramo_dict.get("n_puentes"):
        obras_arte.append(f"{tramo_dict['n_puentes']} puente(s)")
    if tramo_dict.get("n_badenes"):
        obras_arte.append(f"{tramo_dict['n_badenes']} badén(es)")
    obras_arte_txt = ", ".join(obras_arte) if obras_arte else ""

    filas = []
    for idx, (del_km, al_km) in enumerate(grid):
        filas.append({
            "Del Km": f"{del_km:.3f}", "Al Km": f"{al_km:.3f}",
            "Tipo de Superficie": tramo_dict["tipo_superficie"],
            "Estado de Transitabilidad": tramo_dict["estado_actual"],
            "Ancho de Plataforma (m)": tramo_dict["ancho_m"],
            "Norte (WGS84)": tramo_dict.get("gps_lat") if idx == 0 else "",
            "Este (WGS84)": tramo_dict.get("gps_lon") if idx == 0 else "",
            "Altitud (msnm)": tramo_dict.get("altitud_msnm") if idx == 0 else "",
            "Obras Arte / Drenaje / Señalización": obras_arte_txt if idx == 0 else "",
            "Fotos N°": "1" if idx == 0 else "",
        })
    return pd.DataFrame(filas)


def df_ficha_danos(danos_tramo: pd.DataFrame) -> pd.DataFrame:
    """Construye la tabla 1-D (Ficha Técnica de Daños en Camino Vecinal) con datos reales registrados."""
    if danos_tramo.empty:
        return pd.DataFrame(columns=["Del Km", "Al Km", "Longitud (Km)", "Tipo de Daño",
                                      "Nivel de Gravedad", "Clase de Densidad", "Fecha"])
    filas = []
    for r in danos_tramo.itertuples():
        filas.append({
            "Del Km": f"{r.progresiva_inicial:.3f}", "Al Km": f"{r.progresiva_final:.3f}",
            "Longitud (Km)": f"{r.longitud_afectada:.3f}", "Tipo de Daño": r.tipo_dano,
            "Nivel de Gravedad": r.nivel_gravedad, "Clase de Densidad": r.clase_densidad,
            "Fecha": r.fecha_inspeccion,
        })
    return pd.DataFrame(filas)


def _primera_foto_disponible(fotos: list):
    """Devuelve (bytes, etiqueta) de la primera fotografía disponible, descargando desde
    Supabase Storage si el valor almacenado es una URL (y no bytes crudos de vista previa)."""
    if not fotos:
        return None, None
    primera = fotos[0]
    etiqueta = primera.get("pie_foto") or (primera.get("nombre_archivo") or "Fotografía del mantenimiento")
    valor = primera["imagen"]
    contenido = valor if isinstance(valor, (bytes, bytearray)) else descargar_bytes_storage(valor)
    return contenido, etiqueta


def exportar_ficha_excel(tramo_dict: dict, danos_tramo: pd.DataFrame, fotos: list = None) -> bytes:
    """Genera el Excel de la ficha técnica completa: Datos Generales (con foto), 1-B Itinerario, 1-D Daños,
    y una hoja de Evidencia Fotográfica del Mantenimiento con hasta 5 fotos en cuadrícula."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.utils import get_column_letter

    fotos = fotos or []
    wb = Workbook()
    azul = "0B2545"
    dorado = "C9A227"
    gris = "F4F6F9"
    borde = Border(*[Side(style="thin", color="B0B8C4")] * 4)

    def estilo_header(ws, fila, col_ini, col_fin, texto, color_fondo=azul, color_texto="FFFFFF", alto=22):
        ws.merge_cells(start_row=fila, start_column=col_ini, end_row=fila, end_column=col_fin)
        celda = ws.cell(row=fila, column=col_ini, value=texto)
        celda.font = Font(name="Calibri", size=12, bold=True, color=color_texto)
        celda.fill = PatternFill("solid", fgColor=color_fondo)
        celda.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[fila].height = alto

    # ---- HOJA 1: Datos Generales + Foto ----
    ws0 = wb.active
    ws0.title = "Datos Generales"
    for col, w in zip("ABCDE", [22, 30, 4, 28, 30]):
        ws0.column_dimensions[col].width = w
    estilo_header(ws0, 1, 1, 5, "FICHA TÉCNICA DEL TRAMO VIAL", alto=28)
    estilo_header(ws0, 2, 1, 5, "Municipalidad Distrital de Ocongate — Subgerencia de Gestión de Riesgos y Mantenimiento",
                  color_fondo=dorado, color_texto="0B2545", alto=22)

    campos = [
        ("Código de tramo", tramo_dict["codigo"]), ("Nombre del tramo", tramo_dict["nombre"]),
        ("Comunidad / Sector", tramo_dict["comunidad"]),
        ("Distrito / Provincia / Departamento", f"{tramo_dict['distrito']} / {tramo_dict['provincia']} / {tramo_dict['departamento']}"),
        ("Longitud (km)", tramo_dict["longitud_km"]), ("Ancho de plataforma (m)", tramo_dict["ancho_m"]),
        ("Tipo de superficie", tramo_dict["tipo_superficie"]), ("Estado actual", tramo_dict["estado_actual"]),
        ("Responsable de registro", tramo_dict["responsable"]),
        ("Teléfono del responsable", tramo_dict["telefono_responsable"] or "—"),
        ("Fecha de registro", tramo_dict["fecha_registro"]),
    ]
    fila = 4
    for etiqueta, valor in campos:
        c1 = ws0.cell(row=fila, column=1, value=etiqueta)
        c1.font = Font(name="Calibri", size=10, bold=True, color=azul)
        c1.fill = PatternFill("solid", fgColor=gris)
        c1.border = borde
        ws0.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=3)
        c2 = ws0.cell(row=fila, column=2, value=valor)
        c2.font = Font(name="Calibri", size=10)
        c2.border = borde
        ws0.row_dimensions[fila].height = 20
        fila += 1

    foto_bytes, foto_label = _primera_foto_disponible(fotos)
    if foto_bytes:
        try:
            img_buf = BytesIO(foto_bytes)
            xl_img = XLImage(img_buf)
            xl_img.width, xl_img.height = 360, 260
            ws0.add_image(xl_img, "D4")
            ws0.cell(row=fila + 1, column=4, value=f"📷 {foto_label}").font = Font(italic=True, size=9, color="555555")
            if len(fotos) > 1:
                ws0.cell(row=fila + 2, column=4,
                         value=f"Ver todas las fotografías ({len(fotos)}) en la hoja 'Evidencia Fotográfica'."
                         ).font = Font(italic=True, size=8, color="777777")
        except Exception:
            pass

    # ---- HOJA 2: Ficha 1-B Itinerario ----
    ws1 = wb.create_sheet("1-B Itinerario")
    df_b = df_ficha_itinerario(tramo_dict)
    estilo_header(ws1, 1, 1, len(df_b.columns), f"1-B: FICHA DEL ITINERARIO DEL CAMINO VECINAL — Tramo {tramo_dict['codigo']}", alto=26)
    for j, col_name in enumerate(df_b.columns, start=1):
        c = ws1.cell(row=2, column=j, value=col_name)
        c.font = Font(bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=azul)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = borde
        ws1.column_dimensions[get_column_letter(j)].width = 16
    for i, row in df_b.iterrows():
        for j, val in enumerate(row, start=1):
            c = ws1.cell(row=i + 3, column=j, value=val)
            c.border = borde
            c.font = Font(size=9)
            c.alignment = Alignment(horizontal="center")
            c.fill = PatternFill("solid", fgColor="FFFFFF" if i % 2 == 0 else gris)
    ws1.freeze_panes = "A3"

    # ---- HOJA 3: Ficha 1-D Daños ----
    ws2 = wb.create_sheet("1-D Daños")
    df_d = df_ficha_danos(danos_tramo)
    estilo_header(ws2, 1, 1, max(len(df_d.columns), 1), f"1-D: FICHA TÉCNICA DE DAÑOS EN CAMINO VECINAL — Tramo {tramo_dict['codigo']}", alto=26)
    if not df_d.empty:
        for j, col_name in enumerate(df_d.columns, start=1):
            c = ws2.cell(row=2, column=j, value=col_name)
            c.font = Font(bold=True, size=9, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=azul)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = borde
            ws2.column_dimensions[get_column_letter(j)].width = 18
        for i, row in df_d.iterrows():
            for j, val in enumerate(row, start=1):
                c = ws2.cell(row=i + 3, column=j, value=val)
                c.border = borde
                c.font = Font(size=9)
                c.alignment = Alignment(horizontal="center")
                c.fill = PatternFill("solid", fgColor="FFFFFF" if i % 2 == 0 else gris)
    else:
        ws2.cell(row=3, column=1, value="Sin fichas de daño registradas para este tramo.").font = Font(italic=True, size=10)

    # ---- HOJA 4: Evidencia Fotográfica del Mantenimiento (cuadrícula 2 columnas) ----
    if fotos:
        ws3 = wb.create_sheet("Evidencia Fotográfica")
        for col, w in zip("ABCD", [4, 42, 4, 42]):
            ws3.column_dimensions[col].width = w
        estilo_header(ws3, 1, 1, 4, f"EVIDENCIA FOTOGRÁFICA DEL MANTENIMIENTO — Tramo {tramo_dict['codigo']}", alto=26)

        fila_grid = 3
        FOTOS_POR_FILA = 2
        FILAS_POR_FOTO = 16  # alto reservado (en filas) para imagen + pie de foto
        for idx, foto in enumerate(fotos[:MAX_FOTOS_TRAMO]):
            col_base = 2 if (idx % FOTOS_POR_FILA == 0) else (2 + 2 * (idx % FOTOS_POR_FILA))
            fila_base = fila_grid + (idx // FOTOS_POR_FILA) * FILAS_POR_FOTO
            try:
                contenido_foto = foto["imagen"] if isinstance(foto["imagen"], (bytes, bytearray)) \
                    else descargar_bytes_storage(foto["imagen"])
                if not contenido_foto:
                    continue
                img_buf = BytesIO(contenido_foto)
                xl_img = XLImage(img_buf)
                xl_img.width, xl_img.height = 300, 220
                ws3.add_image(xl_img, f"{get_column_letter(col_base)}{fila_base}")
                pie = foto.get("pie_foto") or foto.get("nombre_archivo") or f"Fotografía {idx + 1}"
                fecha_c = foto.get("fecha_carga") or ""
                cap = ws3.cell(row=fila_base + 14, column=col_base,
                                value=f"📷 {pie}" + (f"  ·  {fecha_c}" if fecha_c else ""))
                cap.font = Font(italic=True, size=9, color="555555")
            except Exception:
                continue

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def exportar_ficha_pdf(tramo_dict: dict, danos_tramo: pd.DataFrame, fotos: list = None) -> bytes:
    """Genera el PDF de la ficha técnica completa con evidencia fotográfica del mantenimiento."""
    try:
        from fpdf import FPDF
    except ImportError:
        return None

    fotos = fotos or []
    pdf = FPDF()
    epw = pdf.w - 2 * pdf.l_margin

    # ---- Página 1: Datos generales + foto ----
    pdf.add_page()
    pdf.set_fill_color(11, 37, 69)
    pdf.rect(0, 0, pdf.w, 22, "F")
    pdf.set_xy(10, 6)
    pdf.set_text_color(255, 255, 255)
    pdf.set_font("Helvetica", "B", 13)
    pdf.cell(0, 10, _pdf_safe(f"FICHA TÉCNICA DEL TRAMO {tramo_dict['codigo']}"))
    pdf.set_text_color(0, 0, 0)
    pdf.set_xy(10, 26)
    pdf.set_font("Helvetica", "", 9)
    pdf.multi_cell(epw, 5, _pdf_safe("Municipalidad Distrital de Ocongate — Subgerencia de Gestión de Riesgos y Mantenimiento"))
    pdf.set_x(pdf.l_margin)
    pdf.ln(2)

    pdf.set_font("Helvetica", "B", 10)
    pdf.cell(epw, 7, _pdf_safe("DATOS GENERALES"), border="B")
    pdf.ln(9)
    pdf.set_font("Helvetica", "", 9)
    campos = [
        ("Nombre del tramo", tramo_dict["nombre"]), ("Comunidad / Sector", tramo_dict["comunidad"]),
        ("Ubicación", f"{tramo_dict['distrito']} / {tramo_dict['provincia']} / {tramo_dict['departamento']}"),
        ("Longitud", f"{tramo_dict['longitud_km']} km"), ("Ancho de plataforma", f"{tramo_dict['ancho_m']} m"),
        ("Tipo de superficie", tramo_dict["tipo_superficie"]), ("Estado actual", tramo_dict["estado_actual"]),
        ("Responsable", tramo_dict["responsable"]), ("Fecha de registro", tramo_dict["fecha_registro"]),
    ]
    y_inicio_tabla = pdf.get_y()
    for etiqueta, valor in campos:
        pdf.set_x(pdf.l_margin)
        pdf.set_font("Helvetica", "B", 9)
        pdf.cell(55, 6, _pdf_safe(etiqueta), border=1)
        pdf.set_font("Helvetica", "", 9)
        pdf.cell(epw - 55 - 75, 6, _pdf_safe(valor), border=1)
        pdf.ln()

    foto_bytes, foto_label = _primera_foto_disponible(fotos)
    if foto_bytes:
        try:
            from PIL import Image as PILImage
            pil_img = PILImage.open(BytesIO(foto_bytes)).convert("RGB")
            tmp_path = "/tmp/_ficha_foto_tmp.jpg"
            pil_img.save(tmp_path, format="JPEG", quality=85)
            pdf.image(tmp_path, x=140, y=y_inicio_tabla, w=58)
            pdf.set_xy(140, y_inicio_tabla + 58 * pil_img.height / pil_img.width + 2)
            pdf.set_font("Helvetica", "I", 7)
            pdf.cell(58, 4, _pdf_safe(foto_label))
            os.remove(tmp_path)
        except Exception:
            pass

    # ---- Página 2: Ficha 1-B Itinerario ----
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(epw, 8, _pdf_safe(f"1-B: FICHA DEL ITINERARIO DEL CAMINO VECINAL — Tramo {tramo_dict['codigo']}"))
    pdf.ln(10)
    df_b = df_ficha_itinerario(tramo_dict)
    cols_b = ["Del Km", "Al Km", "Tipo de Superficie", "Estado de Transitabilidad", "Ancho (m)"]
    anchos_b = [22, 22, 45, 50, 25]
    pdf.set_font("Helvetica", "B", 8)
    for c, w in zip(cols_b, anchos_b):
        pdf.cell(w, 7, _pdf_safe(c), border=1, align="C")
    pdf.ln()
    pdf.set_font("Helvetica", "", 8)
    for _, row in df_b.iterrows():
        if pdf.get_y() > 270:
            pdf.add_page()
        pdf.cell(anchos_b[0], 6, _pdf_safe(row["Del Km"]), border=1, align="C")
        pdf.cell(anchos_b[1], 6, _pdf_safe(row["Al Km"]), border=1, align="C")
        pdf.cell(anchos_b[2], 6, _pdf_safe(row["Tipo de Superficie"])[:22], border=1, align="C")
        pdf.cell(anchos_b[3], 6, _pdf_safe(row["Estado de Transitabilidad"])[:24], border=1, align="C")
        pdf.cell(anchos_b[4], 6, _pdf_safe(row["Ancho de Plataforma (m)"]), border=1, align="C")
        pdf.ln()

    # ---- Página 3: Ficha 1-D Daños ----
    pdf.add_page()
    pdf.set_font("Helvetica", "B", 11)
    pdf.cell(epw, 8, _pdf_safe(f"1-D: FICHA TÉCNICA DE DAÑOS EN CAMINO VECINAL — Tramo {tramo_dict['codigo']}"))
    pdf.ln(10)
    df_d = df_ficha_danos(danos_tramo)
    if df_d.empty:
        pdf.set_font("Helvetica", "I", 9)
        pdf.cell(epw, 7, _pdf_safe("Sin fichas de daño registradas para este tramo."))
    else:
        cols_d = list(df_d.columns)
        ancho_d = epw / len(cols_d)
        pdf.set_font("Helvetica", "B", 8)
        for c in cols_d:
            pdf.cell(ancho_d, 7, _pdf_safe(c), border=1, align="C")
        pdf.ln()
        pdf.set_font("Helvetica", "", 8)
        for _, row in df_d.iterrows():
            if pdf.get_y() > 270:
                pdf.add_page()
            for c in cols_d:
                pdf.cell(ancho_d, 6, _pdf_safe(row[c])[:20], border=1, align="C")
            pdf.ln()

    # ---- Página 4: Evidencia Fotográfica del Mantenimiento ----
    if fotos:
        pdf.add_page()
        pdf.set_fill_color(11, 37, 69)
        pdf.rect(0, 0, pdf.w, 22, "F")
        pdf.set_xy(10, 6)
        pdf.set_text_color(255, 255, 255)
        pdf.set_font("Helvetica", "B", 13)
        pdf.cell(0, 10, _pdf_safe(f"EVIDENCIA FOTOGRÁFICA DEL MANTENIMIENTO — Tramo {tramo_dict['codigo']}"))
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(pdf.l_margin, 28)
        pdf.set_font("Helvetica", "", 9)
        pdf.multi_cell(epw, 5, _pdf_safe(
            f"{len(fotos)} fotografía(s) registrada(s) para este tramo, "
            f"cargadas durante el registro y seguimiento del mantenimiento."
        ))

        cols_grid = 2 if len(fotos) <= 4 else 3
        gap = 6
        cell_w = (epw - gap * (cols_grid - 1)) / cols_grid
        cell_img_h = 55 if cols_grid == 2 else 46
        x_start = pdf.l_margin
        y_start = pdf.get_y() + 4

        try:
            from PIL import Image as PILImage
        except ImportError:
            PILImage = None

        if PILImage:
            for idx, foto in enumerate(fotos[:MAX_FOTOS_TRAMO]):
                col_idx = idx % cols_grid
                row_idx = idx // cols_grid
                x = x_start + col_idx * (cell_w + gap)
                y = y_start + row_idx * (cell_img_h + 15)
                try:
                    contenido_foto = foto["imagen"] if isinstance(foto["imagen"], (bytes, bytearray)) \
                        else descargar_bytes_storage(foto["imagen"])
                    if not contenido_foto:
                        continue
                    pil_img = PILImage.open(BytesIO(contenido_foto)).convert("RGB")
                    tmp_path = f"/tmp/_ficha_foto_{idx}.jpg"
                    pil_img.save(tmp_path, format="JPEG", quality=88)

                    ratio = min(cell_w / pil_img.width, cell_img_h / pil_img.height)
                    w_draw = pil_img.width * ratio
                    h_draw = pil_img.height * ratio
                    x_offset = x + (cell_w - w_draw) / 2
                    y_offset = y + (cell_img_h - h_draw) / 2

                    pdf.set_draw_color(224, 228, 235)
                    pdf.rect(x, y, cell_w, cell_img_h)
                    pdf.image(tmp_path, x=x_offset, y=y_offset, w=w_draw, h=h_draw)

                    pie = foto.get("pie_foto") or foto.get("nombre_archivo") or f"Fotografía {idx + 1}"
                    fecha_c = foto.get("fecha_carga") or ""
                    texto_pie = f"{pie}" + (f"  ·  {fecha_c}" if fecha_c else "")
                    pdf.set_xy(x, y + cell_img_h + 1.5)
                    pdf.set_font("Helvetica", "I", 7)
                    pdf.set_text_color(90, 100, 114)
                    pdf.multi_cell(cell_w, 3.5, _pdf_safe(texto_pie), align="C")
                    pdf.set_text_color(0, 0, 0)
                    os.remove(tmp_path)
                except Exception:
                    continue

    return bytes(pdf.output(dest="S"))


# ──────────────────────────────────────────────────────────────────────────────
# ESTILOS
# ──────────────────────────────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────
# IDENTIDAD VISUAL VIALNET
#   Primario   : #0B2545 (azul corporativo — confianza, infraestructura)
#   Acento     : #14B8A6 (teal — tecnología, movilidad)
#   Superficie : #F7F9FC (fondo neutro claro)
#   Tipografía : Inter / system sans-serif (look de software profesional B2B)
# ──────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700;800&display=swap');

    #MainMenu {visibility: hidden;}
    footer {visibility: hidden;}
    header[data-testid="stHeader"] {background: transparent;}

    html, body, [class*="css"] { font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; }

    :root {
        --vn-primary: #0B2545;
        --vn-primary-light: #15396B;
        --vn-accent: #14B8A6;
        --vn-accent-dark: #0D9488;
        --vn-surface: #F7F9FC;
        --vn-border: #E3E8F0;
        --vn-text-muted: #5A6472;
    }

    .stApp { background: var(--vn-surface); }

    /* ── Encabezado principal (marca) ───────────────────────────────────── */
    .main-header {
        background: linear-gradient(135deg, var(--vn-primary) 0%, var(--vn-primary-light) 60%, #0E3A63 100%);
        padding: 1.8rem 2.2rem; border-radius: 12px; color: white; margin-bottom: 1.4rem;
        border-bottom: 4px solid var(--vn-accent);
        box-shadow: 0 8px 24px rgba(11,37,69,0.22);
    }
    .main-header .brand-row { display:flex; align-items:center; gap:0.6rem; margin-bottom:0.2rem; }
    .main-header .brand-mark {
        font-weight:800; font-size:1.6rem; letter-spacing:-0.02em;
    }
    .main-header .brand-mark span { color: var(--vn-accent); }
    .main-header h2 { font-weight:600; font-size:1.05rem; letter-spacing:0.2px; margin:0.3rem 0 0 0; opacity:0.95; }
    .main-header .subtitle { opacity:0.85; font-size:0.88rem; margin-top:0.35rem; }
    .main-header .tagline {
        display:inline-block; background:var(--vn-accent); color:#04231f; font-weight:700;
        font-size:0.72rem; padding:3px 12px; border-radius:20px; letter-spacing:0.4px;
        text-transform:uppercase; margin-top:0.7rem;
    }

    /* ── Banner reutilizable para encabezar cada módulo ─────────────────── */
    .module-header {
        display:flex; align-items:center; gap:0.8rem;
        background: linear-gradient(120deg, var(--vn-primary) 0%, #16406F 100%);
        border-radius: 10px; padding: 1rem 1.4rem; margin-bottom: 0.5rem;
        border-left: 4px solid var(--vn-accent);
        box-shadow: 0 4px 12px rgba(11,37,69,0.16);
    }
    .module-header .mh-icon {
        font-size:1.4rem; line-height:1; background:rgba(20,184,166,0.15);
        padding:0.5rem; border-radius:8px;
    }
    .module-header .mh-title { color:#fff; font-size:1.25rem; font-weight:700; margin:0; letter-spacing:-0.01em; }
    .module-header .mh-sub { color:#B9C6DC; font-size:0.84rem; margin-top:0.15rem; }

    /* ── Tarjetas KPI ────────────────────────────────────────────────────── */
    .kpi-card {
        background: #ffffff; border: 1px solid var(--vn-border); border-top: 3px solid var(--vn-accent);
        border-radius: 10px; padding: 1.05rem 1.2rem; box-shadow: 0 2px 8px rgba(11,37,69,0.06);
        transition: transform 0.15s ease, box-shadow 0.15s ease;
    }
    .kpi-card:hover { transform: translateY(-3px); box-shadow: 0 10px 20px rgba(11,37,69,0.12); }
    .kpi-value { font-size: 1.9rem; font-weight: 800; color: var(--vn-primary); letter-spacing:-0.02em; }
    .kpi-label { font-size: 0.76rem; color: var(--vn-text-muted); text-transform:uppercase; letter-spacing:0.4px; font-weight:600;}

    /* ── Badges de estado ────────────────────────────────────────────────── */
    .badge-bueno { background:#D1FAE5; color:#065F46; padding:4px 14px; border-radius:20px; font-weight:700; font-size:0.76rem;}
    .badge-regular { background:#FEF3C7; color:#92400E; padding:4px 14px; border-radius:20px; font-weight:700; font-size:0.76rem;}
    .badge-malo { background:#FFEDD5; color:#9A3412; padding:4px 14px; border-radius:20px; font-weight:700; font-size:0.76rem;}
    .badge-muymalo { background:#FEE2E2; color:#991B1B; padding:4px 14px; border-radius:20px; font-weight:700; font-size:0.76rem;}

    /* Aviso de autocompletado inteligente */
    .auto-fill-note {
        background:#ECFDF5; border-left:4px solid var(--vn-accent-dark); border-radius:6px;
        padding:0.6rem 1rem; font-size:0.82rem; color:#065F46; margin-bottom:0.7rem;
    }

    /* ── Sidebar ─────────────────────────────────────────────────────────── */
    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, var(--vn-primary) 0%, #0A2038 100%);
    }
    section[data-testid="stSidebar"] * { color: #F0F2F6 !important; }
    section[data-testid="stSidebar"] .stRadio label { font-size: 0.92rem; padding: 3px 0; }
    section[data-testid="stSidebar"] hr { border-color: rgba(255,255,255,0.14); }

    div[data-baseweb="tab-list"] { gap: 6px; }
    button[data-baseweb="tab"] {
        font-weight:600; border-radius: 8px 8px 0 0 !important;
    }
    button[data-baseweb="tab"][aria-selected="true"] {
        background-color: rgba(20,184,166,0.10) !important; border-bottom: 3px solid var(--vn-accent) !important;
    }

    .stButton button[kind="primary"], .stFormSubmitButton button {
        background-color: var(--vn-primary); border: none; font-weight:600; border-radius:8px;
        transition: background-color 0.15s ease, transform 0.1s ease;
    }
    .stButton button[kind="primary"]:hover, .stFormSubmitButton button:hover {
        background-color: var(--vn-accent-dark);
    }

    div[data-testid="stExpander"] {
        border: 1px solid var(--vn-border); border-radius: 10px; box-shadow: 0 1px 4px rgba(11,37,69,0.05);
    }

    h1, h2, h3, h4, h5 { color: var(--vn-primary); font-weight:700; letter-spacing:-0.01em; }
</style>
""", unsafe_allow_html=True)


def badge_estado(estado: str) -> str:
    clase = {"Bueno": "badge-bueno", "Regular": "badge-regular",
             "Malo": "badge-malo", "Muy Malo": "badge-muymalo"}.get(estado, "badge-regular")
    return f'<span class="{clase}">{estado}</span>'


def module_header(icon: str, titulo: str, subtitulo: str = ""):
    """Banner institucional reutilizable para encabezar cada módulo de la plataforma."""
    st.markdown(f"""
    <div class="module-header">
        <div class="mh-icon">{icon}</div>
        <div>
            <p class="mh-title">{titulo}</p>
            {f'<p class="mh-sub">{subtitulo}</p>' if subtitulo else ''}
        </div>
    </div>
    """, unsafe_allow_html=True)


# ──────────────────────────────────────────────────────────────────────────────
# SIDEBAR — MENÚ PRINCIPAL (5 módulos)
# ──────────────────────────────────────────────────────────────────────────────
st.sidebar.markdown("""
<div style="text-align:center; padding:0.6rem 0 0.9rem 0;">
    <div style="font-size:2rem;">🛣️</div>
    <div style="font-weight:800; font-size:1.15rem; letter-spacing:-0.02em;">
        Vial<span style="color:#14B8A6;">Net</span>
    </div>
    <div style="font-size:0.68rem; opacity:0.75; margin-top:0.3rem; line-height:1.35; text-transform:uppercase; letter-spacing:0.4px;">
    Gestión Inteligente de Mantenimiento Vial<br>Municipalidad Distrital de Ocongate
    </div>
</div>
""", unsafe_allow_html=True)
st.sidebar.markdown("---")

_opciones_menu = ["🏠 INICIO", "🛣️ REGISTRO DE TRAMOS", "📂 CONSULTA DE TRAMOS",
                  "📋 FICHA TÉCNICA DE DAÑOS", "🗓️ PROGRAMACIÓN DE MANTENIMIENTO", "📊 REPORTES"]

# Si una acción reciente (ej. guardar una intervención) solicitó redirigir al usuario
# a otro módulo, se aplica aquí ANTES de instanciar el widget de radio.
if "menu_redirect" in st.session_state:
    st.session_state["menu_radio"] = st.session_state.pop("menu_redirect")

menu = st.sidebar.radio(
    "Menú principal", _opciones_menu,
    label_visibility="collapsed", key="menu_radio",
)

st.sidebar.markdown("---")
st.sidebar.caption(f"Sesión: {datetime.now().strftime('%d/%m/%Y %H:%M')}")
st.sidebar.caption("Proyecto Preprofesional — UTEC 2026-I")

# Selector global de tramo (M1-M5) — disponible para filtrar en todos los módulos
_tramos_sidebar = df_tramos()
st.sidebar.markdown("---")
st.sidebar.markdown("**📍 Tramo activo**")
if not _tramos_sidebar.empty:
    _opciones_sb = ["Todos los tramos"] + [
        f"{r.codigo} — {r.nombre}" for r in _tramos_sidebar.sort_values("codigo").itertuples()
    ]
    _sel_sb = st.sidebar.selectbox("Filtrar por tramo", _opciones_sb, label_visibility="collapsed")
    if _sel_sb == "Todos los tramos":
        st.session_state["tramo_activo_codigo"] = None
    else:
        st.session_state["tramo_activo_codigo"] = _sel_sb.split(" — ")[0]
else:
    st.session_state["tramo_activo_codigo"] = None


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 1 — INICIO
# ══════════════════════════════════════════════════════════════════════════════
if menu == "🏠 INICIO":
    st.markdown("""
    <div class="main-header">
        <div class="brand-row">
            <span class="brand-mark">Vial<span>Net</span></span>
        </div>
        <h2>Gestión Preventiva de Trochas Carrozables</h2>
        <p class="subtitle" style="margin:0.3rem 0 0 0;">
        Municipalidad Distrital de Ocongate · Provincia de Quispicanchi · Región Cusco
        </p>
        <span class="tagline">Subgerencia de Gestión de Riesgos y Mantenimiento</span>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("""
    Esta plataforma apoya a la **Subgerencia de Gestión de Riesgos y Mantenimiento** en el
    registro, inspección técnica y seguimiento del estado de la red vial no pavimentada
    (trochas carrozables) del distrito, siguiendo los criterios del **Manual de Conservación
    Vial MCV-2014 (MTC)** y los formatos técnicos de **Provías Descentralizado**.
    """)

    tramos = df_tramos()
    danos = df_danos()

    _tac = st.session_state.get("tramo_activo_codigo")
    if _tac:
        tramos = tramos[tramos["codigo"] == _tac]
        if not danos.empty:
            danos = danos[danos["codigo_tramo"] == _tac]
        st.info(f"📍 Mostrando datos del tramo **{_tac}** únicamente. Cambia el filtro en el panel lateral para ver todos los tramos.")

    total_tramos = len(tramos)
    km_inventariados = round(tramos["longitud_km"].sum(), 2) if not tramos.empty else 0.0

    if not tramos.empty:
        conteo_estado = tramos["estado_actual"].value_counts()
        n_bueno = int(conteo_estado.get("Bueno", 0))
        n_regular = int(conteo_estado.get("Regular", 0))
        n_malo = int(conteo_estado.get("Malo", 0)) + int(conteo_estado.get("Muy Malo", 0))
    else:
        n_bueno = n_regular = n_malo = 0

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{total_tramos}</div>'
                     f'<div class="kpi-label">Tramos registrados</div></div>', unsafe_allow_html=True)
    with c2:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{km_inventariados}</div>'
                     f'<div class="kpi-label">Km inventariados</div></div>', unsafe_allow_html=True)
    with c3:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{n_bueno}</div>'
                     f'<div class="kpi-label">Tramos en buen estado</div></div>', unsafe_allow_html=True)
    with c4:
        st.markdown(f'<div class="kpi-card"><div class="kpi-value">{n_malo}</div>'
                     f'<div class="kpi-label">Tramos en mal / muy mal estado</div></div>', unsafe_allow_html=True)

    st.markdown("###")
    colA, colB = st.columns(2)

    with colA:
        st.markdown("##### Distribución de tramos por estado")
        if not tramos.empty:
            fig = px.pie(
                tramos, names="estado_actual", hole=0.45,
                color="estado_actual",
                color_discrete_map={"Bueno": "#43a047", "Regular": "#fdd835",
                                     "Malo": "#fb8c00", "Muy Malo": "#e53935"},
            )
            fig.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=320)
            st.plotly_chart(fig, use_container_width=True)
        else:
            st.info("Aún no hay tramos registrados. Ve a **🛣️ Registro de Tramos**.")

    with colB:
        st.markdown("##### Kilómetros por comunidad / sector")
        if not tramos.empty:
            agg = tramos.groupby("comunidad")["longitud_km"].sum().reset_index()
            fig2 = px.bar(agg, x="comunidad", y="longitud_km", color="longitud_km",
                          color_continuous_scale="Greens", text_auto=".1f")
            fig2.update_layout(margin=dict(t=10, b=10, l=10, r=10), height=320,
                               yaxis_title="km", xaxis_title="")
            st.plotly_chart(fig2, use_container_width=True)
        else:
            st.info("Sin datos aún.")

    st.markdown("###")
    st.markdown("##### Indicadores rápidos de inspección")
    if not danos.empty:
        c1, c2, c3 = st.columns(3)
        c1.metric("Daños registrados", len(danos))
        c2.metric("ICT promedio de la red", f"{danos['ict'].mean():.1f} / 100")
        c3.metric("Tramos con prioridad Urgente/Alta",
                  int(danos["prioridad"].isin(["Urgente", "Alta"]).sum()))
    else:
        st.info("Aún no se han registrado fichas de daños. Ve a **📋 Ficha Técnica de Daños**.")


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 2 — REGISTRO DE TRAMOS
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "🛣️ REGISTRO DE TRAMOS":
    module_header("🛣️", "Registro de Tramos", "Datos generales del tramo y programación de su mantenimiento, en un solo flujo.")

    tab_nuevo, tab_mant = st.tabs(["🆕 Nuevo tramo", "🛠️ Programar mantenimiento del tramo"])

    with tab_nuevo:
        # ── Fotografías del mantenimiento ──────────────────────────────────
        # Vive FUERA del st.form: así los botones "Quitar" y la vista previa
        # reaccionan al instante, sin esperar al envío del formulario.
        st.markdown("#### 📷 Fotografías del mantenimiento")
        st.caption("Sube entre 1 y 5 fotografías (JPG, JPEG o PNG) que evidencien el estado del tramo o la intervención realizada.")

        st.session_state.setdefault("fotos_nuevo_tramo", [])
        st.session_state.setdefault("fotos_nuevo_tramo_uploader_key", 0)

        _fotos_actuales = st.session_state["fotos_nuevo_tramo"]

        _nuevas = st.file_uploader(
            "Selecciona una o varias fotografías",
            type=["jpg", "jpeg", "png"],
            accept_multiple_files=True,
            label_visibility="collapsed",
            key=f"uploader_fotos_nuevo_tramo_{st.session_state['fotos_nuevo_tramo_uploader_key']}",
        )

        if _nuevas:
            nombres_existentes = {f["nombre_archivo"] for f in _fotos_actuales}
            espacio_disponible = MAX_FOTOS_TRAMO - len(_fotos_actuales)
            agregadas, rechazadas = 0, 0
            for nf in _nuevas:
                if nf.name in nombres_existentes:
                    continue
                if agregadas >= espacio_disponible:
                    rechazadas += 1
                    continue
                st.session_state["fotos_nuevo_tramo"].append({
                    "nombre_archivo": nf.name, "imagen": nf.getvalue(),
                    "pie_foto": "", "fecha_carga": date.today().strftime("%Y-%m-%d"),
                })
                nombres_existentes.add(nf.name)
                agregadas += 1
            if rechazadas > 0:
                st.error(
                    f"⚠️ Solo se permiten un máximo de **{MAX_FOTOS_TRAMO} fotografías** por tramo. "
                    f"Se descartó {rechazadas} imagen adicional." if rechazadas == 1 else
                    f"⚠️ Solo se permiten un máximo de **{MAX_FOTOS_TRAMO} fotografías** por tramo. "
                    f"Se descartaron {rechazadas} imágenes adicionales."
                )
            if agregadas > 0:
                st.session_state["fotos_nuevo_tramo_uploader_key"] += 1
                st.rerun()

        _fotos_actuales = st.session_state["fotos_nuevo_tramo"]
        _n_fotos = len(_fotos_actuales)
        _color_contador = "#0D9488" if _n_fotos > 0 else "#9AA5B1"
        st.markdown(
            f'<div style="display:inline-block; background:{_color_contador}1A; color:{_color_contador}; '
            f'font-weight:700; font-size:0.82rem; padding:4px 14px; border-radius:20px; margin:0.3rem 0 0.8rem 0;">'
            f'📸 Fotos cargadas: {_n_fotos}/{MAX_FOTOS_TRAMO}</div>',
            unsafe_allow_html=True,
        )

        if _n_fotos == 0:
            st.info("Aún no has cargado fotografías. Puedes agregar hasta 5 antes de guardar el tramo (opcional).")
        else:
            cols_prev = st.columns(5)
            for i, foto in enumerate(_fotos_actuales):
                with cols_prev[i % 5]:
                    st.image(BytesIO(foto["imagen"]), use_container_width=True)
                    st.caption(f"{foto['nombre_archivo'][:20]}")
                    if st.button("✕ Quitar", key=f"quitar_foto_nuevo_{i}", use_container_width=True):
                        st.session_state["fotos_nuevo_tramo"].pop(i)
                        st.rerun()

        st.markdown("---")

        with st.form("form_registro_tramo", clear_on_submit=True):
            st.markdown("#### Datos Generales")
            c1, c2, c3 = st.columns(3)
            codigo = c1.text_input("Código de tramo *", placeholder="Ej: M-01")
            nombre = c2.text_input("Nombre del tramo *", placeholder="Ej: Ocongate - Pacchanta")
            comunidad = c3.text_input("Comunidad / Sector *", placeholder="Ej: Pacchanta Baja")

            c4, c5, c6 = st.columns(3)
            distrito = c4.text_input("Distrito *", value="Ocongate")
            provincia = c5.text_input("Provincia *", value="Quispicanchi")
            departamento = c6.text_input("Departamento *", value="Cusco")

            c7, c8, c9 = st.columns(3)
            longitud_km = c7.number_input("Longitud (km) *", min_value=0.0, step=0.1, format="%.2f")
            ancho_m = c8.number_input("Ancho de plataforma (m) *", min_value=0.0, step=0.1, format="%.2f")
            tipo_superficie = c9.selectbox("Tipo de superficie *", TIPOS_SUPERFICIE)

            c10, c11, c12 = st.columns(3)
            estado_actual = c10.selectbox("Estado actual *", ESTADOS_TRAMO)
            responsable = c11.text_input("Responsable de registro *",
                                          value=st.session_state.get("ultimo_responsable", ""))
            telefono_responsable = c12.text_input("Número telefónico del responsable", placeholder="9XXXXXXXX",
                                                    value=st.session_state.get("ultimo_telefono", ""))

            fecha_registro = st.date_input("Fecha de registro *", value=date.today())

            with st.expander("🔧 Datos técnicos avanzados y recursos (opcional)"):
                a1, a2, a3 = st.columns(3)
                altitud_msnm = a1.number_input("Altitud (m s.n.m.)", min_value=0.0, step=10.0, value=0.0)
                estado_drenaje = a2.text_input("Estado del drenaje", placeholder="Ej: Deficiente — cunetas colmatadas")
                senalizacion = a3.text_input("Señalización", placeholder="Ej: Ausente / Parcial / Completa")

                a4, a5, a6 = st.columns(3)
                n_puentes = a4.number_input("N° de puentes", min_value=0, step=1, value=0)
                n_badenes = a5.number_input("N° de badenes", min_value=0, step=1, value=0)
                accesibilidad = a6.text_input("Accesibilidad", placeholder="Ej: Estacional (cerrado dic-mar)")

                a7, a8 = st.columns(2)
                cantera_fuente = a7.text_input("Cantera / fuente de material", placeholder="Ej: Sí — Cantera Pacchanta (2.3 km)")
                fuente_lastre = a8.text_input("Fuente de lastre", placeholder="Ej: Río Ausangate (material aluvial)")

                a9, a10, a11 = st.columns(3)
                ultima_intervencion = a9.date_input("Fecha de última intervención", value=None)
                costo_rutinario = a10.number_input("Costo est. mant. rutinario (S/.)", min_value=0.0, step=1000.0, value=0.0)
                costo_periodico = a11.number_input("Costo est. mant. periódico (S/.)", min_value=0.0, step=1000.0, value=0.0)

                a12, a13 = st.columns(2)
                gps_lat = a12.number_input("Coordenada GPS — Latitud", value=0.0, format="%.4f")
                gps_lon = a13.number_input("Coordenada GPS — Longitud", value=0.0, format="%.4f")

            st.markdown("#### Archivos Adjuntos")
            adjunto_pdf = st.file_uploader("Adjuntar PDF (expediente, ficha, plano, etc.)", type=["pdf"])

            observaciones_tramo = st.text_area(
                "Observaciones técnicas", placeholder="Notas, hallazgos relevantes, condiciones especiales del tramo...")

            enviado = st.form_submit_button("💾 Guardar tramo en la base de datos", use_container_width=True)

        if enviado:
            if not codigo or not nombre or not comunidad or not responsable:
                st.error("Completa los campos obligatorios marcados con *.")
            else:
                try:
                    data = {
                        "codigo": codigo, "nombre": nombre, "comunidad": comunidad,
                        "distrito": distrito, "provincia": provincia, "departamento": departamento,
                        "longitud_km": longitud_km, "ancho_m": ancho_m,
                        "tipo_superficie": tipo_superficie, "estado_actual": estado_actual,
                        "responsable": responsable, "telefono_responsable": telefono_responsable,
                        "fecha_registro": fecha_registro.strftime("%Y-%m-%d"),
                        "adjunto_pdf": adjunto_pdf.getvalue() if adjunto_pdf else None,
                        "adjunto_pdf_nombre": adjunto_pdf.name if adjunto_pdf else None,
                        "altitud_msnm": altitud_msnm if altitud_msnm else None,
                        "estado_drenaje": estado_drenaje or None,
                        "senalizacion": senalizacion or None,
                        "n_puentes": n_puentes if n_puentes else None,
                        "n_badenes": n_badenes if n_badenes else None,
                        "accesibilidad": accesibilidad or None,
                        "cantera_fuente": cantera_fuente or None,
                        "fuente_lastre": fuente_lastre or None,
                        "ultima_intervencion": ultima_intervencion.strftime("%Y-%m-%d") if ultima_intervencion else None,
                        "costo_rutinario": costo_rutinario if costo_rutinario else None,
                        "costo_periodico": costo_periodico if costo_periodico else None,
                        "gps_lat": gps_lat if gps_lat else None,
                        "gps_lon": gps_lon if gps_lon else None,
                        "observaciones": observaciones_tramo or None,
                    }
                    nuevo_id = insertar_tramo(data)

                    fotos_a_guardar = st.session_state.get("fotos_nuevo_tramo", [])
                    if fotos_a_guardar:
                        guardar_fotos_tramo(nuevo_id, fotos_a_guardar)

                    st.session_state["ultimo_responsable"] = responsable
                    st.session_state["ultimo_telefono"] = telefono_responsable
                    n_fotos_ok = len(fotos_a_guardar)
                    st.session_state["fotos_nuevo_tramo"] = []
                    st.session_state["fotos_nuevo_tramo_uploader_key"] += 1

                    st.success(
                        f"✅ Tramo **{codigo} — {nombre}** guardado correctamente en la base de datos"
                        + (f", junto con {n_fotos_ok} fotografía(s)." if n_fotos_ok else ".")
                    )
                except psycopg2.errors.UniqueViolation:
                    st.error(f"⚠️ Ya existe un tramo con el código **{codigo}**. Usa un código distinto.")
                except psycopg2.OperationalError:
                    st.error("⚠️ No se pudo conectar a la base de datos en la nube. Intenta nuevamente en unos segundos.")
                except Exception as e:
                    st.error(f"⚠️ Ocurrió un error inesperado al guardar el tramo: {e}")

        st.markdown("---")
        st.caption(f"Tramos registrados actualmente: **{len(df_tramos())}**")

    with tab_mant:
        st.markdown("#### 🛠️ Programar mantenimiento para un tramo")
        st.caption("El sistema sugiere fechas tentativas según el tipo de mantenimiento, la temporada de lluvias y la criticidad del tramo.")

        with st.expander("ℹ️ Tipos de mantenimiento vial y actividades asociadas", expanded=False):
            for nombre, info in MANTENIMIENTOS.items():
                regla = REGLAS_MANTENIMIENTO.get(nombre, {})
                st.markdown(
                    f'<div style="border-left:5px solid {info["color"]}; padding:0.4rem 0.8rem; margin-bottom:0.5rem;">'
                    f'<b style="color:{info["color"]};">{nombre}</b> '
                    f'<span style="font-size:0.78rem; color:#777;">'
                    f'(frecuencia recomendada: cada {regla.get("frecuencia_dias", "—")} días · '
                    f'duración típica: {regla.get("duracion_dias", "—")} días)</span><br>'
                    f'<span style="font-size:0.86rem;">{info["descripcion"]}</span><br>'
                    f'<span style="font-size:0.82rem; color:#444;">• ' + "<br>• ".join(info["actividades"]) + '</span>'
                    f'</div>', unsafe_allow_html=True,
                )

        tramos_mant = df_tramos()
        danos_mant = df_danos()

        if tramos_mant.empty:
            st.info("Primero registra al menos un tramo en la pestaña **🆕 Nuevo tramo**.")
        else:
            opciones_tramo_m = {f"{row.codigo} — {row.nombre}": row.id for row in tramos_mant.itertuples()}
            tramo_default_id_m = st.session_state.get("ultimo_dano_tramo")
            if tramo_default_id_m is None and st.session_state.get("tramo_activo_codigo"):
                _fila_tac = tramos_mant[tramos_mant["codigo"] == st.session_state["tramo_activo_codigo"]]
                if not _fila_tac.empty:
                    tramo_default_id_m = int(_fila_tac.iloc[0]["id"])

            s1, s2 = st.columns(2)
            sel_tramo_label_m = s1.selectbox(
                "Tramo", list(opciones_tramo_m.keys()),
                index=list(opciones_tramo_m.values()).index(tramo_default_id_m) if tramo_default_id_m in opciones_tramo_m.values() else 0,
                key="prog_sel_tramo",
            )
            tramo_id_sel = opciones_tramo_m[sel_tramo_label_m]
            tramo_sel_dict = tramos_mant[tramos_mant["id"] == tramo_id_sel].iloc[0].to_dict()
            s2.text_input("Ubicación del tramo", value=f"{tramo_sel_dict['comunidad']}, {tramo_sel_dict['distrito']} ({tramo_sel_dict['longitud_km']} km)", disabled=True)

            danos_tramo = danos_mant[danos_mant["tramo_id"] == tramo_id_sel] if not danos_mant.empty else pd.DataFrame()
            opciones_dano = {"— Sin vincular a ficha de daño —": None}
            if not danos_tramo.empty:
                for row in danos_tramo.itertuples():
                    opciones_dano[f"#{row.id} · {row.tipo_dano} · ICT {row.ict} · sugerido: {row.tipo_mantenimiento}"] = row.id
            sel_dano_label = st.selectbox("Ficha de daño vinculada (opcional)", list(opciones_dano.keys()), key="prog_sel_dano")
            dano_id_sel = opciones_dano[sel_dano_label]

            sugerido_tipo, componente_sugerido, prioridad_sugerida = None, "Superficie", "Media"
            if dano_id_sel is not None:
                fila_dano = danos_tramo[danos_tramo["id"] == dano_id_sel].iloc[0]
                sugerido_tipo = fila_dano["tipo_mantenimiento"]
                componente_sugerido = determinar_componente(fila_dano["tipo_dano"])
                prioridad_sugerida = fila_dano["prioridad"]

            s3, s4, s5 = st.columns(3)
            tipo_mant = s3.selectbox(
                "Tipo de mantenimiento *", list(MANTENIMIENTOS.keys()),
                index=list(MANTENIMIENTOS.keys()).index(sugerido_tipo) if sugerido_tipo in MANTENIMIENTOS else 0,
                key="prog_sel_tipo",
                help="Al vincular una ficha de daño se preselecciona el tipo sugerido (editable).",
            )
            componente_int = s4.selectbox(
                "Componente vial intervenido *", COMPONENTES_VIA,
                index=COMPONENTES_VIA.index(componente_sugerido) if componente_sugerido in COMPONENTES_VIA else 0,
                key="prog_sel_componente",
            )
            prioridad_int = s5.selectbox(
                "Prioridad de intervención *", PRIORIDADES_INTERVENCION,
                index=PRIORIDADES_INTERVENCION.index(prioridad_sugerida) if prioridad_sugerida in PRIORIDADES_INTERVENCION else 1,
                key="prog_sel_prioridad",
            )

            sugerencia = sugerir_programacion(tipo_mant, componente_int, tramo_sel_dict, prioridad_int)

            st.markdown(
                f'<div style="background:#0b254508; border-left:4px solid #0b2545; border-radius:6px; '
                f'padding:0.8rem 1rem; margin:0.6rem 0;">'
                f'💡 <b>Sugerencia automática:</b> inicio tentativo <b>{sugerencia["fecha_inicio"].strftime("%d/%m/%Y")}</b>, '
                f'fin tentativo <b>{sugerencia["fecha_fin"].strftime("%d/%m/%Y")}</b> '
                f'({sugerencia["duracion_dias"]} días) · frecuencia recomendada cada {sugerencia["frecuencia_dias"]} días.'
                f'</div>', unsafe_allow_html=True,
            )
            for adv in sugerencia["advertencias"]:
                if adv.startswith("🌧️") or adv.startswith("⚠️"):
                    st.warning(adv)
                else:
                    st.info(adv)

            st.markdown("##### Confirmar / ajustar la programación")
            with st.form("form_programacion", clear_on_submit=True):
                actividades_sel = st.multiselect(
                    "Actividades a ejecutar *", MANTENIMIENTOS[tipo_mant]["actividades"],
                    default=MANTENIMIENTOS[tipo_mant]["actividades"],
                )

                c1, c2, c3 = st.columns(3)
                fecha_inicio_form = c1.date_input("Fecha tentativa de inicio *", value=sugerencia["fecha_inicio"])
                fecha_fin_form = c2.date_input("Fecha tentativa de fin *", value=sugerencia["fecha_fin"])
                duracion_dias = c3.number_input("Plazo estimado (días)", min_value=1.0,
                                                 value=float(sugerencia["duracion_dias"]), step=1.0)

                c4, c5 = st.columns(2)
                expediente = c4.text_input("N° de Expediente Técnico / código de intervención", placeholder="Ej: ET-MDO-006")
                costo_estimado = c5.number_input(
                    "Costo estimado / presupuesto (S/.)", min_value=0.0, step=500.0,
                    value=float(sugerencia["costo_referencial"]) if sugerencia["costo_referencial"] else 10000.0,
                )

                estado_int = st.selectbox("Estado de la intervención", ESTADOS_INTERVENCION)

                responsable_int = st.text_input(
                    "Responsable / área encargada",
                    value=st.session_state.get("ultimo_responsable_mant", "Subgerencia de Gestión de Riesgos y Mantenimiento"))
                observaciones_int = st.text_area("Descripción técnica de los trabajos a ejecutar / Observaciones")
                foto_evidencia_int = st.file_uploader("Evidencia fotográfica o archivo asociado (opcional)",
                                                       type=["jpg", "jpeg", "png"])

                guardar_int = st.form_submit_button("💾 Guardar programación", use_container_width=True)

            if guardar_int:
                try:
                    insertar_intervencion({
                        "tramo_id": tramo_id_sel, "dano_id": dano_id_sel, "tipo_mantenimiento": tipo_mant,
                        "componente": componente_int, "prioridad": prioridad_int,
                        "actividades": "; ".join(actividades_sel), "expediente": expediente,
                        "fecha_programada": fecha_inicio_form.strftime("%Y-%m-%d"),
                        "fecha_fin": fecha_fin_form.strftime("%Y-%m-%d"),
                        "duracion_dias": duracion_dias, "costo_estimado": costo_estimado, "estado": estado_int,
                        "responsable": responsable_int, "observaciones": observaciones_int,
                        "foto_evidencia": img_to_blob(foto_evidencia_int, carpeta="fotos-evidencia-mantenimiento"),
                    })
                except Exception as e:
                    st.error(f"⚠️ No se pudo guardar la programación de mantenimiento: {e}")
                    st.stop()
                st.session_state["ultimo_responsable_mant"] = responsable_int
                for k in ["ultimo_dano_id", "ultimo_dano_tramo", "ultimo_dano_mant", "ultimo_dano_dias"]:
                    st.session_state.pop(k, None)
                # Confirmación + redirección automática a Programación de Mantenimiento,
                # para que el Gantt y el resumen reflejen la intervención de inmediato.
                st.session_state["intervencion_guardada_info"] = (
                    f"Intervención de **{tipo_mant}** ({componente_int}) programada para el tramo "
                    f"**{tramo_sel_dict['codigo']} — {tramo_sel_dict['nombre']}** "
                    f"del {fecha_inicio_form.strftime('%d/%m/%Y')} al {fecha_fin_form.strftime('%d/%m/%Y')}."
                )
                st.session_state["tramo_activo_codigo"] = tramo_sel_dict["codigo"]
                st.session_state["menu_redirect"] = "🗓️ PROGRAMACIÓN DE MANTENIMIENTO"
                st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 3 — CONSULTA DE TRAMOS
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📂 CONSULTA DE TRAMOS":
    module_header("📂", "Consulta de Tramos", "Búsqueda, filtros, edición, fotografías y descarga de fichas técnicas por tramo.")
    tramos = df_tramos()

    if tramos.empty:
        st.info("No hay tramos registrados todavía.")
    else:
        st.markdown("#### Filtros")
        f1, f2, f3, f4 = st.columns(4)
        busq_codigo = f1.text_input("Buscar por código")
        busq_nombre = f2.text_input("Buscar por nombre")
        filtro_comunidad = f3.selectbox("Filtrar por comunidad", ["Todas"] + sorted(tramos["comunidad"].dropna().unique().tolist()))
        filtro_estado = f4.selectbox("Filtrar por estado", ["Todos"] + ESTADOS_TRAMO)

        min_km, max_km = float(tramos["longitud_km"].min()), float(tramos["longitud_km"].max())
        if min_km == max_km:
            max_km += 1
        rango_long = st.slider("Filtrar por longitud (km)", min_value=float(min_km), max_value=float(max_km),
                                value=(float(min_km), float(max_km)))

        df_f = tramos.copy()
        if busq_codigo:
            df_f = df_f[df_f["codigo"].str.contains(busq_codigo, case=False, na=False)]
        if busq_nombre:
            df_f = df_f[df_f["nombre"].str.contains(busq_nombre, case=False, na=False)]
        if filtro_comunidad != "Todas":
            df_f = df_f[df_f["comunidad"] == filtro_comunidad]
        if filtro_estado != "Todos":
            df_f = df_f[df_f["estado_actual"] == filtro_estado]
        df_f = df_f[(df_f["longitud_km"] >= rango_long[0]) & (df_f["longitud_km"] <= rango_long[1])]

        st.markdown(f"#### Resultados ({len(df_f)} tramo(s))")
        st.dataframe(
            df_f[["codigo", "nombre", "comunidad", "distrito", "longitud_km", "ancho_m",
                  "tipo_superficie", "estado_actual", "responsable", "fecha_registro"]],
            use_container_width=True, hide_index=True,
        )

        st.markdown("---")
        st.markdown("#### Detalle / Edición / Eliminación de un tramo")
        opciones = {f"{row.codigo} — {row.nombre}": row.id for row in df_f.itertuples()}
        if opciones:
            seleccion = st.selectbox("Selecciona un tramo", list(opciones.keys()))
            tramo_id = opciones[seleccion]

            conn = get_conn()
            row = conn.execute("SELECT * FROM tramos WHERE id=?", (tramo_id,)).fetchone()
            cols = [d[0] for d in conn.execute("SELECT * FROM tramos LIMIT 1").description]
            conn.close()
            tramo_dict = dict(zip(cols, row))
            fotos_tramo = obtener_fotos_tramo(tramo_id)

            tabs = st.tabs(["📄 Ficha", "📷 Fotografías", "✏️ Editar", "🗑️ Eliminar", "⬇️ Descargar"])

            with tabs[0]:
                c1, c2 = st.columns(2)
                with c1:
                    st.write(f"**Código:** {tramo_dict['codigo']}")
                    st.write(f"**Nombre:** {tramo_dict['nombre']}")
                    st.write(f"**Comunidad:** {tramo_dict['comunidad']}")
                    st.write(f"**Distrito / Provincia / Dpto:** {tramo_dict['distrito']} / {tramo_dict['provincia']} / {tramo_dict['departamento']}")
                    st.write(f"**Longitud:** {tramo_dict['longitud_km']} km")
                    st.write(f"**Fecha de registro:** {tramo_dict['fecha_registro']}")
                with c2:
                    st.write(f"**Ancho de plataforma:** {tramo_dict['ancho_m']} m")
                    st.write(f"**Tipo de superficie:** {tramo_dict['tipo_superficie']}")
                    st.markdown(f"**Estado actual:** {badge_estado(tramo_dict['estado_actual'])}", unsafe_allow_html=True)
                    st.write(f"**Responsable:** {tramo_dict['responsable']} ({tramo_dict['telefono_responsable'] or 's/n'})")
                    st.write(f"**Fecha de última intervención:** {tramo_dict.get('ultima_intervencion') or 'No registrada'}")
                if tramo_dict.get("observaciones"):
                    st.markdown("**Observaciones técnicas:**")
                    st.info(tramo_dict["observaciones"])
                if tramo_dict["adjunto_pdf"]:
                    st.link_button("⬇️ Descargar PDF adjunto", url=tramo_dict["adjunto_pdf"],
                                    use_container_width=False)
                    st.caption(f"Archivo: {tramo_dict['adjunto_pdf_nombre'] or 'adjunto.pdf'}")

            with tabs[1]:
                _n_fotos_tab = len(fotos_tramo)
                _color_ct = "#0D9488" if _n_fotos_tab > 0 else "#9AA5B1"
                st.markdown(
                    f'<div style="display:inline-block; background:{_color_ct}1A; color:{_color_ct}; '
                    f'font-weight:700; font-size:0.82rem; padding:4px 14px; border-radius:20px; margin-bottom:0.8rem;">'
                    f'📸 Fotos cargadas: {_n_fotos_tab}/{MAX_FOTOS_TRAMO}</div>',
                    unsafe_allow_html=True,
                )
                if not fotos_tramo:
                    st.info("Este tramo aún no tiene fotografías del mantenimiento cargadas.")
                else:
                    cols_gal = st.columns(min(len(fotos_tramo), 5))
                    for i, foto in enumerate(fotos_tramo):
                        with cols_gal[i % len(cols_gal)]:
                            st.image(blob_to_img(foto["imagen"]), use_container_width=True)
                            pie_mostrado = foto["pie_foto"] or foto["nombre_archivo"] or f"Foto {i + 1}"
                            st.caption(f"{pie_mostrado}")
                            if foto["fecha_carga"]:
                                st.caption(f"🗓️ {foto['fecha_carga']}")
                            if st.button("🗑️ Eliminar", key=f"del_foto_{foto['id']}", use_container_width=True):
                                eliminar_foto_tramo(foto["id"])
                                st.rerun()

                st.markdown("---")
                if _n_fotos_tab < MAX_FOTOS_TRAMO:
                    _nuevas_extra = st.file_uploader(
                        f"Agregar más fotografías (máx. {MAX_FOTOS_TRAMO - _n_fotos_tab} adicional(es))",
                        type=["jpg", "jpeg", "png"], accept_multiple_files=True,
                        key=f"uploader_fotos_extra_{tramo_id}",
                    )
                    if _nuevas_extra:
                        espacio = MAX_FOTOS_TRAMO - _n_fotos_tab
                        a_guardar = [
                            {"nombre_archivo": nf.name, "imagen": nf.getvalue(),
                             "pie_foto": "", "fecha_carga": date.today().strftime("%Y-%m-%d")}
                            for nf in _nuevas_extra[:espacio]
                        ]
                        if len(_nuevas_extra) > espacio:
                            st.error(f"⚠️ Solo se permiten un máximo de {MAX_FOTOS_TRAMO} fotografías por tramo. "
                                      f"Se descartaron {len(_nuevas_extra) - espacio} imagen(es) adicional(es).")
                        if a_guardar:
                            guardar_fotos_tramo(tramo_id, a_guardar)
                            st.rerun()
                else:
                    st.caption(f"Se alcanzó el máximo de {MAX_FOTOS_TRAMO} fotografías para este tramo.")

            with tabs[2]:
                with st.form(f"editar_{tramo_id}"):
                    e1, e2, e3 = st.columns(3)
                    n_codigo = e1.text_input("Código", value=tramo_dict["codigo"])
                    n_nombre = e2.text_input("Nombre", value=tramo_dict["nombre"])
                    n_comunidad = e3.text_input("Comunidad", value=tramo_dict["comunidad"])
                    e4, e5, e6 = st.columns(3)
                    n_distrito = e4.text_input("Distrito", value=tramo_dict["distrito"])
                    n_provincia = e5.text_input("Provincia", value=tramo_dict["provincia"])
                    n_departamento = e6.text_input("Departamento", value=tramo_dict["departamento"])
                    e7, e8, e9 = st.columns(3)
                    n_long = e7.number_input("Longitud (km)", value=float(tramo_dict["longitud_km"]), step=0.1)
                    n_ancho = e8.number_input("Ancho (m)", value=float(tramo_dict["ancho_m"]), step=0.1)
                    n_sup = e9.selectbox("Tipo de superficie", TIPOS_SUPERFICIE,
                                          index=TIPOS_SUPERFICIE.index(tramo_dict["tipo_superficie"]) if tramo_dict["tipo_superficie"] in TIPOS_SUPERFICIE else 0)
                    e10, e11 = st.columns(2)
                    n_estado = e10.selectbox("Estado actual", ESTADOS_TRAMO,
                                              index=ESTADOS_TRAMO.index(tramo_dict["estado_actual"]) if tramo_dict["estado_actual"] in ESTADOS_TRAMO else 0)
                    n_resp = e11.text_input("Responsable", value=tramo_dict["responsable"])
                    n_tel = st.text_input("Teléfono", value=tramo_dict["telefono_responsable"] or "")
                    guardar = st.form_submit_button("💾 Guardar cambios")
                if guardar:
                    try:
                        actualizar_tramo(tramo_id, {
                            "codigo": n_codigo, "nombre": n_nombre, "comunidad": n_comunidad,
                            "distrito": n_distrito, "provincia": n_provincia, "departamento": n_departamento,
                            "longitud_km": n_long, "ancho_m": n_ancho, "tipo_superficie": n_sup,
                            "estado_actual": n_estado, "responsable": n_resp, "telefono_responsable": n_tel,
                        })
                        st.success("✅ Tramo actualizado. Recarga la página o vuelve a filtrar para ver los cambios.")
                    except psycopg2.errors.UniqueViolation:
                        st.error(f"⚠️ Ya existe otro tramo con el código **{n_codigo}**. Usa un código distinto.")
                    except Exception as e:
                        st.error(f"⚠️ Ocurrió un error inesperado al actualizar el tramo: {e}")

            with tabs[3]:
                st.warning("Esta acción eliminará el tramo junto con sus fichas de daño, intervenciones "
                           "programadas y fotografías asociadas (incluyendo los archivos en Storage).")
                if st.button("🗑️ Confirmar eliminación", key=f"del_{tramo_id}"):
                    try:
                        eliminar_tramo(tramo_id)
                        st.success("Tramo eliminado. Vuelve a cargar la consulta.")
                    except Exception as e:
                        st.error(f"⚠️ Ocurrió un error al eliminar el tramo: {e}")

            with tabs[4]:
                danos_tramo_dl = df_danos()
                danos_tramo_dl = (danos_tramo_dl[danos_tramo_dl["tramo_id"] == tramo_id]
                                   if not danos_tramo_dl.empty else danos_tramo_dl)

                st.markdown(
                    f'<div style="background:#0b254508; border-left:4px solid #0b2545; '
                    f'border-radius:6px; padding:0.7rem 1rem; margin-bottom:0.8rem;">'
                    f'<b>Vista previa — Ficha Técnica del Tramo {tramo_dict["codigo"]}</b></div>',
                    unsafe_allow_html=True,
                )

                col_prev1, col_prev2 = st.columns([1, 1.4])
                with col_prev1:
                    if fotos_tramo:
                        foto_bytes_prev, foto_label_prev = _primera_foto_disponible(fotos_tramo)
                        st.image(blob_to_img(foto_bytes_prev), caption=foto_label_prev, use_container_width=True)
                        if len(fotos_tramo) > 1:
                            st.caption(f"+ {len(fotos_tramo) - 1} fotografía(s) más incluida(s) en el documento descargable.")
                    else:
                        st.info("Este tramo no tiene fotografías cargadas todavía. Agrega una en el módulo de Registro.")

                with col_prev2:
                    st.markdown(f"**Código:** {tramo_dict['codigo']} &nbsp;·&nbsp; **Nombre:** {tramo_dict['nombre']}", unsafe_allow_html=True)
                    st.markdown(f"**Comunidad:** {tramo_dict['comunidad']} &nbsp;·&nbsp; **Longitud:** {tramo_dict['longitud_km']} km", unsafe_allow_html=True)
                    st.markdown(f"**Estado actual:** {badge_estado(tramo_dict['estado_actual'])}", unsafe_allow_html=True)
                    st.markdown(f"**Superficie:** {tramo_dict['tipo_superficie']} &nbsp;·&nbsp; **Ancho:** {tramo_dict['ancho_m']} m", unsafe_allow_html=True)
                    st.caption(f"Fichas de daño registradas: {len(danos_tramo_dl)}")

                st.markdown("##### Vista previa — 1-B Ficha del Itinerario del Camino Vecinal")
                st.dataframe(df_ficha_itinerario(tramo_dict).head(8), use_container_width=True, hide_index=True)
                st.caption(f"Mostrando 8 de {len(generar_grid_progresiva(tramo_dict['longitud_km']))} tramos de progresiva. El archivo descargable incluye la tabla completa.")

                st.markdown("##### Vista previa — 1-D Ficha Técnica de Daños")
                df_d_prev = df_ficha_danos(danos_tramo_dl)
                if df_d_prev.empty:
                    st.caption("Sin fichas de daño registradas aún para este tramo.")
                else:
                    st.dataframe(df_d_prev, use_container_width=True, hide_index=True)

                st.markdown("---")
                st.markdown("##### Descargar ficha técnica completa")
                d1, d2 = st.columns(2)
                with d1:
                    excel_bytes = exportar_ficha_excel(tramo_dict, danos_tramo_dl, fotos_tramo)
                    st.download_button(
                        "📊 Descargar ficha completa (Excel)",
                        data=excel_bytes,
                        file_name=f"Ficha_Tecnica_{tramo_dict['codigo']}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"xlsx_{tramo_id}",
                        use_container_width=True,
                    )
                with d2:
                    pdf_ficha_bytes = exportar_ficha_pdf(tramo_dict, danos_tramo_dl, fotos_tramo)
                    download_button_bytes(
                        pdf_ficha_bytes,
                        f"Ficha_Tecnica_{tramo_dict['codigo']}.pdf",
                        "📄 Descargar ficha completa (PDF)",
                        "application/pdf",
                        f"pdf_{tramo_id}",
                    )
                st.caption("Ambos archivos incluyen: datos generales, ficha 1-B (itinerario), ficha 1-D (daños registrados) "
                           "y la sección de Evidencia Fotográfica del Mantenimiento con las fotos cargadas.")


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 4 — FICHA TÉCNICA DE DAÑOS
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📋 FICHA TÉCNICA DE DAÑOS":
    module_header("📋", "Ficha Técnica de Daños en Camino Vecinal", "Formato basado en las fichas 1-B / 1-D de Provías Descentralizado (Itinerario y Daños).")

    tramos = df_tramos()
    if tramos.empty:
        st.info("Primero registra al menos un tramo en **🛣️ Registro de Tramos**.")
    else:
        opciones = {f"{row.codigo} — {row.nombre} ({row.longitud_km} km)": (row.id, row.longitud_km)
                    for row in tramos.itertuples()}
        _claves_op = list(opciones.keys())
        _tac_ficha = st.session_state.get("tramo_activo_codigo")
        _idx_default = next((i for i, k in enumerate(_claves_op) if k.startswith(f"{_tac_ficha} —")), 0) if _tac_ficha else 0
        seleccion = st.selectbox("Tramo a inspeccionar", _claves_op, index=_idx_default)
        if _tac_ficha:
            st.markdown('<div class="auto-fill-note">📍 Se preseleccionó el tramo activo definido en el panel lateral.</div>', unsafe_allow_html=True)
        tramo_id, longitud_tramo = opciones[seleccion]

        conn = get_conn()
        row = conn.execute("SELECT * FROM tramos WHERE id=?", (tramo_id,)).fetchone()
        cols_t = [d[0] for d in conn.execute("SELECT * FROM tramos LIMIT 1").description]
        conn.close()
        tramo_dict_ficha = dict(zip(cols_t, row))
        fotos_tramo_ficha = obtener_fotos_tramo(tramo_id)

        def _fmt(v, sufijo=""):
            if v is None or v == "":
                return "No registrado"
            return f"{v}{sufijo}"

        st.markdown("---")
        st.markdown(f"### 📍 Tramo {tramo_dict_ficha['codigo']} — {tramo_dict_ficha['nombre']}")

        col_dt, col_rl = st.columns(2)
        with col_dt:
            st.markdown("##### 🔧 Datos técnicos")
            st.markdown(f"""
            <div class="kpi-card" style="padding:0.9rem 1.1rem;">
            <table style="width:100%; font-size:0.88rem;">
            <tr><td style="color:#5a6472;">Longitud del tramo</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['longitud_km'], ' km')}</td></tr>
            <tr><td style="color:#5a6472;">Ancho de plataforma</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['ancho_m'], ' m')}</td></tr>
            <tr><td style="color:#5a6472;">Altitud</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['altitud_msnm'], ' m s.n.m.')}</td></tr>
            <tr><td style="color:#5a6472;">Tipo de superficie</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['tipo_superficie'])}</td></tr>
            <tr><td style="color:#5a6472;">Estado del drenaje</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['estado_drenaje'])}</td></tr>
            <tr><td style="color:#5a6472;">Señalización</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['senalizacion'])}</td></tr>
            <tr><td style="color:#5a6472;">N° de puentes</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['n_puentes'])}</td></tr>
            <tr><td style="color:#5a6472;">N° de badenes</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['n_badenes'])}</td></tr>
            </table>
            </div>
            """, unsafe_allow_html=True)

        with col_rl:
            st.markdown("##### 🚧 Recursos y logística")
            costo_km_rut = (tramo_dict_ficha["costo_rutinario"] / tramo_dict_ficha["longitud_km"]
                             if tramo_dict_ficha["costo_rutinario"] and tramo_dict_ficha["longitud_km"] else None)
            st.markdown(f"""
            <div class="kpi-card" style="padding:0.9rem 1.1rem;">
            <table style="width:100%; font-size:0.88rem;">
            <tr><td style="color:#5a6472;">Accesibilidad</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['accesibilidad'])}</td></tr>
            <tr><td style="color:#5a6472;">Cantera / fuente mat.</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['cantera_fuente'])}</td></tr>
            <tr><td style="color:#5a6472;">Fuente de lastre</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['fuente_lastre'])}</td></tr>
            <tr><td style="color:#5a6472;">Última intervención</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['ultima_intervencion'])}</td></tr>
            <tr><td style="color:#5a6472;">Costo est. rutinario</td><td style="text-align:right; font-weight:600;">{('S/. ' + format(tramo_dict_ficha['costo_rutinario'], ',.0f')) if tramo_dict_ficha['costo_rutinario'] else 'No registrado'}</td></tr>
            <tr><td style="color:#5a6472;">Costo est. periódico</td><td style="text-align:right; font-weight:600;">{('S/. ' + format(tramo_dict_ficha['costo_periodico'], ',.0f')) if tramo_dict_ficha['costo_periodico'] else 'No registrado'}</td></tr>
            <tr><td style="color:#5a6472;">Costo/km rutinario</td><td style="text-align:right; font-weight:600;">{('S/. ' + format(costo_km_rut, ',.0f')) if costo_km_rut else 'No registrado'}</td></tr>
            <tr><td style="color:#5a6472;">Coordenadas GPS</td><td style="text-align:right; font-weight:600;">{(str(tramo_dict_ficha['gps_lat']) + ', ' + str(tramo_dict_ficha['gps_lon'])) if tramo_dict_ficha['gps_lat'] else 'No registrado'}</td></tr>
            </table>
            </div>
            """, unsafe_allow_html=True)

        st.markdown("##### 📉 Curva de deterioro del IEC (Índice de Estado del Camino)")
        st.plotly_chart(curva_iec_deterioro(tramo_dict_ficha), use_container_width=True)
        st.caption(
            "Curva referencial según el Manual de Conservación Vial MCV-2014 (MTC). La estrella marca la "
            "posición estimada del tramo según su estado actual y el tiempo transcurrido desde su última intervención."
        )

        st.markdown("---")
        st.markdown("### 📑 Expediente Técnico Resumido del Tramo")
        st.caption("Fichas 1-B (Itinerario) y 1-D (Daños) según formato Provías Descentralizado, con evidencia fotográfica y descarga del expediente completo.")

        danos_tramo_exp = df_danos()
        danos_tramo_exp = (danos_tramo_exp[danos_tramo_exp["tramo_id"] == tramo_id]
                            if not danos_tramo_exp.empty else danos_tramo_exp)

        exp_col1, exp_col2 = st.columns([1, 1.4])
        with exp_col1:
            st.markdown("##### 📷 Evidencia fotográfica")
            if fotos_tramo_ficha:
                _mini_cols = st.columns(min(len(fotos_tramo_ficha), 3))
                for i, foto in enumerate(fotos_tramo_ficha[:3]):
                    with _mini_cols[i]:
                        st.image(blob_to_img(foto["imagen"]), use_container_width=True)
                if len(fotos_tramo_ficha) > 3:
                    st.caption(f"+ {len(fotos_tramo_ficha) - 3} fotografía(s) más en el expediente descargable.")
            else:
                st.info("Este tramo aún no tiene fotografías cargadas.")
        with exp_col2:
            st.markdown("##### 🗂️ Datos de gestión técnica")
            estado_mant_actual = (danos_tramo_exp.sort_values("fecha_inspeccion", ascending=False).iloc[0]["tipo_mantenimiento"]
                                   if not danos_tramo_exp.empty else "Sin evaluar")
            presupuesto_ref = tramo_dict_ficha.get("costo_rutinario") or 0
            st.markdown(f"""
            <div class="kpi-card" style="padding:0.9rem 1.1rem;">
            <table style="width:100%; font-size:0.88rem;">
            <tr><td style="color:#5a6472;">Estado actual</td><td style="text-align:right;">{badge_estado(tramo_dict_ficha['estado_actual'])}</td></tr>
            <tr><td style="color:#5a6472;">Mantenimiento sugerido (última ficha)</td><td style="text-align:right; font-weight:600;">{estado_mant_actual}</td></tr>
            <tr><td style="color:#5a6472;">Presupuesto referencial rutinario</td><td style="text-align:right; font-weight:600;">{('S/. ' + format(presupuesto_ref, ',.0f')) if presupuesto_ref else 'No registrado'}</td></tr>
            <tr><td style="color:#5a6472;">Ubicación</td><td style="text-align:right; font-weight:600;">{tramo_dict_ficha['comunidad']}, {tramo_dict_ficha['distrito']}</td></tr>
            <tr><td style="color:#5a6472;">Responsable</td><td style="text-align:right; font-weight:600;">{_fmt(tramo_dict_ficha['responsable'])}</td></tr>
            <tr><td style="color:#5a6472;">Fichas de daño registradas</td><td style="text-align:right; font-weight:600;">{len(danos_tramo_exp)}</td></tr>
            </table>
            </div>
            """, unsafe_allow_html=True)
            if tramo_dict_ficha.get("observaciones"):
                st.caption(f"📝 {tramo_dict_ficha['observaciones']}")

        tab_1b, tab_1d = st.tabs(["1-B: Ficha del Itinerario del Camino Vecinal", "1-D: Ficha Técnica de Daños en Camino Vecinal"])
        with tab_1b:
            st.dataframe(df_ficha_itinerario(tramo_dict_ficha), use_container_width=True, hide_index=True, height=320)
        with tab_1d:
            df_1d_view = df_ficha_danos(danos_tramo_exp)
            if df_1d_view.empty:
                st.caption("Sin fichas de daño registradas para este tramo todavía.")
            else:
                st.dataframe(df_1d_view, use_container_width=True, hide_index=True)

        dl1, dl2 = st.columns(2)
        with dl1:
            excel_bytes_exp = exportar_ficha_excel(tramo_dict_ficha, danos_tramo_exp, fotos_tramo_ficha)
            st.download_button(
                "📊 Descargar expediente completo (Excel)", data=excel_bytes_exp,
                file_name=f"Expediente_Tecnico_{tramo_dict_ficha['codigo']}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"exp_xlsx_{tramo_id}", use_container_width=True,
            )
        with dl2:
            pdf_bytes_exp = exportar_ficha_pdf(tramo_dict_ficha, danos_tramo_exp, fotos_tramo_ficha)
            download_button_bytes(
                pdf_bytes_exp, f"Expediente_Tecnico_{tramo_dict_ficha['codigo']}.pdf",
                "📄 Descargar expediente completo (PDF)", "application/pdf", f"exp_pdf_{tramo_id}",
            )
        st.caption("El expediente incluye: datos generales, ficha 1-B (itinerario), ficha 1-D (daños registrados) "
                   "y la sección de Evidencia Fotográfica del Mantenimiento con las fotos cargadas.")

        st.markdown("---")

        with st.form("form_ficha_dano", clear_on_submit=True):
            st.markdown("#### Progresiva y ubicación del daño")
            c1, c2, c3 = st.columns(3)
            prog_inicial = c1.number_input("Progresiva inicial (km)", min_value=0.0, step=0.05, format="%.3f")
            prog_final = c2.number_input("Progresiva final (km)", min_value=0.0, step=0.05, format="%.3f")
            longitud_afectada = c3.number_input("Longitud afectada (km)", min_value=0.0, step=0.01, format="%.3f")

            st.markdown("#### Características del daño")
            c4, c5, c6 = st.columns(3)
            tipo_dano = c4.selectbox("Tipo de daño", TIPOS_DANO)
            clase_densidad = c5.selectbox("Clase de densidad", CLASES_DENSIDAD,
                                           help="Solo aplica obligatoriamente si el tipo de daño es 'Baches o Huecos'.")
            fecha_inspeccion = c6.date_input("Fecha de inspección", value=date.today())

            observaciones = st.text_area("Observaciones técnicas", placeholder="Describe lo observado en campo...")
            foto_dano = st.file_uploader("Fotografía del daño", type=["jpg", "jpeg", "png"])

            calcular = st.form_submit_button("⚙️ Calcular clasificación automática y guardar", use_container_width=True)

        if calcular:
            falla = clasificar_falla(tipo_dano)
            peso = PESO_DANO.get(tipo_dano, 1)
            gravedad = calcular_gravedad(longitud_afectada, longitud_tramo, peso)
            estado_tramo_calc = calcular_estado(gravedad)
            transitabilidad = calcular_transitabilidad(estado_tramo_calc)
            prioridad = calcular_prioridad(estado_tramo_calc, transitabilidad)
            tiempo_rep = calcular_tiempo_reparacion(tipo_dano, max(longitud_afectada, 0.05))
            ict = calcular_ict(gravedad, clase_densidad, falla)
            pct_deterioro = round(100 - ict, 1)
            necesidad = calcular_necesidad_intervencion(prioridad)
            tipo_mantenimiento = determinar_tipo_mantenimiento(tipo_dano, gravedad, ict, transitabilidad)

            data = {
                "tramo_id": tramo_id, "progresiva_inicial": prog_inicial, "progresiva_final": prog_final,
                "longitud_afectada": longitud_afectada, "tipo_dano": tipo_dano, "tipo_falla": falla,
                "nivel_gravedad": gravedad, "clase_densidad": clase_densidad, "estado_tramo": estado_tramo_calc,
                "transitabilidad": transitabilidad, "necesidad_intervencion": necesidad,
                "tiempo_estimado_dias": tiempo_rep, "fecha_inspeccion": fecha_inspeccion.strftime("%Y-%m-%d"),
                "observaciones": observaciones, "ict": ict, "pct_deterioro": pct_deterioro,
                "prioridad": prioridad, "tipo_mantenimiento": tipo_mantenimiento,
                "foto_dano": img_to_blob(foto_dano, carpeta="fotos-danos"),
            }
            try:
                dano_id = insertar_dano(data)
            except Exception as e:
                st.error(f"⚠️ No se pudo guardar la ficha de daño en la base de datos: {e}")
                st.stop()
            st.session_state["ultimo_dano_id"] = dano_id
            st.session_state["ultimo_dano_tramo"] = tramo_id
            st.session_state["ultimo_dano_mant"] = tipo_mantenimiento
            st.session_state["ultimo_dano_dias"] = tiempo_rep

            st.success("✅ Ficha de daño registrada. Resultado de la clasificación automática:")
            r1, r2, r3, r4 = st.columns(4)
            r1.metric("Tipo de falla", falla)
            r2.metric("Nivel de gravedad", gravedad)
            r3.markdown(f"**Estado del tramo**<br>{badge_estado(estado_tramo_calc)}", unsafe_allow_html=True)
            r4.metric("Transitabilidad", transitabilidad)

            r5, r6, r7, r8 = st.columns(4)
            r5.metric("Prioridad de intervención", prioridad)
            r6.metric("Tiempo estimado de reparación", f"{tiempo_rep} días")
            r7.metric("ICT (Índice de Condición)", f"{ict} / 100")
            r8.metric("% de deterioro", f"{pct_deterioro}%")

            color_mant = MANTENIMIENTOS[tipo_mantenimiento]["color"]
            st.markdown(
                f'<div style="background:{color_mant}15; border-left:5px solid {color_mant}; '
                f'border-radius:8px; padding:0.8rem 1rem; margin-top:0.5rem;">'
                f'<b style="color:{color_mant};">Tipo de mantenimiento recomendado: {tipo_mantenimiento}</b><br>'
                f'<span style="font-size:0.88rem;">{MANTENIMIENTOS[tipo_mantenimiento]["descripcion"]}</span></div>',
                unsafe_allow_html=True,
            )
            st.info(f"**Necesidad de intervención:** {necesidad}")
            st.caption("👉 Ve al módulo **🗓️ Programación de Mantenimiento** para programar esta intervención con expediente, fecha y costo.")

        st.markdown("---")
        st.markdown("#### Histórico de fichas de daños registradas")
        danos = df_danos()
        if danos.empty:
            st.caption("Aún no hay fichas registradas.")
        else:
            cols_show = ["codigo_tramo", "nombre_tramo", "progresiva_inicial", "progresiva_final",
                         "longitud_afectada", "tipo_dano", "tipo_falla", "nivel_gravedad",
                         "estado_tramo", "transitabilidad", "prioridad", "tipo_mantenimiento", "ict", "pct_deterioro",
                         "tiempo_estimado_dias", "fecha_inspeccion"]
            st.dataframe(danos[cols_show].rename(columns={
                "codigo_tramo": "Código", "nombre_tramo": "Tramo", "progresiva_inicial": "Prog. Inicial",
                "progresiva_final": "Prog. Final", "longitud_afectada": "Long. Afectada (km)",
                "tipo_dano": "Tipo de Daño", "tipo_falla": "Falla", "nivel_gravedad": "Gravedad",
                "estado_tramo": "Estado", "transitabilidad": "Transitabilidad", "prioridad": "Prioridad",
                "tipo_mantenimiento": "Mantenimiento Sugerido",
                "ict": "ICT", "pct_deterioro": "% Deterioro", "tiempo_estimado_dias": "Días Rep.",
                "fecha_inspeccion": "Fecha",
            }), use_container_width=True, hide_index=True)

            st.markdown("##### 🏆 Ranking de tramos críticos (menor ICT)")
            ranking = danos.groupby(["codigo_tramo", "nombre_tramo"])["ict"].mean().reset_index()
            ranking = ranking.sort_values("ict").rename(columns={"ict": "ICT promedio"})
            ranking.index = range(1, len(ranking) + 1)
            st.dataframe(ranking, use_container_width=True)


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 5 — PROGRAMACIÓN DE MANTENIMIENTO (5 tipos de mantenimiento)
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "🗓️ PROGRAMACIÓN DE MANTENIMIENTO":
    module_header(
        "🗓️", "Programación de Mantenimiento Vial",
        "Visión general de toda la conservación vial programada, con sugerencia automática de fechas, "
        "alertas por temporada de lluvias y criticidad."
    )

    if "intervencion_guardada_info" in st.session_state:
        st.success("✅ " + st.session_state.pop("intervencion_guardada_info") + " Ya está reflejada en el cronograma de abajo.")

    tramos = df_tramos()
    danos = df_danos()
    intervenciones = df_intervenciones()

    if not intervenciones.empty:
        if "fecha_fin" not in intervenciones.columns or intervenciones["fecha_fin"].isna().all():
            intervenciones["fecha_fin"] = (pd.to_datetime(intervenciones["fecha_programada"]) + pd.to_timedelta(
                intervenciones["duracion_dias"], unit="D")).astype(str)
        else:
            fallback = (pd.to_datetime(intervenciones["fecha_programada"]) +
                        pd.to_timedelta(intervenciones["duracion_dias"], unit="D")).astype(str)
            intervenciones["fecha_fin"] = intervenciones["fecha_fin"].fillna(fallback)

    # ════════════════════════════════════════════════════════════════════════
    # PARTE A — VISTA GENERAL DE PLANIFICACIÓN
    # ════════════════════════════════════════════════════════════════════════
    st.markdown("### 📊 A. Vista general de planificación")

    if intervenciones.empty:
        st.info("Aún no hay intervenciones programadas. Registra la primera en la sección **B** más abajo.")
    else:
        st.markdown("##### Totales por tipo de mantenimiento")
        cols_kpi = st.columns(len(MANTENIMIENTOS))
        for col, (nombre, info) in zip(cols_kpi, MANTENIMIENTOS.items()):
            subset = intervenciones[intervenciones["tipo_mantenimiento"] == nombre]
            n = len(subset)
            costo_total = subset["costo_estimado"].sum() if n else 0
            with col:
                st.markdown(f"""
                <div class="kpi-card" style="border-top-color:{info['color']};">
                <div class="kpi-value" style="color:{info['color']}; font-size:1.5rem;">{n}</div>
                <div class="kpi-label" style="font-size:0.72rem;">{nombre}</div>
                <div style="font-size:0.75rem; color:#5a6472; margin-top:0.2rem;">S/. {costo_total:,.0f}</div>
                </div>
                """, unsafe_allow_html=True)

        k1, k2, k3, k4 = st.columns(4)
        k1.metric("Total intervenciones", len(intervenciones))
        k2.metric("Costo total programado", f"S/. {intervenciones['costo_estimado'].sum():,.0f}")
        k3.metric("Pendientes / Programadas", int(intervenciones["estado"].isin(["Pendiente", "Programada"]).sum()))
        k4.metric("Prioridad Alta / Urgente", int(intervenciones["prioridad"].isin(["Alta", "Urgente"]).sum()))

        st.markdown("""
        <div style="background:linear-gradient(135deg,#0b2545,#13315c); border-radius:8px 8px 0 0;
                    padding:0.9rem 1.2rem; margin-top:0.5rem;">
            <span style="color:#fff; font-size:1.05rem; font-weight:700; font-family:Georgia,serif;">
            📅 Gantt Chart — Programación de Mantenimiento Vial</span>
        </div>
        """, unsafe_allow_html=True)
        st.caption("Incluye los 5 tipos de mantenimiento, todos los tramos y todas las intervenciones registradas hasta ahora — ordenadas cronológicamente.")

        gantt_df = intervenciones.copy()
        gantt_df["fecha_fin_dt"] = pd.to_datetime(gantt_df["fecha_fin"])
        gantt_df["fecha_inicio_dt"] = pd.to_datetime(gantt_df["fecha_programada"])
        gantt_df = gantt_df.sort_values("fecha_inicio_dt")
        # Etiqueta de fila: una barra por intervención (tramo + tipo), igual que filas de tarea en un Gantt clásico
        gantt_df["tarea"] = gantt_df["codigo_tramo"] + " · " + gantt_df["tipo_mantenimiento"]
        orden_tareas = gantt_df.sort_values("fecha_inicio_dt")["tarea"].tolist()

        fig_g = px.timeline(
            gantt_df, x_start="fecha_inicio_dt", x_end="fecha_fin_dt",
            y="tarea", color="tipo_mantenimiento",
            color_discrete_map={k: v["color"] for k, v in MANTENIMIENTOS.items()},
            category_orders={"tarea": orden_tareas},
            hover_data=["nombre_tramo", "componente", "expediente", "estado", "prioridad", "costo_estimado", "duracion_dias"],
        )
        fig_g.update_traces(marker_line_width=0, width=0.55)
        fig_g.update_yaxes(autorange="reversed", title="", showgrid=True, gridcolor="#e8ebf0")
        fig_g.update_xaxes(title="", showgrid=True, gridcolor="#e8ebf0", side="top",
                            tickfont=dict(color="#0b2545", size=11))
        fig_g.update_layout(
            height=max(280, 42 * len(orden_tareas) + 80),
            legend_title="Tipo de mantenimiento",
            plot_bgcolor="#ffffff", paper_bgcolor="#ffffff",
            margin=dict(t=10, b=10, l=10, r=10),
            font=dict(family="Calibri, Arial", size=12, color="#23272f"),
            bargap=0.35,
        )
        st.plotly_chart(fig_g, use_container_width=True)
        st.caption("💡 Pasa el cursor sobre cada barra para ver tramo, componente, expediente, estado, prioridad y costo de la intervención.")

        st.markdown("##### 📋 Cronograma de Mantenimiento — Detalle por tipo")
        st.caption("Los 5 tipos de mantenimiento definidos en el proyecto, con su programación actual (o pendiente de programar).")

        for nombre_tipo, info_tipo in MANTENIMIENTOS.items():
            subset_tipo = intervenciones[intervenciones["tipo_mantenimiento"] == nombre_tipo].sort_values("fecha_programada")
            st.markdown(
                f'<div style="border-left:5px solid {info_tipo["color"]}; padding:0.3rem 0.7rem; '
                f'margin-top:0.6rem; font-weight:700; color:{info_tipo["color"]};">{nombre_tipo} '
                f'<span style="font-size:0.78rem; color:#777; font-weight:400;">({len(subset_tipo)} intervención(es))</span></div>',
                unsafe_allow_html=True,
            )
            if subset_tipo.empty:
                st.caption("　Sin intervenciones programadas para este tipo de mantenimiento.")
            else:
                tabla_tipo = subset_tipo[["nombre_tramo", "fecha_programada", "fecha_fin", "estado", "responsable", "observaciones"]].rename(
                    columns={"nombre_tramo": "Nombre / Tramo", "fecha_programada": "Fecha tentativa inicio",
                             "fecha_fin": "Fecha tentativa fin", "estado": "Estado", "responsable": "Responsable",
                             "observaciones": "Observaciones"}
                )
                st.dataframe(tabla_tipo, use_container_width=True, hide_index=True)


        cc1, cc2 = st.columns(2)

        with cc1:
            st.markdown("###### Costo programado por tipo de mantenimiento")
            costo_mant = intervenciones.groupby("tipo_mantenimiento")["costo_estimado"].sum().reset_index()
            fig_pie = px.pie(costo_mant, names="tipo_mantenimiento", values="costo_estimado", hole=0.4,
                              color="tipo_mantenimiento",
                              color_discrete_map={k: v["color"] for k, v in MANTENIMIENTOS.items()})
            fig_pie.update_layout(height=300, margin=dict(t=10, b=10, l=10, r=10))
            st.plotly_chart(fig_pie, use_container_width=True)
        with cc2:
            st.markdown("###### Intervenciones por componente vial")
            comp_count = intervenciones["componente"].fillna("Sin especificar").value_counts().reset_index()
            comp_count.columns = ["Componente", "N°"]
            fig_comp = px.bar(comp_count, x="Componente", y="N°", color="Componente", text_auto=True)
            fig_comp.update_layout(showlegend=False, height=300, margin=dict(t=10, b=10, l=10, r=10))
            st.plotly_chart(fig_comp, use_container_width=True)

        st.markdown("##### 📍 Resumen por tramo: historial y programación futura")
        tramo_filtro = st.selectbox("Selecciona un tramo", sorted(intervenciones["codigo_tramo"].dropna().unique()), key="resumen_tramo_sel")
        interv_t = intervenciones[intervenciones["codigo_tramo"] == tramo_filtro].sort_values("fecha_programada")
        tramo_info = tramos[tramos["codigo"] == tramo_filtro].iloc[0]

        st.markdown(
            f"**{tramo_filtro} — {tramo_info['nombre']}** &nbsp;·&nbsp; "
            f'<span style="font-size:0.85rem; color:#5a6472;">{tramo_info["comunidad"]}, {tramo_info["distrito"]} · '
            f'{tramo_info["longitud_km"]} km · Registrado: {tramo_info["fecha_registro"]} · '
            f'Última intervención: {tramo_info["ultima_intervencion"] or "No registrada"}</span> · '
            f"Estado: {badge_estado(tramo_info['estado_actual'])}",
            unsafe_allow_html=True,
        )

        resumen_cols = ["tipo_mantenimiento", "componente", "actividades", "fecha_programada", "fecha_fin",
                         "duracion_dias", "costo_estimado", "estado", "prioridad", "responsable"]
        st.dataframe(
            interv_t[resumen_cols].rename(columns={
                "tipo_mantenimiento": "Tipo de mantenimiento", "componente": "Componente",
                "actividades": "Actividades", "fecha_programada": "Inicio tentativo", "fecha_fin": "Fin tentativo",
                "duracion_dias": "Días", "costo_estimado": "Costo (S/.)", "estado": "Estado",
                "prioridad": "Prioridad", "responsable": "Responsable",
            }), use_container_width=True, hide_index=True,
        )

        with st.expander("📋 Ver tabla completa de todas las intervenciones (todos los tramos)"):
            st.dataframe(
                intervenciones[["codigo_tramo", "nombre_tramo", "comunidad_tramo", "longitud_tramo_km",
                                 "tipo_mantenimiento", "componente", "fecha_registro_tramo",
                                 "ultima_intervencion_tramo", "fecha_programada", "fecha_fin", "duracion_dias",
                                 "estado", "prioridad", "costo_estimado", "responsable"]]
                .rename(columns={
                    "codigo_tramo": "Código", "nombre_tramo": "Tramo", "comunidad_tramo": "Ubicación",
                    "longitud_tramo_km": "Long. (km)", "tipo_mantenimiento": "Tipo Mant.",
                    "componente": "Componente", "fecha_registro_tramo": "F. Registro",
                    "ultima_intervencion_tramo": "Última Interv.", "fecha_programada": "Inicio",
                    "fecha_fin": "Fin", "duracion_dias": "Días", "prioridad": "Prioridad",
                    "costo_estimado": "Costo (S/.)",
                }), use_container_width=True, hide_index=True,
            )
            u1, u2, u3 = st.columns(3)
            opciones_upd = {f"#{r.id} · {r.codigo_tramo} · {r.tipo_mantenimiento} · {r.fecha_programada}": r.id
                            for r in intervenciones.itertuples()}
            sel_upd = u1.selectbox("Intervención", list(opciones_upd.keys()), key="upd_interv")
            nuevo_estado_upd = u2.selectbox("Nuevo estado", ESTADOS_INTERVENCION, key="upd_estado")
            if u3.button("🔄 Actualizar estado", use_container_width=True):
                actualizar_estado_intervencion(opciones_upd[sel_upd], nuevo_estado_upd)
                st.success("Estado actualizado. Vuelve a cargar la tabla para ver el cambio.")

        st.markdown("##### 🗂️ Fichas de planificación detalladas del tramo seleccionado")
        for _, interv in interv_t.iterrows():
            color = MANTENIMIENTOS.get(interv["tipo_mantenimiento"], {}).get("color", "#0b2545")
            col_card, col_img = st.columns([3, 1])
            with col_card:
                st.markdown(f"""
                <div style="border-left:5px solid {color}; background:#fff; border-radius:6px;
                            padding:0.9rem 1.1rem; margin-bottom:0.8rem; box-shadow:0 1px 3px rgba(0,0,0,0.06);">
                <b style="color:{color};">{interv['tipo_mantenimiento']}</b> · Componente: <b>{interv['componente'] or '—'}</b>
                &nbsp;·&nbsp; Prioridad: <b>{interv['prioridad'] or '—'}</b><br>
                <span style="font-size:0.85rem;">
                📅 {interv['fecha_programada']} → {str(interv['fecha_fin'])[:10]} ({interv['duracion_dias']} días) &nbsp;|&nbsp;
                💰 S/. {interv['costo_estimado']:,.0f} &nbsp;|&nbsp; Estado: <b>{interv['estado']}</b><br>
                📄 Expediente: {interv['expediente'] or '—'} &nbsp;|&nbsp; Responsable: {interv['responsable'] or '—'}<br>
                🛠️ Actividades: {interv['actividades'] or '—'}<br>
                📝 {interv['observaciones'] or 'Sin observaciones adicionales.'}
                </span>
                </div>
                """, unsafe_allow_html=True)
            with col_img:
                if interv.get("foto_evidencia"):
                    st.image(blob_to_img(interv["foto_evidencia"]), caption="Evidencia", use_container_width=True)

    st.markdown("---")
    st.info("👉 Para registrar una nueva intervención de mantenimiento, ve al módulo **🛣️ Registro de Tramos** → pestaña **🛠️ Programar mantenimiento del tramo**.")


# ══════════════════════════════════════════════════════════════════════════════
# MÓDULO 6 — REPORTES
# ══════════════════════════════════════════════════════════════════════════════
elif menu == "📊 REPORTES":
    module_header("📊", "Reportes", "Inventario, daños, costos e intervenciones, con gráficos interactivos y exportación a Excel/PDF.")
    tramos = df_tramos()
    danos = df_danos()

    _tac_r = st.session_state.get("tramo_activo_codigo")
    if _tac_r:
        tramos = tramos[tramos["codigo"] == _tac_r]
        if not danos.empty:
            danos = danos[danos["codigo_tramo"] == _tac_r]
        st.info(f"📍 Reportes filtrados al tramo **{_tac_r}**. Cambia el filtro lateral para ver la red completa.")

    if tramos.empty:
        st.info("No hay datos suficientes para generar reportes todavía.")
    else:
        tab1, tab2, tab3, tab4, tab5 = st.tabs(
            ["📦 Inventario y Estado", "⚠️ Daños", "💰 Costos e Intervenciones",
             "🗂️ Histórico de Expedientes", "⬇️ Exportar"])

        # ---- TAB 1: Inventario y estado de la red ----
        with tab1:
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("##### Estado de la red vial (N° de tramos)")
                fig = px.bar(tramos["estado_actual"].value_counts().reset_index(),
                             x="estado_actual", y="count",
                             color="estado_actual",
                             color_discrete_map={"Bueno": "#43a047", "Regular": "#fdd835",
                                                  "Malo": "#fb8c00", "Muy Malo": "#e53935"},
                             text_auto=True)
                fig.update_layout(showlegend=False, xaxis_title="", yaxis_title="N° tramos", height=340)
                st.plotly_chart(fig, use_container_width=True)
            with c2:
                st.markdown("##### Kilómetros afectados por estado")
                km_estado = tramos.groupby("estado_actual")["longitud_km"].sum().reset_index()
                fig2 = px.pie(km_estado, names="estado_actual", values="longitud_km", hole=0.4,
                              color="estado_actual",
                              color_discrete_map={"Bueno": "#43a047", "Regular": "#fdd835",
                                                   "Malo": "#fb8c00", "Muy Malo": "#e53935"})
                fig2.update_layout(height=340)
                st.plotly_chart(fig2, use_container_width=True)

            st.markdown("##### Inventario completo de tramos")
            st.dataframe(tramos, use_container_width=True, hide_index=True)

        # ---- TAB 2: Daños ----
        with tab2:
            if danos.empty:
                st.info("No hay fichas de daños registradas aún.")
            else:
                c1, c2 = st.columns(2)
                with c1:
                    st.markdown("##### Daños por tipo")
                    fig3 = px.bar(danos["tipo_dano"].value_counts().reset_index(),
                                  x="tipo_dano", y="count", color="count",
                                  color_continuous_scale="Reds", text_auto=True)
                    fig3.update_layout(xaxis_title="", yaxis_title="N° de casos", height=360)
                    st.plotly_chart(fig3, use_container_width=True)
                with c2:
                    st.markdown("##### Daños por nivel de gravedad")
                    fig4 = px.pie(danos, names="nivel_gravedad", hole=0.4,
                                  color="nivel_gravedad",
                                  color_discrete_map={"Sin Deterioro": "#43a047", "Leve": "#9ccc65",
                                                       "Moderada": "#fb8c00", "Severa": "#e53935"})
                    fig4.update_layout(height=360)
                    st.plotly_chart(fig4, use_container_width=True)

                st.markdown("##### Histograma de % de deterioro (ICT invertido)")
                fig5 = px.histogram(danos, x="pct_deterioro", nbins=10, color_discrete_sequence=["#c62828"])
                fig5.update_layout(xaxis_title="% de deterioro", yaxis_title="Frecuencia", height=320)
                st.plotly_chart(fig5, use_container_width=True)

                st.markdown("##### Kilómetros afectados por tipo de daño")
                km_dano = danos.groupby("tipo_dano")["longitud_afectada"].sum().reset_index()
                fig6 = px.bar(km_dano.sort_values("longitud_afectada", ascending=False),
                              x="tipo_dano", y="longitud_afectada", color="longitud_afectada",
                              color_continuous_scale="Oranges", text_auto=".2f")
                fig6.update_layout(xaxis_title="", yaxis_title="km afectados", height=340)
                st.plotly_chart(fig6, use_container_width=True)

                st.markdown("##### Tramos críticos (Prioridad Urgente / Alta)")
                criticos = danos[danos["prioridad"].isin(["Urgente", "Alta"])][
                    ["codigo_tramo", "nombre_tramo", "tipo_dano", "estado_tramo",
                     "transitabilidad", "prioridad", "ict", "tiempo_estimado_dias"]]
                st.dataframe(criticos, use_container_width=True, hide_index=True)

        # ---- TAB 3: Costos / necesidades de mantenimiento ----
        with tab3:
            if danos.empty:
                st.info("No hay fichas de daños registradas aún.")
            else:
                COSTO_REFERENCIAL_KM = {  # S/. por km — referencial MCV-2014 / benchmarks de campo
                    "Deformación": 9500, "Erosión": 11000, "Baches o Huecos": 14500,
                    "Encalaminado": 8000, "Lodazal": 17500, "Cruce de Agua": 22000,
                    "Hundimiento": 26000, "Pérdida de Plataforma": 31000,
                    "Deslizamiento": 35000, "Falla de Drenaje": 19500,
                }
                danos["costo_estimado"] = danos.apply(
                    lambda r: round(r["longitud_afectada"] * COSTO_REFERENCIAL_KM.get(r["tipo_dano"], 12000), 0),
                    axis=1,
                )
                c1, c2, c3 = st.columns(3)
                c1.metric("Costo estimado total de intervención", f"S/. {danos['costo_estimado'].sum():,.0f}")
                c2.metric("Km totales afectados", f"{danos['longitud_afectada'].sum():.2f} km")
                c3.metric("Días totales estimados de reparación", f"{danos['tiempo_estimado_dias'].sum():.0f} días")

                st.markdown("##### Costo estimado por tipo de daño")
                costo_tipo = danos.groupby("tipo_dano")["costo_estimado"].sum().reset_index()
                fig7 = px.bar(costo_tipo.sort_values("costo_estimado", ascending=False),
                              x="tipo_dano", y="costo_estimado", color="costo_estimado",
                              color_continuous_scale="Blues", text_auto=".2s")
                fig7.update_layout(xaxis_title="", yaxis_title="S/.", height=340)
                st.plotly_chart(fig7, use_container_width=True)

                st.markdown("##### Necesidades de mantenimiento por tramo")
                necesidades = danos.groupby(["codigo_tramo", "nombre_tramo", "necesidad_intervencion"]).size().reset_index(name="N° fichas")
                st.dataframe(necesidades, use_container_width=True, hide_index=True)

        # ---- TAB 4: Histórico de expedientes ----
        with tab4:
            st.markdown("##### Expedientes técnicos ejecutados — Municipalidad Distrital de Ocongate")
            df_exp = pd.DataFrame(EXPEDIENTES_HISTORICOS)
            st.dataframe(
                df_exp.style.format({"Costo total ejecutado (S/.)": "S/. {:,.2f}"}),
                use_container_width=True, hide_index=True,
            )
            c1, c2 = st.columns(2)
            with c1:
                st.markdown("##### Inversión ejecutada por año")
                inv_anio = df_exp.groupby("Año")["Costo total ejecutado (S/.)"].sum().reset_index()
                fig_exp1 = px.bar(inv_anio, x="Año", y="Costo total ejecutado (S/.)",
                                  color="Costo total ejecutado (S/.)", color_continuous_scale="Blues",
                                  text_auto=".2s")
                fig_exp1.update_layout(height=340)
                st.plotly_chart(fig_exp1, use_container_width=True)
            with c2:
                st.markdown("##### Costo ejecutado por expediente")
                fig_exp2 = px.bar(df_exp.sort_values("Costo total ejecutado (S/.)", ascending=True),
                                  x="Costo total ejecutado (S/.)", y="Expediente", orientation="h",
                                  color="Costo total ejecutado (S/.)", color_continuous_scale="Tealgrn",
                                  text_auto=".2s")
                fig_exp2.update_layout(height=340)
                st.plotly_chart(fig_exp2, use_container_width=True)

            total_historico = df_exp["Costo total ejecutado (S/.)"].sum()
            st.metric("Inversión histórica total ejecutada (2021-2024)", f"S/. {total_historico:,.2f}")
            st.caption(
                "Esta información histórica sirve como línea base de costos reales de intervención "
                "correctiva, en contraste con el ahorro proyectado mediante el enfoque preventivo (PMA)."
            )

        # ---- TAB 5: Exportar ----
        with tab5:
            st.markdown("#### Exportar reportes")
            dfs_excel = {"Inventario_Tramos": tramos}
            if not danos.empty:
                dfs_excel["Fichas_Danos"] = danos.drop(columns=["foto_dano"], errors="ignore")
                dfs_excel["Ranking_Criticos"] = (
                    danos.groupby(["codigo_tramo", "nombre_tramo"])["ict"].mean()
                    .reset_index().sort_values("ict").rename(columns={"ict": "ICT_promedio"})
                )

            c1, c2 = st.columns(2)
            with c1:
                st.download_button(
                    "⬇️ Descargar reporte completo (Excel)",
                    data=exportar_excel(dfs_excel),
                    file_name=f"VialNet_Reporte_{date.today().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with c2:
                pdf_bytes = exportar_pdf_resumen(
                    "VialNet — Reporte de Inventario de Trochas Carrozables, Ocongate",
                    tramos[["codigo", "nombre", "comunidad", "longitud_km", "estado_actual"]],
                    texto_intro=f"Total de tramos: {len(tramos)} | Km inventariados: {tramos['longitud_km'].sum():.2f} km",
                )
                download_button_bytes(
                    pdf_bytes,
                    f"VialNet_Reporte_{date.today().strftime('%Y%m%d')}.pdf",
                    "⬇️ Descargar reporte resumen (PDF)",
                    "application/pdf",
                    "pdf_main",
                )


st.markdown("---")
st.caption("Plataforma desarrollada como demostración técnica — Proyecto Preprofesional UTEC 2026-I · "
           "Frank Puma Mamani (202220055) · Municipalidad Distrital de Ocongate.")
